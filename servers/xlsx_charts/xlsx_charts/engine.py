"""XLSX Charts engine — pure openpyxl logic, zero MCP imports."""

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string

from shared.file_utils import resolve_path
from shared.live_edit import notify_reload
from shared.platform_utils import open_file
from shared.progress import fail, ok
from shared.version_control import snapshot

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_CELL_RE = re.compile(r"^[A-Z]{1,3}\d+$")

CHART_TYPES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
    "scatter": ScatterChart,
}


def _validate_cell(address: str) -> bool:
    return bool(_CELL_RE.match(address.upper()))


def _open_wb(path: Path, progress: list[dict[str, Any]]) -> tuple[Any, dict[str, Any] | None]:
    """Load workbook; return (wb, None) on success or (None, error_dict) on fail."""
    if not path.exists():
        progress.append(fail("File not found", str(path)))
        return None, {
            "success": False,
            "error": f"File not found: {path}",
            "hint": "Check that file_path is absolute and the file exists.",
            "progress": progress,
            "token_estimate": 20,
        }
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        progress.append(fail(f"Wrong file type: {path.suffix}"))
        return None, {
            "success": False,
            "error": f"Expected .xlsx file, got {path.suffix}",
            "hint": "Use the correct server for this file type.",
            "progress": progress,
            "token_estimate": 20,
        }
    wb = openpyxl.load_workbook(str(path))
    return wb, None


def _check_sheet(
    wb: Any, sheet_name: str, progress: list[dict[str, Any]], backup: str | None
) -> tuple[Any, dict[str, Any] | None]:
    """Return (ws, None) or (None, error_dict) if sheet missing."""
    if sheet_name not in wb.sheetnames:
        progress.append(fail(f"Sheet '{sheet_name}' not found"))
        return None, {
            "success": False,
            "error": f"Sheet '{sheet_name}' not found",
            "hint": "Use list_sheets to get available sheet names.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }
    return wb[sheet_name], None


def _parse_range_ref(ws: Any, data_range: str) -> Reference:
    """Parse 'Sheet1!A1:D10' or 'A1:D10' into an openpyxl Reference."""
    if "!" in data_range:
        _, cell_range = data_range.split("!", 1)
    else:
        cell_range = data_range
    return Reference(ws, range_string=f"'{ws.title}'!{cell_range}")


def _parse_cell_range(range_str: str) -> tuple[int, int, int, int]:
    """Parse 'A1:D10' into (min_row, min_col, max_row, max_col) 1-indexed."""
    top, bot = range_str.upper().split(":")
    top_col_str = re.sub(r"\d", "", top)
    top_row = int(re.sub(r"[A-Z]", "", top))
    bot_col_str = re.sub(r"\d", "", bot)
    bot_row = int(re.sub(r"[A-Z]", "", bot))
    return (
        top_row,
        column_index_from_string(top_col_str),
        bot_row,
        column_index_from_string(bot_col_str),
    )


# ---------------------------------------------------------------------------
# Tool implementations
# ---------------------------------------------------------------------------


def add_chart(
    file_path: str,
    sheet_name: str,
    chart_type: str,
    data_range: str,
    title: str,
    anchor_cell: str,
    width: float = 15.0,
    height: float = 10.0,
    open_after: bool = False,
) -> dict[str, Any]:
    """Create a chart from a data range and place it on the sheet."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        if chart_type not in CHART_TYPES:
            progress.append(fail(f"Unsupported chart type: {chart_type}"))
            return {
                "success": False,
                "error": f"Unsupported chart type: {chart_type}",
                "hint": f"Allowed types: {', '.join(sorted(CHART_TYPES))}",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        progress.append(ok(f"Opened {path.name}"))

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        chart_cls = CHART_TYPES[chart_type]
        chart = chart_cls()
        chart.title = title
        chart.width = width
        chart.height = height

        data_ref = _parse_range_ref(ws, data_range)
        chart.add_data(data_ref, titles_from_data=True)

        anchor = anchor_cell.upper()
        ws.add_chart(chart, anchor)

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(notify_reload(str(path), "xlsx"))
        progress.append(ok(f"Created {chart_type} chart '{title}'", f"anchored at {anchor}"))

        result: dict[str, Any] = {
            "success": True,
            "op": "add_chart",
            "sheet": sheet_name,
            "chart_type": chart_type,
            "title": title,
            "anchor_cell": anchor,
            "data_range": data_range,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def delete_chart(
    file_path: str,
    sheet_name: str,
    chart_index: int,
    open_after: bool = False,
) -> dict[str, Any]:
    """Remove a chart from a sheet by its zero-based index."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        progress.append(ok(f"Opened {path.name}"))

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        charts = ws._charts  # type: ignore[attr-defined]
        if chart_index < 0 or chart_index >= len(charts):
            progress.append(
                fail(f"Chart index {chart_index} out of range", f"Sheet has {len(charts)} chart(s)")
            )
            return {
                "success": False,
                "error": f"chart_index {chart_index} out of range (0-{len(charts) - 1})",
                "hint": "Use add_chart to see chart indices.",
                "backup": backup,
                "progress": progress,
                "token_estimate": 15,
            }

        deleted_title = getattr(charts[chart_index], "title", f"chart_{chart_index}")
        del charts[chart_index]

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"Deleted chart '{deleted_title}'", sheet_name))
        result: dict[str, Any] = {
            "success": True,
            "op": "delete_chart",
            "sheet": sheet_name,
            "chart_index": chart_index,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def update_chart(
    file_path: str,
    sheet_name: str,
    chart_index: int,
    title: str = "",
    data_range: str = "",
    open_after: bool = False,
) -> dict[str, Any]:
    """Update chart title and/or data range. Uses delete-then-add for data changes."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        progress.append(ok(f"Opened {path.name}"))

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        charts = ws._charts  # type: ignore[attr-defined]
        if chart_index < 0 or chart_index >= len(charts):
            progress.append(
                fail(f"Chart index {chart_index} out of range", f"Sheet has {len(charts)} chart(s)")
            )
            return {
                "success": False,
                "error": f"chart_index {chart_index} out of range (0-{len(charts) - 1})",
                "hint": "Use add_chart to see chart indices.",
                "backup": backup,
                "progress": progress,
                "token_estimate": 15,
            }

        chart = charts[chart_index]

        if title:
            chart.title = title
            progress.append(ok(f"Updated chart title to '{title}'"))

        if data_range:
            data_ref = _parse_range_ref(ws, data_range)
            chart.series.clear()
            chart.add_data(data_ref, titles_from_data=True)
            progress.append(ok("Updated chart data range", data_range))

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        result: dict[str, Any] = {
            "success": True,
            "op": "update_chart",
            "sheet": sheet_name,
            "chart_index": chart_index,
            "title": title,
            "data_range": data_range,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def add_pivot_table(
    file_path: str,
    sheet_name: str,
    source_range: str,
    dest_cell: str,
    rows: str,
    cols: str,
    values: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Create a summary pivot-style table from source data.

    rows/cols/values are column header names from source_range.
    Writes aggregated (summed) values to dest_cell location.
    """
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        progress.append(ok(f"Opened {path.name}"))

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        # Parse source range
        if "!" in source_range:
            src_sheet_name, cell_range = source_range.split("!", 1)
            src_sheet_name = src_sheet_name.strip("'")
            if src_sheet_name not in wb.sheetnames:
                progress.append(fail(f"Source sheet '{src_sheet_name}' not found"))
                return {
                    "success": False,
                    "error": f"Sheet '{src_sheet_name}' not found",
                    "hint": "Check the source_range sheet name.",
                    "backup": backup,
                    "progress": progress,
                    "token_estimate": 15,
                }
            src_ws = wb[src_sheet_name]
        else:
            cell_range = source_range
            src_ws = ws

        min_row, min_col, max_row, max_col = _parse_cell_range(cell_range)

        # Read header row
        headers: list[str] = []
        for col in range(min_col, max_col + 1):
            cell_val = src_ws.cell(row=min_row, column=col).value
            headers.append(str(cell_val) if cell_val is not None else "")

        # Validate column names
        for col_name in [rows, cols, values]:
            if col_name not in headers:
                progress.append(fail(f"Column '{col_name}' not in source headers"))
                return {
                    "success": False,
                    "error": f"Column '{col_name}' not found in source range headers",
                    "hint": f"Available headers: {', '.join(headers)}",
                    "backup": backup,
                    "progress": progress,
                    "token_estimate": 15,
                }

        rows_idx = headers.index(rows) + min_col
        cols_idx = headers.index(cols) + min_col
        vals_idx = headers.index(values) + min_col

        # Read data rows and aggregate
        aggregated: dict[str, dict[str, float]] = {}
        col_keys: list[str] = []

        for r in range(min_row + 1, max_row + 1):
            row_key = str(src_ws.cell(row=r, column=rows_idx).value or "")
            col_key = str(src_ws.cell(row=r, column=cols_idx).value or "")
            val = src_ws.cell(row=r, column=vals_idx).value
            num_val = float(val) if isinstance(val, (int, float)) else 0.0

            if row_key not in aggregated:
                aggregated[row_key] = {}
            aggregated[row_key][col_key] = aggregated[row_key].get(col_key, 0.0) + num_val

            if col_key not in col_keys:
                col_keys.append(col_key)

        # Write pivot table to dest_cell
        dest_addr = dest_cell.upper()
        dest_row = int(re.sub(r"[A-Z]", "", dest_addr))
        dest_col = column_index_from_string(re.sub(r"\d", "", dest_addr))

        # Header row: blank, then column keys
        ws.cell(row=dest_row, column=dest_col).value = rows
        for ci, ck in enumerate(col_keys):
            ws.cell(row=dest_row, column=dest_col + 1 + ci).value = ck

        # Data rows
        for ri, (rk, col_data) in enumerate(aggregated.items()):
            ws.cell(row=dest_row + 1 + ri, column=dest_col).value = rk
            for ci, ck in enumerate(col_keys):
                ws.cell(row=dest_row + 1 + ri, column=dest_col + 1 + ci).value = col_data.get(
                    ck, 0.0
                )

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(
            ok(
                f"Created pivot summary at {dest_cell}",
                f"{len(aggregated)} row groups × {len(col_keys)} column groups",
            )
        )
        result: dict[str, Any] = {
            "success": True,
            "op": "add_pivot_table",
            "sheet": sheet_name,
            "dest_cell": dest_cell,
            "row_groups": len(aggregated),
            "col_groups": len(col_keys),
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def set_cell_style(
    file_path: str,
    sheet_name: str,
    cell_address: str,
    font_name: str = "",
    font_size: float = 0,
    bold: bool = False,
    fill_color: str = "",
    number_format: str = "",
    open_after: bool = False,
) -> dict[str, Any]:
    """Apply font, fill color, and number format to a cell."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        addr = cell_address.upper()
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid cell address: {cell_address}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {cell_address}",
                "hint": "Use Excel notation like B5 or C12.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        progress.append(ok(f"Opened {path.name}"))

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        cell = ws[addr]

        # Build font
        existing_font = cell.font
        new_font_kwargs: dict[str, Any] = {
            "name": font_name if font_name else (existing_font.name if existing_font else None),
            "size": font_size if font_size > 0 else (existing_font.size if existing_font else None),
            "bold": bold if bold else (existing_font.bold if existing_font else None),
        }
        cell.font = Font(**{k: v for k, v in new_font_kwargs.items() if v is not None})

        # Apply fill color
        if fill_color:
            clean_color = fill_color.lstrip("#")
            if len(clean_color) == 6:
                cell.fill = PatternFill(
                    start_color=clean_color,
                    end_color=clean_color,
                    fill_type="solid",
                )

        # Apply number format
        if number_format:
            cell.number_format = number_format

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"Styled cell {addr}", sheet_name))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_cell_style",
            "sheet": sheet_name,
            "cell": addr,
            "font_name": font_name,
            "font_size": font_size,
            "bold": bold,
            "fill_color": fill_color,
            "number_format": number_format,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }
