"""XLSX Formulas engine — pure openpyxl logic, zero MCP imports."""

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from shared.file_utils import resolve_path
from shared.live_edit import notify_reload
from shared.progress import fail, ok, warn
from shared.version_control import snapshot

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_CELL_RE = re.compile(r"^[A-Z]{1,3}\d+$")
_RANGE_RE = re.compile(r"^[A-Z]{1,3}\d+:[A-Z]{1,3}\d+$")

COLOR_MAP: dict[str, str] = {
    "green": "00FF00",
    "red": "FF0000",
    "yellow": "FFFF00",
    "blue": "0000FF",
}

VALID_RULES = {"greater_than", "less_than", "between", "equal_to"}
VALID_COLORS = set(COLOR_MAP.keys())


def _validate_cell(address: str) -> bool:
    return bool(_CELL_RE.match(address.upper()))


def _validate_range(address: str) -> bool:
    return bool(_RANGE_RE.match(address.upper()))


def _open_wb(path: Path, progress: list[dict[str, Any]]) -> tuple[Any, dict[str, Any] | None]:
    """Load workbook; return (wb, None) on success or (None, error_dict) on fail."""
    if not path.exists():
        progress.append(fail("File not found", str(path)))
        return None, {
            "success": False,
            "error": f"File not found: {path}",
            "hint": "Check that file_path is absolute and the file exists.",
            "progress": progress,
            "token_estimate": 20,
        }
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        progress.append(fail(f"Wrong file type: {path.suffix}"))
        return None, {
            "success": False,
            "error": f"Expected .xlsx file, got {path.suffix}",
            "hint": "Use the correct server for this file type.",
            "progress": progress,
            "token_estimate": 20,
        }
    wb = openpyxl.load_workbook(str(path))
    return wb, None


def _check_sheet(wb: Any, sheet_name: str, progress: list[dict[str, Any]], backup: str | None) -> tuple[Any, dict[str, Any] | None]:
    """Return (ws, None) or (None, error_dict) if sheet missing."""
    if sheet_name not in wb.sheetnames:
        progress.append(fail(f"Sheet '{sheet_name}' not found"))
        return None, {
            "success": False,
            "error": f"Sheet '{sheet_name}' not found",
            "hint": "Use list_sheets to get available sheet names.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }
    return wb[sheet_name], None


# ---------------------------------------------------------------------------
# Tool implementations
# ---------------------------------------------------------------------------


def set_formula(
    file_path: str,
    sheet_name: str,
    cell_address: str,
    formula: str,
) -> dict[str, Any]:
    """Write a formula string to a cell. Formula must start with '='."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    addr = cell_address.upper()
    try:
        if not formula.startswith("="):
            progress.append(fail("Formula must start with '='", formula))
            return {
                "success": False,
                "error": "Formula must start with '='",
                "hint": "Prefix the formula with '=', e.g. '=SUM(B2:B10)'.",
                "progress": progress,
                "token_estimate": 15,
            }
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid cell address: {cell_address}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {cell_address}",
                "hint": "Use Excel notation like B5 or C12.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        ws[addr] = formula
        wb.save(str(path))
        wb.close()

        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_formula",
            "sheet": sheet_name,
            "cell": addr,
            "formula": formula,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def set_named_range(
    file_path: str,
    sheet_name: str,
    range_name: str,
    range_address: str,
) -> dict[str, Any]:
    """Define a named range in the workbook."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        if not range_name.replace("_", "").isalnum():
            progress.append(fail(f"Invalid range name: {range_name}"))
            return {
                "success": False,
                "error": f"Invalid range name: {range_name}",
                "hint": "Range names must be alphanumeric (underscores allowed).",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        if sheet_name not in wb.sheetnames:
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": "Use list_sheets to get available sheet names.",
                "backup": backup,
                "progress": progress,
                "token_estimate": 15,
            }

        attr_text = f"'{sheet_name}'!{range_address}"
        defn = DefinedName(range_name, attr_text=attr_text)
        wb.defined_names[range_name] = defn
        wb.save(str(path))
        wb.close()

        progress.append(ok(f"Defined named range '{range_name}'", attr_text))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_named_range",
            "range_name": range_name,
            "sheet": sheet_name,
            "range_address": range_address,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def set_conditional_format(
    file_path: str,
    sheet_name: str,
    range_address: str,
    rule: str,
    value: float,
    color: str,
    value2: float = 0.0,
) -> dict[str, Any]:
    """Apply a color-based conditional formatting rule to a range."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        if rule not in VALID_RULES:
            progress.append(fail(f"Unknown rule: {rule}"))
            return {
                "success": False,
                "error": f"Unknown rule: {rule}",
                "hint": f"Allowed rules: {', '.join(sorted(VALID_RULES))}",
                "progress": progress,
                "token_estimate": 15,
            }
        if color not in VALID_COLORS:
            progress.append(fail(f"Unknown color: {color}"))
            return {
                "success": False,
                "error": f"Unknown color: {color}",
                "hint": f"Allowed colors: {', '.join(sorted(VALID_COLORS))}",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        hex_color = COLOR_MAP[color]
        fill = PatternFill(
            start_color=hex_color,
            end_color=hex_color,
            fill_type="solid",
        )

        op_map = {
            "greater_than": "greaterThan",
            "less_than": "lessThan",
            "between": "between",
            "equal_to": "equal",
        }
        operator = op_map[rule]

        if rule == "between":
            rule_obj = CellIsRule(
                operator=operator,
                formula=[str(value), str(value2)],
                fill=fill,
            )
        else:
            rule_obj = CellIsRule(
                operator=operator,
                formula=[str(value)],
                fill=fill,
            )

        ws.conditional_formatting.add(range_address, rule_obj)
        wb.save(str(path))
        wb.close()

        progress.append(
            ok(
                f"Applied '{rule}' conditional format",
                f"{range_address} → {color}",
            )
        )
        result: dict[str, Any] = {
            "success": True,
            "op": "set_conditional_format",
            "sheet": sheet_name,
            "range": range_address,
            "rule": rule,
            "value": value,
            "color": color,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def set_data_validation(
    file_path: str,
    sheet_name: str,
    range_address: str,
    validation_type: str,
    formula1: str = "",
    formula2: str = "",
) -> dict[str, Any]:
    """Add data validation (list, decimal, or whole) to a cell range."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    valid_types = {"list", "decimal", "whole"}
    try:
        if validation_type not in valid_types:
            progress.append(fail(f"Unknown validation type: {validation_type}"))
            return {
                "success": False,
                "error": f"Unknown validation_type: {validation_type}",
                "hint": f"Allowed types: {', '.join(sorted(valid_types))}",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        dv = DataValidation(
            type=validation_type,
            formula1=formula1 or None,
            formula2=formula2 or None,
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.sqref = range_address

        wb.save(str(path))
        wb.close()

        progress.append(
            ok(
                f"Set '{validation_type}' data validation",
                f"{range_address}",
            )
        )
        result: dict[str, Any] = {
            "success": True,
            "op": "set_data_validation",
            "sheet": sheet_name,
            "range": range_address,
            "validation_type": validation_type,
            "formula1": formula1,
            "formula2": formula2,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def freeze_panes(
    file_path: str,
    sheet_name: str,
    cell_address: str,
) -> dict[str, Any]:
    """Freeze rows/columns at cell_address. Empty string to unfreeze."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        addr = cell_address.upper() if cell_address else ""
        if addr and not _validate_cell(addr):
            progress.append(fail(f"Invalid cell address: {cell_address}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {cell_address}",
                "hint": "Use Excel notation like B2. Empty string to unfreeze.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        ws.freeze_panes = addr if addr else None
        wb.save(str(path))
        wb.close()

        action = f"frozen at {addr}" if addr else "unfrozen"
        progress.append(ok(f"Panes {action}", sheet_name))
        result: dict[str, Any] = {
            "success": True,
            "op": "freeze_panes",
            "sheet": sheet_name,
            "cell_address": addr,
            "frozen": bool(addr),
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }


def set_autofilter(
    file_path: str,
    sheet_name: str,
    range_address: str,
) -> dict[str, Any]:
    """Enable AutoFilter on the specified range."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    try:
        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        ws.auto_filter.ref = range_address
        wb.save(str(path))
        wb.close()

        progress.append(ok(f"AutoFilter set on {range_address}", sheet_name))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_autofilter",
            "sheet": sheet_name,
            "range": range_address,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": str(e),
            "hint": "Use restore_version to undo if a snapshot was taken.",
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }
