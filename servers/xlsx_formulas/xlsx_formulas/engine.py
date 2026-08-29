"""XLSX Formulas engine — pure openpyxl logic, zero MCP imports."""

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from shared.file_utils import drop_snapshot_if_unwritten, hint_for_error, resolve_path, scrub_repr, sheet_names_hint
from shared.live_edit import notify_reload
from shared.platform_utils import get_max_cells, open_file
from shared.progress import fail, info, ok, warn
from shared.version_control import snapshot

# openpyxl writes the formula string into the sheet and stops there -- it has no
# calculation engine, and the cached-result slot Excel and LibreOffice fill in
# when they save stays empty. So a caller that writes =SUM(B2:B10) and then
# reads the cell back gets `value: null`, which reads as a failed write. It was
# not: the formula is in the file and any spreadsheet application will compute
# it on open. Every tool here that writes a formula says so, once, in the same
# response as the success.
UNCALCULATED_NOTE = (
    "Stored, not computed. This server writes the formula text; it has no calculation engine, so the cell "
    "has no cached result until Excel or LibreOffice opens the file and saves it. read_cell() reports such "
    "a cell as type 'formula_uncalculated'."
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_CELL_RE = re.compile(r"^[A-Z]{1,3}\d+$")
_RANGE_RE = re.compile(r"^[A-Z]{1,3}\d+:[A-Z]{1,3}\d+$")

COLOR_MAP: dict[str, str] = {
    "green": "00FF00",
    "red": "FF0000",
    "yellow": "FFFF00",
    "blue": "0000FF",
}

VALID_RULES = {"greater_than", "less_than", "between", "equal_to"}
VALID_COLORS = set(COLOR_MAP.keys())

# set_conditional_format's docstring reads "rule: gt/lt/between/eq" and has done
# since it was written: the 80-character budget cannot hold "greater_than
# less_than between equal_to" *and* the four colour names, so the rules were
# abbreviated. The schema carries no enum and no description, so that line is
# the only vocabulary a caller ever sees -- and a coverage sweep read it, sent
# rule="gt", and was told "Unknown rule: gt". Rather than drop the colours from
# the documentation to make room, the short forms the docs promise are accepted.
# auto_sum already normalises its own function_name the same way.
RULE_ALIASES = {"gt": "greater_than", "lt": "less_than", "eq": "equal_to"}


def _validate_cell(address: str) -> bool:
    return bool(_CELL_RE.match(address.upper()))


def _validate_range(address: str) -> bool:
    return bool(_RANGE_RE.match(address.upper()))


def _open_wb(path: Path, progress: list[dict[str, Any]]) -> tuple[Any, dict[str, Any] | None]:
    """Load workbook; return (wb, None) on success or (None, error_dict) on fail."""
    if not path.exists():
        progress.append(fail("File not found", str(path)))
        return None, {
            "success": False,
            "error": f"File not found: {path}",
            "hint": "Check that file_path is absolute and the file exists.",
            "progress": progress,
            "token_estimate": 20,
        }
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        progress.append(fail(f"Wrong file type: {path.suffix}"))
        return None, {
            "success": False,
            "error": f"Expected .xlsx file, got {path.suffix}",
            "hint": "Use the correct server for this file type.",
            "progress": progress,
            "token_estimate": 20,
        }
    wb = openpyxl.load_workbook(str(path))
    return wb, None


def _check_sheet(
    wb: Any, sheet_name: str, progress: list[dict[str, Any]], backup: str | None
) -> tuple[Any, dict[str, Any] | None]:
    """Return (ws, None) or (None, error_dict) if sheet missing."""
    if sheet_name not in wb.sheetnames:
        available_sheets = list(wb.sheetnames)
        progress.append(fail(f"Sheet '{sheet_name}' not found"))
        return None, {
            "success": False,
            "error": f"Sheet '{sheet_name}' not found",
            "hint": sheet_names_hint(available_sheets),
            "backup": backup,
            "progress": progress,
            "token_estimate": 15,
        }
    return wb[sheet_name], None


# ---------------------------------------------------------------------------
# Tool implementations
# ---------------------------------------------------------------------------


def set_formula(
    file_path: str,
    sheet_name: str,
    cell_address: str,
    formula: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Write a formula string to a cell. Formula must start with '='."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    addr = cell_address.upper()
    path: Path | None = None
    try:
        if not formula.startswith("="):
            progress.append(fail("Formula must start with '='", formula))
            return {
                "success": False,
                "error": "Formula must start with '='",
                "hint": "Prefix the formula with '=', e.g. '=SUM(B2:B10)'.",
                "progress": progress,
                "token_estimate": 15,
            }
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid cell address: {cell_address}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {cell_address}",
                "hint": "Use Excel notation like B5 or C12.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        ws[addr] = formula
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_formula",
            "sheet": sheet_name,
            "cell": addr,
            "formula": formula,
            "calculated": False,
            "note": UNCALCULATED_NOTE,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def set_named_range(
    file_path: str,
    sheet_name: str,
    range_name: str,
    range_address: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Define a named range in the workbook."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        if not range_name.replace("_", "").isalnum():
            progress.append(fail(f"Invalid range name: {range_name}"))
            return {
                "success": False,
                "error": f"Invalid range name: {range_name}",
                "hint": "Range names must be alphanumeric (underscores allowed).",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        # The sheet name was prefixed unconditionally, so a caller who passed
        # range_address="Sheet1!$M$2:$M$10" -- the form every Excel reference
        # they have ever seen -- got "'Sheet1'!Sheet1!$M$2:$M$10" written into
        # the workbook and success:true back. Excel rejects that definedName
        # and nothing in the response said anything was wrong.
        address = range_address.strip()
        if "!" in address:
            qualifier, _, bare = address.rpartition("!")
            named_sheet = qualifier.strip().strip("'")
            if named_sheet and named_sheet != sheet_name:
                progress.append(fail(f"range_address names sheet '{named_sheet}'"))
                return {
                    "success": False,
                    "error": f"range_address names sheet '{named_sheet}' but sheet_name is '{sheet_name}'",
                    "hint": (
                        f"Pass the bare range, e.g. range_address='{bare}', and name the sheet "
                        "in sheet_name — or set sheet_name to the sheet the address refers to."
                    ),
                    "backup": drop_snapshot_if_unwritten(backup, path, progress),
                    "progress": progress,
                    "token_estimate": 30,
                }
            address = bare
            progress.append(info("Sheet qualifier removed", f"{range_address} → {address}"))

        attr_text = f"'{sheet_name}'!{address}"
        defn = DefinedName(range_name, attr_text=attr_text)
        wb.defined_names[range_name] = defn
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"Defined named range '{range_name}'", attr_text))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_named_range",
            "range_name": range_name,
            "sheet": sheet_name,
            "range_address": address,
            "reference": attr_text,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def set_conditional_format(
    file_path: str,
    sheet_name: str,
    range_address: str,
    rule: str,
    value: float,
    color: str,
    value2: float = 0.0,
    open_after: bool = False,
) -> dict[str, Any]:
    """Apply a color-based conditional formatting rule to a range."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        rule = RULE_ALIASES.get(rule, rule)
        if rule not in VALID_RULES:
            progress.append(fail(f"Unknown rule: {rule}"))
            return {
                "success": False,
                "error": f"Unknown rule: {rule}",
                "hint": "Allowed rules: between, equal_to (eq), greater_than (gt), less_than (lt)",
                "progress": progress,
                "token_estimate": 15,
            }
        if color not in VALID_COLORS:
            progress.append(fail(f"Unknown color: {color}"))
            return {
                "success": False,
                "error": f"Unknown color: {color}",
                "hint": f"Allowed colors: {', '.join(sorted(VALID_COLORS))}",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        hex_color = COLOR_MAP[color]
        fill = PatternFill(
            start_color=hex_color,
            end_color=hex_color,
            fill_type="solid",
        )

        op_map = {
            "greater_than": "greaterThan",
            "less_than": "lessThan",
            "between": "between",
            "equal_to": "equal",
        }
        operator = op_map[rule]

        if rule == "between":
            rule_obj = CellIsRule(
                operator=operator,
                formula=[str(value), str(value2)],
                fill=fill,
            )
        else:
            rule_obj = CellIsRule(
                operator=operator,
                formula=[str(value)],
                fill=fill,
            )

        ws.conditional_formatting.add(range_address, rule_obj)
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(
            ok(
                f"Applied '{rule}' conditional format",
                f"{range_address} → {color}",
            )
        )
        result: dict[str, Any] = {
            "success": True,
            "op": "set_conditional_format",
            "sheet": sheet_name,
            "range": range_address,
            "rule": rule,
            "value": value,
            "color": color,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def set_data_validation(
    file_path: str,
    sheet_name: str,
    range_address: str,
    validation_type: str,
    formula1: str = "",
    formula2: str = "",
    open_after: bool = False,
) -> dict[str, Any]:
    """Add data validation (list, decimal, or whole) to a cell range."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    valid_types = {"list", "decimal", "whole"}
    path: Path | None = None
    try:
        if validation_type not in valid_types:
            progress.append(fail(f"Unknown validation type: {validation_type}"))
            return {
                "success": False,
                "error": f"Unknown validation_type: {validation_type}",
                "hint": f"Allowed types: {', '.join(sorted(valid_types))}",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        dv = DataValidation(
            type=validation_type,  # type: ignore[reportArgumentType]
            formula1=formula1 or None,
            formula2=formula2 or None,
            allow_blank=True,
            # openpyxl defaults showErrorMessage to False, which makes Excel
            # draw the dropdown and then accept anything typed over it. A
            # validation that validates nothing is not what this tool is for,
            # and the response gave no sign: the rule was in the XML and
            # success was true.
            showErrorMessage=True,
            errorTitle="Invalid entry",
            error="That value is not allowed in this cell.",
        )
        ws.add_data_validation(dv)
        dv.sqref = range_address

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(
            ok(
                f"Set '{validation_type}' data validation",
                f"{range_address}",
            )
        )
        result: dict[str, Any] = {
            "success": True,
            "op": "set_data_validation",
            "sheet": sheet_name,
            "range": range_address,
            "validation_type": validation_type,
            "formula1": formula1,
            "formula2": formula2,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def freeze_panes(
    file_path: str,
    sheet_name: str,
    cell_address: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Freeze rows/columns at cell_address. Empty string to unfreeze."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        addr = cell_address.upper() if cell_address else ""
        if addr and not _validate_cell(addr):
            progress.append(fail(f"Invalid cell address: {cell_address}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {cell_address}",
                "hint": "Use Excel notation like B2. Empty string to unfreeze.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        ws.freeze_panes = addr if addr else None
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        action = f"frozen at {addr}" if addr else "unfrozen"
        progress.append(ok(f"Panes {action}", sheet_name))
        result: dict[str, Any] = {
            "success": True,
            "op": "freeze_panes",
            "sheet": sheet_name,
            "cell_address": addr,
            "frozen": bool(addr),
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def set_autofilter(
    file_path: str,
    sheet_name: str,
    range_address: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Enable AutoFilter on the specified range."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        ws.auto_filter.ref = range_address
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"AutoFilter set on {range_address}", sheet_name))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_autofilter",
            "sheet": sheet_name,
            "range": range_address,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def fill_formula_down(
    file_path: str,
    sheet_name: str,
    formula: str,
    start_cell: str,
    end_row: int,
    open_after: bool = False,
) -> dict[str, Any]:
    """Fill formula down from start_cell to end_row, adjusting row references."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        if not formula.startswith("="):
            progress.append(fail("Formula must start with '='", formula))
            return {
                "success": False,
                "error": "Formula must start with '='",
                "hint": "Prefix the formula with '=', e.g. '=B2*C2'.",
                "progress": progress,
                "token_estimate": 15,
            }

        addr = start_cell.upper()
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid start_cell: {start_cell}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {start_cell}",
                "hint": "Use Excel notation like D2 or B5.",
                "progress": progress,
                "token_estimate": 15,
            }

        # Parse column letters and start row from start_cell
        col_match = re.match(r"^([A-Z]+)(\d+)$", addr)
        if not col_match:
            progress.append(fail(f"Cannot parse start_cell: {start_cell}"))
            return {
                "success": False,
                "error": f"Cannot parse cell address: {start_cell}",
                "hint": "Use Excel notation like D2.",
                "progress": progress,
                "token_estimate": 15,
            }
        col_letters = col_match.group(1)
        start_row = int(col_match.group(2))

        if end_row < start_row:
            progress.append(fail(f"end_row {end_row} is before start_row {start_row}"))
            return {
                "success": False,
                "error": f"end_row ({end_row}) must be >= start_row ({start_row})",
                "hint": "end_row is the last row number to fill to.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        # Write formula to each row, adjusting row numbers in cell references
        cells_filled = 0
        for target_row in range(start_row, end_row + 1):
            adjusted = re.sub(
                r"([A-Z]+)(" + str(start_row) + r")\b",
                lambda m, tr=target_row: m.group(1) + str(tr),
                formula,
            )
            ws[f"{col_letters}{target_row}"] = adjusted
            cells_filled += 1

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(
            ok(
                f"Filled formula down {col_letters}{start_row}:{col_letters}{end_row}",
                f"{cells_filled} cells",
            )
        )
        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "fill_formula_down",
            "sheet": sheet_name,
            "column": col_letters,
            "start_row": start_row,
            "end_row": end_row,
            "cells_filled": cells_filled,
            "calculated": False,
            "note": UNCALCULATED_NOTE,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def auto_sum(
    file_path: str,
    sheet_name: str,
    data_range: str,
    sum_cell: str,
    function_name: str = "SUM",
    open_after: bool = False,
) -> dict[str, Any]:
    """Write a SUM/AVERAGE/COUNT/MAX/MIN formula for data_range into sum_cell."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    valid_functions = {"SUM", "AVERAGE", "COUNT", "MAX", "MIN"}
    path: Path | None = None
    try:
        fn = function_name.upper()
        if fn not in valid_functions:
            progress.append(fail(f"Unknown function: {function_name}"))
            return {
                "success": False,
                "error": f"Unknown function_name: {function_name}",
                "hint": f"Allowed functions: {', '.join(sorted(valid_functions))}",
                "progress": progress,
                "token_estimate": 15,
            }

        if not _validate_range(data_range.upper()):
            progress.append(fail(f"Invalid data_range: {data_range}"))
            return {
                "success": False,
                "error": f"Invalid range address: {data_range}",
                "hint": "Use Excel range notation like B2:B20.",
                "progress": progress,
                "token_estimate": 15,
            }

        addr = sum_cell.upper()
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid sum_cell: {sum_cell}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {sum_cell}",
                "hint": "Use Excel notation like B21.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws, err = _check_sheet(wb, sheet_name, progress, backup)
        if err:
            return err

        formula = f"={fn}({data_range})"
        ws[addr] = formula
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"Set {addr} = {formula}", sheet_name))
        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "auto_sum",
            "sheet": sheet_name,
            "sum_cell": addr,
            "formula": formula,
            "data_range": data_range,
            "function_name": fn,
            "calculated": False,
            "note": UNCALCULATED_NOTE,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def convert_to_values(
    file_path: str,
    sheet_name: str,
    range_address: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Replace formula cells with their calculated values in a range."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        if not _validate_range(range_address.upper()):
            progress.append(fail(f"Invalid range: {range_address}"))
            return {
                "success": False,
                "error": f"Invalid range address: {range_address}",
                "hint": "Use Excel range notation like B2:D10.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        wb, err = _open_wb(path, progress)
        if err:
            return err

        # Load a second copy with data_only=True to read evaluated values.
        # read_only matters here: the writable copy above already materialises
        # every cell, and a second full load doubles that. On a 16,834 x 16
        # sheet the pair ran long enough for the transport to drop mid-call --
        # the same failure read_cell had, for the same reason. This copy is only
        # ever read from (cell_v.value below), so it can stream.
        wb_values = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            wb.close()
            wb_values.close()
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "progress": progress,
                "token_estimate": 15,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        ws = wb[sheet_name]
        ws_values = wb_values[sheet_name]

        # data_only=True only returns a value that Excel/LibreOffice already
        # cached in the file — openpyxl never evaluates formulas itself. A
        # formula written by set_formula/auto_sum/fill_formula_down and never
        # opened by a real spreadsheet app has no cached value at all, so
        # cell_v.value is None. Blindly assigning that would silently wipe
        # the formula and leave the cell empty while reporting success.
        # Skip those cells instead of destroying their content.
        converted = 0
        skipped: list[str] = []
        # Coordinates are derived from the range bounds rather than read off the
        # cell. A streaming worksheet yields EmptyCell for blank cells, and
        # EmptyCell has no .coordinate -- reading it raised
        # "'EmptyCell' object has no attribute 'coordinate'" and failed the whole
        # call. iter_rows with explicit bounds behaves identically in both modes.
        # range_boundaries is typed as returning optionals for open-ended ranges
        # like "A:C". _validate_range above already requires both endpoints, so
        # these are always concrete -- coerced so the arithmetic below type-checks.
        bounds = range_boundaries(range_address.upper())
        min_col, min_row, max_col, max_row = (int(b or 1) for b in bounds)
        rows = ws_values.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        for row_offset, row in enumerate(rows):
            for col_offset, cell_v in enumerate(row):
                coord = f"{get_column_letter(min_col + col_offset)}{min_row + row_offset}"
                cell_w = ws[coord]
                # Only replace formula cells (value starts with "=")
                if isinstance(cell_w.value, str) and cell_w.value.startswith("="):
                    if cell_v.value is None:
                        skipped.append(coord)
                        continue
                    cell_w.value = cell_v.value
                    converted += 1

        wb_values.close()
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        # "Converted formulas to values" sat above "0 formulas replaced" -- the
        # headline claiming the thing its own detail said had not happened, on a
        # success:true reply. Not a rare corner, either: openpyxl writes formulas
        # without computing them, so a workbook produced anywhere in this fleet
        # has no cached values at all and this tool can convert nothing in it. A
        # round-15 phase hit exactly that on a file the xlsx server had just
        # written, and had to read skipped_no_cached_value to find out.
        if converted:
            progress.append(
                ok(
                    f"Converted formulas to values in {range_address}",
                    f"{converted} formula{'s' if converted != 1 else ''} replaced",
                )
            )
        else:
            progress.append(
                warn(
                    f"No formulas converted in {range_address}",
                    f"{len(skipped)} formula cell(s) had no cached value to write"
                    if skipped
                    else "that range holds no formula cells",
                )
            )
        # Every skipped cell address was listed in full, twice -- once joined
        # into this warning and once as the response field. A sweep converting a
        # filled-down column over 16,834 rows skipped all of them (openpyxl
        # writes no cached value), so the response carried ~33,000 cell
        # references: 260 KB, a token_estimate near 65,000, against the ~10-12k
        # context this server is built for. The count is the information; the
        # addresses are a sample.
        cap = get_max_cells()
        skipped_shown = skipped[:cap]
        if skipped:
            progress.append(
                warn(
                    f"Skipped {len(skipped)} formula cell(s) with no cached value",
                    ", ".join(skipped_shown) + (" ..." if len(skipped) > cap else ""),
                )
            )
        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "convert_to_values",
            "sheet": sheet_name,
            "range": range_address,
            "formulas_converted": converted,
            "skipped_no_cached_value": skipped_shown,
            "skipped_count": len(skipped),
            "truncated": len(skipped) > cap,
            "backup": backup,
            "progress": progress,
        }
        if skipped and converted:
            result["hint"] = (
                "Some formula cells had no cached value (never opened in Excel/LibreOffice) "
                "and were left as formulas rather than being overwritten with a blank value."
            )
        elif skipped:
            # Nothing at all was converted, so say what to do rather than only
            # what happened. Every workbook these servers write lands here.
            result["hint"] = (
                f"None of the {len(skipped)} formula cell(s) had a cached value, so nothing was "
                "converted and the file is unchanged. Formulas written by this server are never "
                "calculated -- openpyxl stores the text, not a result. Open the workbook in "
                "Excel or LibreOffice and save it once so the values are cached, then call "
                "convert_to_values again."
            )
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }
