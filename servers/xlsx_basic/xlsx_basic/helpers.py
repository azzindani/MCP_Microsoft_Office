"""XLSX Basic helpers — private utilities and advanced sheet operations."""

from __future__ import annotations

import logging
import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from shared.counts import counted
from shared.file_utils import drop_snapshot_if_unwritten, hint_for_error, resolve_path, scrub_repr, sheet_names_hint
from shared.live_edit import notify_reload
from shared.platform_utils import get_max_search_results, open_file
from shared.progress import fail, info, ok
from shared.version_control import snapshot

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Address validation helpers (used by engine.py)
# ---------------------------------------------------------------------------

_CELL_RE = re.compile(r"^[A-Z]{1,3}\d+$")
_RANGE_RE = re.compile(r"^[A-Z]{1,3}\d+:[A-Z]{1,3}\d+$")


def _validate_cell(address: str) -> bool:
    return bool(_CELL_RE.match(address.upper()))


def _validate_range(address: str) -> bool:
    return bool(_RANGE_RE.match(address.upper()))


# A canonical decimal and nothing else.
#
# set_cell's MCP signature is `value: str`, so every number arrives as text, and
# openpyxl stores text as text -- =SUM() over the column returns 0 and a chart
# built on it plots nothing. The obvious fix, float(value) on anything that
# parses, is worse than the bug: the values people most need kept as text are
# exactly the ones that parse. 07030 is a New Jersey ZIP code, not 7030; a part
# number like 1E5 becomes 100000.0; "1,234" and " 42 " are display formatting
# that should survive a round trip untouched.
#
# So: optional minus, no leading zeros (except a bare 0 or 0.x), no leading
# plus, no surrounding whitespace, no exponent, no thousands separator.
# Anything else stays text. Exponent form is refused even though Python reads
# it unambiguously -- a sheet full of silently converted part codes is a worse
# outcome than a caller having to send a real float.
_CANONICAL_NUMBER = re.compile(r"^-?(?:0|[1-9][0-9]*)(?:\.[0-9]+)?$")


# Renaming a sheet used to be `ws.title = new` and nothing else, which left
# every reference to the old name pointing at a sheet that no longer existed:
# chart series ('Data'!$A$2:$A$4), defined names, and cross-sheet formulas
# (=Data!B2) all survived the rename untouched. Excel draws the chart empty and
# shows #REF! in the formula, and the tool reported success.
#
# A sheet name appears in a reference either quoted ('Q3 Revenue'!A1) or bare
# (Data!A1). Both have to be rewritten, and neither may match a longer name
# that merely starts the same way -- renaming "Data" must not touch "Data2!A1".
# Hence the lookbehind on the bare form: a name character (or a quote, or the
# dot of a table reference) before it means this is not our sheet.
_SHEET_NAME_NEEDS_QUOTES = re.compile(r"[^A-Za-z0-9_]")


def _retarget_sheet_name(text: str, old: str, new: str) -> str:
    """Rewrite references to `old` in one formula/reference string."""
    esc = re.escape(old)
    replacement = f"'{new}'" if _SHEET_NAME_NEEDS_QUOTES.search(new) else new
    text = re.sub(r"'" + esc + r"'(?=!)", replacement, text)
    return re.sub(r"(?<![A-Za-z0-9_.'])" + esc + r"(?=!)", replacement, text)


def retarget_sheet_references(wb: Any, old: str, new: str) -> dict[str, int]:
    """Point charts, defined names and formulas at a renamed sheet.

    Returns how many of each were rewritten, so the caller can say what it
    touched rather than claiming a bare success.
    """
    counts = {"formulas": 0, "defined_names": 0, "chart_series": 0}

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                v = cell.value
                # `old in v`, not `f"{old}!" in v`: a quoted reference puts the
                # bang after the closing quote ('Old Name'!A1), so the stricter
                # pre-filter skipped every sheet whose name needed quoting --
                # exactly the names most likely to be renamed. The regexes below
                # do the precise matching; this only avoids running them on
                # formulas that cannot possibly match.
                if isinstance(v, str) and v.startswith("=") and old in v:
                    updated = _retarget_sheet_name(v, old, new)
                    if updated != v:
                        cell.value = updated
                        counts["formulas"] += 1

        # Charts are anchored to the sheet they are drawn on, which is often not
        # the sheet holding their data, so every sheet has to be walked.
        for chart in getattr(sheet, "_charts", []):
            for series in getattr(chart, "series", []):
                for holder in (
                    getattr(series, "val", None),
                    getattr(series, "cat", None),
                    getattr(series, "tx", None),
                ):
                    if holder is None:
                        continue
                    for kind in ("numRef", "strRef", "multiLvlStrRef"):
                        ref = getattr(holder, kind, None)
                        if ref is None:
                            continue
                        formula = getattr(ref, "f", None)
                        if not formula:
                            continue
                        updated = _retarget_sheet_name(formula, old, new)
                        if updated != formula:
                            ref.f = updated
                            counts["chart_series"] += 1

    for defined in wb.defined_names.values():
        text = getattr(defined, "attr_text", None)
        if not text:
            continue
        updated = _retarget_sheet_name(text, old, new)
        if updated != text:
            defined.attr_text = updated
            counts["defined_names"] += 1

    return counts


# openpyxl's copy_worksheet copies cells, styles and dimensions and documents
# that it does NOT copy charts, images or other drawing objects. copy_sheet
# passed that omission straight through: a sheet carrying a chart and a picture
# produced a copy with neither, reported as a plain success. The cells were
# right, so nothing looked wrong until you opened the file.
#
# Deep-copying the chart and re-anchoring it does work, and the copy's series
# are retargeted at the new sheet -- which is what Excel does when you duplicate
# a sheet by hand: the duplicate's chart plots the duplicate's data, not the
# original's.
def clone_drawings(source_ws: Any, target_ws: Any, old_name: str, new_name: str) -> dict[str, int]:
    """Copy charts and images onto a freshly copied sheet. Returns counts."""
    import copy as _copy

    counts = {"charts": 0, "images": 0}

    for chart in getattr(source_ws, "_charts", []):
        cloned = _copy.deepcopy(chart)
        for series in getattr(cloned, "series", []):
            for holder in (
                getattr(series, "val", None),
                getattr(series, "cat", None),
                getattr(series, "tx", None),
            ):
                if holder is None:
                    continue
                for kind in ("numRef", "strRef", "multiLvlStrRef"):
                    ref = getattr(holder, kind, None)
                    if ref is None:
                        continue
                    formula = getattr(ref, "f", None)
                    if formula:
                        ref.f = _retarget_sheet_name(formula, old_name, new_name)
        target_ws.add_chart(cloned, _anchor_cell(chart))
        counts["charts"] += 1

    for image in getattr(source_ws, "_images", []):
        target_ws.add_image(_copy.deepcopy(image), _anchor_cell(image))
        counts["images"] += 1

    return counts


def _anchor_cell(drawing: Any, default: str = "A1") -> str:
    """Where a chart or image sits, as an address, falling back to A1.

    The anchor is a OneCellAnchor/TwoCellAnchor whose _from carries 0-based col
    and row. Losing the position is much better than losing the drawing, so
    anything unexpected lands at A1 rather than raising.
    """
    anchor = getattr(drawing, "anchor", None)
    corner = getattr(anchor, "_from", None)
    if corner is None:
        return default
    try:
        return f"{get_column_letter(int(corner.col) + 1)}{int(corner.row) + 1}"
    except Exception:
        return default


def coerce_cell_value(value: Any) -> tuple[Any, str]:
    """Return (value to store, what it was stored as).

    The second element is reported to the caller as `stored_type`, so someone
    who sent "07030" can see it stayed text rather than discovering it in a
    broken SUM three steps later.
    """
    if value is None:
        return value, "empty"
    # bool before int: isinstance(True, int) is True in Python, so a boolean
    # would otherwise be reported as a number.
    if isinstance(value, bool):
        return value, "boolean"
    if isinstance(value, (int, float)):
        return value, "number"
    if not isinstance(value, str):
        return value, "other"
    if not _CANONICAL_NUMBER.match(value):
        return value, "text"
    return (float(value) if "." in value else int(value)), "number"


def _last_cell(ws: Any) -> str:
    """Return the Excel address of the last used cell in the worksheet."""
    col_letter = get_column_letter(ws.max_column) if ws.max_column else "A"
    row = ws.max_row or 1
    return f"{col_letter}{row}"


def _cell_count_for_range(range_address: str) -> int:
    """Return number of cells in an A1:C5 style range address."""
    top, bot = range_address.upper().split(":")
    top_col = re.sub(r"\d", "", top)
    top_row = int(re.sub(r"[A-Z]", "", top))
    bot_col = re.sub(r"\d", "", bot)
    bot_row = int(re.sub(r"[A-Z]", "", bot))
    cols = column_index_from_string(bot_col) - column_index_from_string(top_col) + 1
    rows = bot_row - top_row + 1
    return cols * rows


# ---------------------------------------------------------------------------
# Advanced sheet operations
# ---------------------------------------------------------------------------


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _header_row_index(rows: list[list[Any]]) -> int | None:
    """Index of the first row holding anything, or None if every row is blank.

    `has_header` used to mean "skip rows[0]" -- the *physical* first row, not
    the header. A sheet whose row 1 is blank therefore had its header sorted
    into the body as an ordinary value, and the write-back (which also started
    at a fixed row 2) landed the sorted block on top of it. success: true,
    rows_sorted correct, header gone:

        1  ⌀      ⌀            1  ⌀      ⌀
        2  name   qty   sort   2  alpha  3
        3  beta   2     ---->  3  beta   2
        4  alpha  3            4  gamma  1
        5  gamma  1            5  name   qty     <- header, sorted as data

    A leading blank or title row is ordinary in a real workbook, and
    insert_row(1) creates one, so two normal calls in sequence corrupted the
    file with no failure anywhere.
    """
    for i, row in enumerate(rows):
        if any(not _is_blank(v) for v in row):
            return i
    return None


def _sort_key(value: Any) -> tuple[int, float, str]:
    """Order a cell without ever comparing across types.

    The key was `(value is None, value)`, which hands mixed types straight to
    `<` and raises out of the tool as a bare
    `'<' not supported between instances of 'int' and 'str'`. A column holding
    numbers plus one stray text cell -- a stringified total, an "n/a", a header
    the caller did not declare -- is the common case, not the exotic one.

    Ranking by type first keeps every comparison within a single type: numbers
    before booleans before dates before text, and the second and third slots
    are only ever read against a value of the same rank.
    """
    if isinstance(value, bool):
        return (1, float(value), "")
    if isinstance(value, (int, float)):
        return (0, float(value), "")
    if isinstance(value, timedelta):
        return (0, value.total_seconds(), "")
    if isinstance(value, (datetime, date, time)):
        return (2, 0.0, value.isoformat())
    return (3, 0.0, str(value).casefold())


def sort_sheet(
    file_path: str,
    sheet_name: str,
    column: str,
    ascending: bool = True,
    has_header: bool = True,
    open_after: bool = False,
) -> dict[str, Any]:
    """Sort all rows in a sheet by the values in a given column."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        col_upper = column.upper()
        try:
            col_idx = column_index_from_string(col_upper) - 1
        except Exception:
            progress.append(fail(f"Invalid column letter: {column}"))
            return {
                "success": False,
                "error": f"Invalid column letter: {column}",
                "hint": "Use a single column letter like 'A', 'B', or 'C'.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]

        # A merged region spans rows, so reordering the rows underneath it
        # cannot preserve it -- and its non-anchor cells are read-only, so the
        # write-back below would fail on them anyway. Previously the None-skip
        # hid that: the sort "succeeded" and left the merged area holding
        # whichever rows happened to land under it.
        merged = [str(rng) for rng in ws.merged_cells.ranges]
        if merged:
            progress.append(fail(f"Sheet has {len(merged)} merged region(s)"))
            return {
                "success": False,
                "error": f"Cannot sort '{sheet_name}': it has merged cells ({', '.join(merged[:5])})",
                "hint": "Unmerge the region first — a merged block cannot follow the rows it spans.",
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 30,
            }

        # Read all rows as lists of cell values
        all_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        if not all_rows:
            progress.append(info("Sheet is empty — nothing to sort"))
            wb.close()
            return {
                "success": True,
                "op": "sort_sheet",
                "sheet": sheet_name,
                "rows_sorted": 0,
                "backup": backup,
                "progress": progress,
                "token_estimate": 15,
            }

        # column_index_from_string accepts far more than a column that exists.
        # "QTY" -- a caller passing the header name, which the docstring's
        # `column='A'` invites -- is a perfectly valid column string resolving
        # to index 12347, so the guard above lets it through and r[col_idx]
        # then indexes off the end of every row as `list index out of range`.
        width = max(len(r) for r in all_rows)
        if col_idx >= width:
            headers = [str(v) for v in all_rows[0][:width] if not _is_blank(v)]
            named = f" This sheet's column names are: {', '.join(headers)}." if headers else ""
            progress.append(fail(f"Column {col_upper} is beyond the data"))
            return {
                "success": False,
                "error": (
                    f"Column '{column}' is outside this sheet: it holds "
                    f"{width} column(s), A to {get_column_letter(width)}"
                ),
                "hint": (
                    f"Nothing was written. column= takes a column LETTER, not a header name."
                    f" Pass one between A and {get_column_letter(width)}.{named}"
                ),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 30,
            }

        if has_header:
            header_idx = _header_row_index(all_rows)
            if header_idx is None:
                progress.append(info("Sheet holds no values — nothing to sort"))
                wb.close()
                return {
                    "success": True,
                    "op": "sort_sheet",
                    "sheet": sheet_name,
                    "rows_sorted": 0,
                    "backup": backup,
                    "progress": progress,
                    "token_estimate": 15,
                }
            data_start = header_idx + 1
        else:
            data_start = 0
        data_rows = all_rows[data_start:]

        def _cell(row: list[Any]) -> Any:
            # Rows come back ragged when trailing cells were never written.
            return row[col_idx] if col_idx < len(row) else None

        # Blanks sink to the bottom in BOTH directions -- reverse= applied to a
        # (is_none, value) tuple floated them to the top on a descending sort,
        # which contradicted the comment that stood here.
        present = [r for r in data_rows if not _is_blank(_cell(r))]
        blank = [r for r in data_rows if _is_blank(_cell(r))]
        present.sort(key=lambda r: _sort_key(_cell(r)), reverse=not ascending)
        sorted_rows = present + blank

        # Write back cell by cell.
        #
        # Assigning .value rather than passing value= to ws.cell(): openpyxl's
        # cell() skips the assignment entirely when the value is None
        #
        #     cell = self._get_cell(row, column)
        #     if value is not None:
        #         cell.value = value
        #
        # so every blank cell in the sorted data left the *previous* occupant of
        # that address in place. Sorting three rows by column A turned
        #
        #     b, 2, ⌀          a, 1, ⌀
        #     c, ⌀, keep  into  b, 2, keep     <- c's note, on b's row
        #     a, 1, ⌀          c, 1, keep     <- b's n, on c's row
        #
        # with success:true and the ordering itself correct. On a real sheet
        # that is mass silent corruption: a sweep measured 541 blanks in one
        # column come back holding the value of whatever row had been there.
        start_row = data_start + 1
        for r_offset, row_vals in enumerate(sorted_rows):
            for c_offset, val in enumerate(row_vals):
                # MergedCell.value is read-only, but the guard above has
                # already refused any sheet that has a merged region.
                ws.cell(row=start_row + r_offset, column=c_offset + 1).value = val  # type: ignore[reportAttributeAccessIssue]

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "sort_sheet",
            "sheet": sheet_name,
            "column": col_upper,
            "ascending": ascending,
            "rows_sorted": len(sorted_rows),
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            # Everything this function can raise past the guards above comes
            # from the column it was told to sort by.
            "hint": hint_for_error(e, path, argument="column"),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def rename_sheet(
    file_path: str,
    old_name: str,
    new_name: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Rename a worksheet tab from old_name to new_name."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        if not new_name:
            progress.append(fail("new_name cannot be empty"))
            return {
                "success": False,
                "error": "new_name cannot be empty",
                "hint": "Provide a non-empty name for the sheet.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))
        if old_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{old_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{old_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }
        if new_name in wb.sheetnames:
            progress.append(fail(f"Sheet '{new_name}' already exists"))
            return {
                "success": False,
                "error": f"Sheet '{new_name}' already exists",
                "hint": "Choose a different name -- add_sheet cannot replace a sheet. Use list_sheets to see which names are taken, or rename_sheet to free this one.",
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        # Retarget before the title changes: the reference strings still carry
        # the old name either way, but doing it first keeps the two steps from
        # depending on each other's order in a way a later edit could break.
        retargeted = retarget_sheet_references(wb, old_name, new_name)
        wb[old_name].title = new_name
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"Renamed sheet '{old_name}' → '{new_name}'"))
        moved = sum(retargeted.values())
        if moved:
            progress.append(
                ok(
                    f"Repointed {moved} reference(s) at '{new_name}'",
                    f"{retargeted['formulas']} formula(s), {retargeted['defined_names']} defined name(s), "
                    f"{retargeted['chart_series']} chart series",
                )
            )
        result: dict[str, Any] = {
            "success": True,
            "op": "rename_sheet",
            "old_name": old_name,
            "new_name": new_name,
            # Say what else moved. A rename that silently broke every chart and
            # cross-sheet formula still reported a bare success.
            "references_updated": retargeted,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def find_duplicates(
    file_path: str,
    sheet_name: str,
    column: str,
    has_header: bool = True,
) -> dict[str, Any]:
    """Find duplicate values in a column. Returns rows where a value repeats."""
    progress: list[dict[str, Any]] = []
    try:
        col_upper = column.upper()
        try:
            col_idx = column_index_from_string(col_upper)
        except Exception:
            progress.append(fail(f"Invalid column letter: {column}"))
            return {
                "success": False,
                "error": f"Invalid column letter: {column}",
                "hint": "Use a single column letter like 'A', 'B', or 'C'.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            wb.close()
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]

        # Collect (value, row_number) pairs from the target column
        value_rows: dict[Any, list[int]] = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if has_header and row_idx == 1:
                continue
            if col_idx - 1 < len(row):
                val = row[col_idx - 1]
                if val is not None:
                    value_rows.setdefault(val, []).append(row_idx)

        wb.close()

        # Keep only values that appear more than once.
        #
        # The count of distinct duplicated values was capped at 100 and each
        # value's row list was not capped at all, so a low-cardinality column
        # answered with every row number it had: two values carrying 15,101 and
        # 1,733 rows put ~16,800 integers in one response and the transport
        # truncated it. What a caller needs is how many, not which 15,101.
        max_rows = get_max_search_results()
        duplicates = []
        rows_truncated = False
        for v, rows in value_rows.items():
            if len(rows) <= 1:
                continue
            entry: dict[str, Any] = {"value": v, "count": len(rows), "rows": rows[:max_rows]}
            if len(rows) > max_rows:
                entry["rows_truncated"] = True
                rows_truncated = True
            duplicates.append(entry)

        # Count before capping. This used to read len(duplicates) *after* the
        # slice below, so a column with thousands of repeated values answered
        # "duplicate_count: 100" -- wrong exactly when the column was worst, and
        # wrong in the one field a caller reads to decide what to do next.
        # `truncated` was already correct; the number beside it was not.
        #
        # The cap now comes from the same helper as the per-value row cap above.
        # It was a typed-in 100 sitting next to a derived limit, so the two
        # disagreed (50 vs 100 by default) and MCP_CONSTRAINED_MODE reached only
        # one of them.
        total_duplicate_values = len(duplicates)
        max_values = get_max_search_results()
        truncated = False
        if total_duplicate_values > max_values:
            duplicates = duplicates[:max_values]
            truncated = True

        progress.append(
            ok(
                f"Found {total_duplicate_values} duplicate value(s) in column {col_upper}"
                + (f", returning {len(duplicates)}" if truncated else ""),
                sheet_name,
            )
        )
        result: dict[str, Any] = {
            "success": True,
            "sheet": sheet_name,
            "column": col_upper,
            # How many exist, and how many came back -- two different numbers
            # whenever truncated is true, and conflating them was the defect.
            "duplicate_count": total_duplicate_values,
            "duplicates_returned": len(duplicates),
            "duplicates": duplicates,
            **counted(len(duplicates), total_duplicate_values),
            "rows_truncated": rows_truncated,
            "max_duplicates_returned": max_values,
            "max_rows_per_value": max_rows,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": "Check that file_path points to a valid .xlsx file.",
            "progress": progress,
            "token_estimate": 15,
        }


def copy_sheet(
    file_path: str,
    source_sheet: str,
    new_sheet_name: str,
    open_after: bool = False,
) -> dict[str, Any]:
    """Copy a worksheet within the same workbook to a new sheet name."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        if not new_sheet_name:
            progress.append(fail("new_sheet_name cannot be empty"))
            return {
                "success": False,
                "error": "new_sheet_name cannot be empty",
                "hint": "Provide a non-empty name for the copied sheet.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))
        if source_sheet not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{source_sheet}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{source_sheet}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }
        if new_sheet_name in wb.sheetnames:
            progress.append(fail(f"Sheet '{new_sheet_name}' already exists"))
            return {
                "success": False,
                "error": f"Sheet '{new_sheet_name}' already exists",
                "hint": "Choose a different name -- add_sheet cannot replace a sheet. Use list_sheets to see which names are taken, or rename_sheet to free this one.",
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        source_ws = wb[source_sheet]
        copied_ws = wb.copy_worksheet(source_ws)
        copied_ws.title = new_sheet_name
        drawings = clone_drawings(source_ws, copied_ws, source_sheet, new_sheet_name)

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"Copied '{source_sheet}' → '{new_sheet_name}'"))
        if drawings["charts"] or drawings["images"]:
            progress.append(
                ok(
                    f"Copied {drawings['charts']} chart(s) and {drawings['images']} image(s)",
                    f"chart series now read from '{new_sheet_name}'",
                )
            )
        result: dict[str, Any] = {
            "success": True,
            "op": "copy_sheet",
            "source_sheet": source_sheet,
            "new_sheet_name": new_sheet_name,
            # openpyxl copies no drawings at all, so a caller had no way to know
            # the copy was missing its chart short of opening the file.
            "drawings_copied": drawings,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }
