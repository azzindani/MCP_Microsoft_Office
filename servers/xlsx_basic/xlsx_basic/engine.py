"""XLSX Basic engine — pure openpyxl logic, zero MCP imports."""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from shared.counts import counted
from shared.file_utils import drop_snapshot_if_unwritten, hint_for_error, resolve_path, scrub_repr, sheet_names_hint
from shared.live_edit import notify_reload
from shared.platform_utils import get_max_cells, get_max_search_results, open_file
from shared.progress import fail, ok, warn
from shared.version_control import snapshot

from .helpers import _cell_count_for_range, _last_cell, _validate_cell, _validate_range, coerce_cell_value

logger = logging.getLogger(__name__)


def _coord(row: int, col: int) -> str:
    """Address of a cell from its 1-based position.

    Every read here streams the sheet (read_only=True), and a streaming
    worksheet yields EmptyCell for a blank cell. EmptyCell carries a value of
    None but has no .coordinate, so reading the address off the cell raised
    "'EmptyCell' object has no attribute 'coordinate'" and failed the whole
    call -- get_sheet_summary on any sheet with a gap in its header row, and
    read_cell_range on any range containing one blank cell. A coverage sweep
    hit the first of those on a workbook whose columns ran A-P and then X.

    xlsx_formulas' convert_to_values already derives its coordinates from the
    range bounds for exactly this reason. Deriving them here too means no
    caller of this module can reintroduce the crash by dropping a None check.
    """
    return f"{get_column_letter(col)}{row}"


# ---------------------------------------------------------------------------
# Tool implementations
# ---------------------------------------------------------------------------


def list_sheets(file_path: str) -> dict[str, Any]:
    """Return sheet names and row/col dimensions for every sheet."""
    progress: list[dict[str, Any]] = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            progress.append(fail(f"Wrong file type: {path.suffix}"))
            return {
                "success": False,
                "error": f"Expected .xlsx file, got {path.suffix}",
                "hint": "Use the correct server for this file type.",
                "progress": progress,
                "token_estimate": 20,
            }

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append(
                {
                    "name": name,
                    "max_row": ws.max_row or 0,
                    "max_col": ws.max_column or 0,
                    "last_cell": _last_cell(ws),
                }
            )
        wb.close()

        progress.append(ok(f"Listed sheets for {path.name}", f"{len(sheets)} sheet(s)"))
        result: dict[str, Any] = {
            "success": True,
            "file": str(path),
            "sheet_count": len(sheets),
            "sheets": sheets,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": "Check that file_path points to a valid .xlsx file.",
            "progress": progress,
            "token_estimate": 15,
        }


def get_sheet_summary(file_path: str, sheet_name: str) -> dict[str, Any]:
    """Return dimensions, header row, and first-column sample for a sheet."""
    progress: list[dict[str, Any]] = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            wb.close()
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        last_cell = _last_cell(ws)

        # Header row — first row as list of cell dicts
        header_row: list[dict[str, Any]] = []
        first_col_sample: list[Any] = []
        sample_count = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if row_idx == 1:
                for col_idx, cell in enumerate(row, start=1):
                    header_row.append({"cell": _coord(row_idx, col_idx), "value": cell.value})
            else:
                # First column sample — up to 5 non-empty values
                if sample_count < 5 and row:
                    first_cell = row[0]
                    if first_cell.value is not None:
                        first_col_sample.append({"cell": _coord(row_idx, 1), "value": first_cell.value})
                        sample_count += 1

        wb.close()

        more_rows = max_row - 1 - sample_count
        if more_rows > 0:
            first_col_sample.append(f"... {more_rows} more rows")

        progress.append(ok(f"Summarised sheet '{sheet_name}'", f"{max_row} rows, {max_col} cols"))
        result: dict[str, Any] = {
            "success": True,
            "sheet": sheet_name,
            "dimensions": {
                "rows": max_row,
                "cols": max_col,
                "last_cell": last_cell,
            },
            "header_row": header_row,
            "first_col_sample": first_col_sample,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": "Check that file_path points to a valid .xlsx file.",
            "progress": progress,
            "token_estimate": 15,
        }


def read_cell(file_path: str, sheet_name: str, cell_address: str) -> dict[str, Any]:
    """Return value, formula, and data type for a single cell."""
    progress: list[dict[str, Any]] = []
    addr = cell_address.upper()
    try:
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid cell address: {cell_address}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {cell_address}",
                "hint": "Use Excel notation like B5 or C12.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        # read_only is not an optimisation here, it is what keeps the server
        # alive. Without it openpyxl builds a Python object for every cell in
        # the file, so reading one cell out of a 16,834 x 16 sheet allocated
        # ~510 MB against this container's 512 MB limit and killed the process
        # -- taking all twelve Office sub-servers down with it, since they
        # share one container. Streaming the same read peaks at 37 MB.
        wb_val = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if sheet_name not in wb_val.sheetnames:
            available_sheets = list(wb_val.sheetnames)
            wb_val.close()
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "progress": progress,
                "token_estimate": 15,
            }
        ws_val = wb_val[sheet_name]
        value = ws_val[addr].value
        wb_val.close()

        # Second pass for the formula string: the cached value and the formula
        # never come from the same load, so both are needed -- but both stream.
        wb_form = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
        ws_form = wb_form[sheet_name]
        raw = ws_form[addr].value
        formula = raw if isinstance(raw, str) and raw.startswith("=") else None
        wb_form.close()

        # Determine type
        if value is None and formula is not None:
            # A cell holding a formula is not an empty cell. openpyxl has no
            # calculation engine, so a formula this server wrote and nothing
            # has opened since carries no cached result -- and reporting that
            # as `value: null, type: "empty"` reads as "the write did not
            # land", which is the opposite of what happened. The formula is
            # right there in the same response.
            cell_type = "formula_uncalculated"
        elif value is None:
            cell_type = "empty"
        elif isinstance(value, bool):
            cell_type = "boolean"
        elif isinstance(value, (int, float)):
            cell_type = "number"
        else:
            cell_type = "string"

        progress.append(ok(f"Read cell {addr}", sheet_name))
        result: dict[str, Any] = {
            "success": True,
            "sheet": sheet_name,
            "cell": addr,
            "value": value,
            "formula": formula,
            "type": cell_type,
            "progress": progress,
        }
        if cell_type == "formula_uncalculated":
            result["note"] = (
                f"{addr} holds the formula {formula!r} and no cached result. Excel and LibreOffice write the "
                "computed value into the file when they save; this server stores formulas without evaluating "
                "them, so there is nothing to read back until the file is opened in one of them."
            )
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": "Check that file_path points to a valid .xlsx file.",
            "progress": progress,
            "token_estimate": 15,
        }


def read_cell_range(file_path: str, sheet_name: str, range_address: str) -> dict[str, Any]:
    """Return a bounded 2D cell array. Max 200 cells."""
    progress: list[dict[str, Any]] = []
    rng = range_address.upper()
    try:
        if not _validate_range(rng):
            progress.append(fail(f"Invalid range address: {range_address}"))
            return {
                "success": False,
                "error": f"Invalid range address: {range_address}",
                "hint": "Use Excel notation like A1:D10.",
                "progress": progress,
                "token_estimate": 15,
            }

        cell_count = _cell_count_for_range(rng)
        max_cells = get_max_cells()
        if cell_count > max_cells:
            progress.append(fail(f"Range too large: {cell_count} cells"))
            return {
                "success": False,
                "error": f"Range {rng} contains {cell_count} cells (max {max_cells})",
                "hint": f"Split into smaller ranges of {max_cells} cells or fewer.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        # Streamed for the same reason as read_cell: this range is capped at 200
        # cells, but a non-read_only load materialises the whole sheet regardless
        # of how little of it the caller asked for.
        wb_val = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if sheet_name not in wb_val.sheetnames:
            available_sheets = list(wb_val.sheetnames)
            wb_val.close()
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "progress": progress,
                "token_estimate": 15,
            }
        ws_val = wb_val[sheet_name]

        wb_form = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
        ws_form = wb_form[sheet_name]

        # _validate_range above requires both endpoints, so the bounds are
        # always concrete; the coercion is for the type checker.
        min_col, min_row, _max_col, _max_row = (int(b or 1) for b in range_boundaries(rng))

        data: list[list[dict[str, Any]]] = []
        for row_offset, (row_val, row_form) in enumerate(zip(ws_val[rng], ws_form[rng])):
            row_data = []
            for col_offset, (cell_v, cell_f) in enumerate(zip(row_val, row_form)):
                raw = cell_f.value
                formula = raw if isinstance(raw, str) and raw.startswith("=") else None
                row_data.append(
                    {
                        "cell": _coord(min_row + row_offset, min_col + col_offset),
                        "value": cell_v.value,
                        "formula": formula,
                    }
                )
            data.append(row_data)

        wb_val.close()
        wb_form.close()

        progress.append(ok(f"Read range {rng}", f"{cell_count} cells from '{sheet_name}'"))
        result: dict[str, Any] = {
            "success": True,
            "sheet": sheet_name,
            "range": rng,
            "cell_count": cell_count,
            "data": data,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": "Check that file_path points to a valid .xlsx file.",
            "progress": progress,
            "token_estimate": 15,
        }


def search_cells(
    file_path: str,
    sheet_name: str,
    query: str,
    max_results: int = 20,
) -> dict[str, Any]:
    """Scan cell values for text matching query. Returns matching addresses only."""
    progress: list[dict[str, Any]] = []
    try:
        if not query:
            progress.append(fail("Query cannot be empty"))
            return {
                "success": False,
                "error": "query cannot be empty",
                "hint": "Provide a non-empty search string.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        cap = min(max_results, get_max_search_results())

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            wb.close()
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]
        matches: list[dict[str, Any]] = []
        query_lower = query.lower()
        total_scanned = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            for col_idx, cell in enumerate(row, start=1):
                total_scanned += 1
                if cell.value is not None and query_lower in str(cell.value).lower():
                    matches.append({"cell": _coord(row_idx, col_idx), "value": cell.value})
                    if len(matches) > cap:
                        break
            if len(matches) > cap:
                break

        wb.close()

        # One past the cap. `len(matches) >= cap` cannot distinguish "exactly
        # cap matches exist" from "more exist", so a sheet with precisely that
        # many came back truncated -- the third copy of this mistake in this
        # repo, after docx_basic and docx_tables.
        truncated = len(matches) > cap
        found = len(matches)
        matches = matches[:cap]
        progress.append(
            ok(
                f"Found {len(matches)} match(es) for '{query}'",
                f"scanned {total_scanned} cells",
            )
        )
        result: dict[str, Any] = {
            "success": True,
            "sheet": sheet_name,
            "query": query,
            "matches": matches,
            # The cells read, which the loop stops early, so it is named for
            # what it measures rather than for the size of the sheet. It is not
            # the denominator for the matches; that one is below.
            "total_cells_scanned": total_scanned,
            **counted(len(matches), found, exact=not truncated),
            "progress": progress,
        }
        if not matches:
            result["hint"] = "Try a different search term or check the sheet name."
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": "Check that file_path points to a valid .xlsx file.",
            "progress": progress,
            "token_estimate": 15,
        }


def set_cell(
    file_path: str, sheet_name: str, cell_address: str, value: Any, open_after: bool = False
) -> dict[str, Any]:
    """Write a value to a single cell by address."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    addr = cell_address.upper()
    path: Path | None = None
    try:
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid cell address: {cell_address}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {cell_address}",
                "hint": "Use Excel notation like B5 or C12.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]
        stored, stored_type = coerce_cell_value(value)
        ws[addr] = stored
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_cell",
            "sheet": sheet_name,
            "cell": addr,
            # The stored value, not the string that arrived. Echoing the input
            # while storing something else would be its own small lie.
            "value": stored,
            "stored_type": stored_type,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def set_range(
    file_path: str,
    sheet_name: str,
    start_cell: str,
    data: list[list[Any]],
    open_after: bool = False,
) -> dict[str, Any]:
    """Write a 2D list of values starting at start_cell."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    addr = start_cell.upper()
    path: Path | None = None
    try:
        if not _validate_cell(addr):
            progress.append(fail(f"Invalid start cell: {start_cell}"))
            return {
                "success": False,
                "error": f"Invalid cell address: {start_cell}",
                "hint": "Use Excel notation like B5 or C12.",
                "progress": progress,
                "token_estimate": 15,
            }

        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]

        # Parse start cell into row/col
        col_letter = re.sub(r"\d", "", addr)
        start_row = int(re.sub(r"[A-Z]", "", addr))
        start_col = column_index_from_string(col_letter)

        total_cells = 0
        for r_offset, row_data in enumerate(data):
            for c_offset, val in enumerate(row_data):
                # .value rather than ws.cell(value=...), which skips the write
                # when the value is None and so leaves whatever the cell held
                # before -- counted in cells_written all the same. "" already
                # clears correctly; a JSON null did not.
                #
                # Same coercion as set_cell: two tools that write cells must not
                # disagree about whether "10" is a number.
                stored, _ = coerce_cell_value(val)
                ws.cell(row=start_row + r_offset, column=start_col + c_offset).value = stored
                total_cells += 1

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "set_range",
            "sheet": sheet_name,
            "start_cell": addr,
            "rows_written": len(data),
            "cells_written": total_cells,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def insert_row(file_path: str, sheet_name: str, row_index: int, open_after: bool = False) -> dict[str, Any]:
    """Insert an empty row at row_index (1-based), shifting existing rows down."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]
        ws.insert_rows(row_index)
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "insert_row",
            "sheet": sheet_name,
            "row_index": row_index,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def delete_row(file_path: str, sheet_name: str, row_index: int, open_after: bool = False) -> dict[str, Any]:
    """Remove row at row_index (1-based), shifting remaining rows up."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            available_sheets = list(wb.sheetnames)
            progress.append(fail(f"Sheet '{sheet_name}' not found"))
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found",
                "hint": sheet_names_hint(available_sheets),
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        if row_index < 1 or row_index > max_row:
            progress.append(fail(f"Row {row_index} out of range", f"sheet has {max_row} rows"))
            return {
                "success": False,
                "error": f"row_index {row_index} out of range (1-{max_row})",
                "hint": "Use list_sheets to get current row count.",
                "backup": drop_snapshot_if_unwritten(backup, path, progress),
                "progress": progress,
                "token_estimate": 15,
            }

        ws.delete_rows(row_index)
        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(notify_reload(str(path), "xlsx"))
        result: dict[str, Any] = {
            "success": True,
            "op": "delete_row",
            "sheet": sheet_name,
            "row_index": row_index,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }


def add_sheet(file_path: str, sheet_name: str = "", open_after: bool = False) -> dict[str, Any]:
    """Create a new worksheet, optionally with a given name."""
    progress: list[dict[str, Any]] = []
    backup: str | None = None
    path: Path | None = None
    try:
        path = resolve_path(file_path)
        if not path.exists():
            progress.append(fail("File not found", str(path)))
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": progress,
                "token_estimate": 20,
            }

        backup = snapshot(str(path))
        progress.append(ok("Snapshot saved", Path(backup).name))

        wb = openpyxl.load_workbook(str(path))

        if sheet_name:
            if sheet_name in wb.sheetnames:
                progress.append(warn(f"Sheet '{sheet_name}' already exists"))
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' already exists",
                    "hint": "Choose a different name -- add_sheet cannot replace a sheet. Use list_sheets to see which names are taken, or rename_sheet to free this one.",
                    "backup": drop_snapshot_if_unwritten(backup, path, progress),
                    "progress": progress,
                    "token_estimate": 15,
                }
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb.create_sheet()
            sheet_name = ws.title

        wb.save(str(path))
        wb.close()
        if open_after:
            open_file(path)

        progress.append(ok(f"Created sheet '{sheet_name}'"))
        result: dict[str, Any] = {
            "success": True,
            "op": "add_sheet",
            "sheet": sheet_name,
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = len(str(result)) // 4
        return result

    except Exception as e:
        progress.append(fail(str(e)))
        return {
            "success": False,
            "error": scrub_repr(e),
            "hint": hint_for_error(e, path),
            "backup": drop_snapshot_if_unwritten(backup, path, progress),
            "progress": progress,
            "token_estimate": 15,
        }
