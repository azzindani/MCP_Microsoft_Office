"""XLSX New engine — create Excel workbooks from scratch. Zero MCP imports."""

from __future__ import annotations

import logging
import re
import sys
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# sys.path bootstrap so 'shared' is importable when run directly
# ---------------------------------------------------------------------------
_ROOT = Path(__file__).resolve().parents[2]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

from shared.file_utils import embed_content  # noqa: E402
from shared.platform_utils import open_file, resolve_output_path  # noqa: E402
from shared.progress import fail, info, ok, warn  # noqa: E402

logging.basicConfig(stream=sys.stderr, level=logging.WARNING)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _ensure_parent(path: Path) -> None:
    """Create parent directories if they do not exist."""
    path.parent.mkdir(parents=True, exist_ok=True)


def _token_estimate(obj: Any) -> int:
    return len(str(obj)) // 4


def _write_headers(ws: Any, headers: list[Any]) -> None:
    """Write header row in row 1, bold."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)


def _write_rows(ws: Any, rows: list[list[Any]], start_row: int = 2) -> None:
    """Write data rows starting at start_row."""
    for row_idx, row_data in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _fit_columns(ws: Any, scan_rows: int = 200, max_width: int = 60) -> None:
    """Widen each column to fit what is in it.

    openpyxl leaves every column at the default 8.43 characters, so a sheet
    these tools created was unreadable the moment anyone looked at it: a report
    header row rendered as "platform spends impressioclicks" -- "impressions"
    running into the column beside it -- and "Google Ads" as "Google Ad". The
    values are all in the file, which is why every structural check passed; it
    shows up only in a render or a print.

    Only the first `scan_rows` rows are measured, so create_from_csv on a
    16,834-row export does not walk the whole sheet to size four columns, and
    the width is capped so one long cell cannot push the rest off the page.
    """
    from openpyxl.utils import get_column_letter

    widths: dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, scan_rows)):
        for cell in row:
            text = str(cell.value) if cell.value is not None else ""
            if text and len(text) > widths.get(cell.column, 0):
                widths[cell.column] = len(text)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, max_width)


def _fit_to_one_page_wide(ws: Any) -> None:
    """Print the sheet's full width on a single page.

    _fit_columns above sizes columns to their contents, which is what makes a
    generated sheet readable -- and on a document-shaped sheet it can push the
    total past the printable width. create_invoice's description column
    auto-fits to 41 characters, taking the table to 77 characters (~7.7in)
    against roughly 6.9in of printable A4, so the Total column printed on a
    second page: every amount, the subtotal and the grand total, separated from
    the labels that named them. The money was on page 2 of a two-page invoice.

    Scaling to one page wide is what a person does in Excel for this, and it is
    inert when the content already fits. Only applied to sheets meant to be read
    as a document -- a raw data export of 16 columns would be scaled to nothing.
    """
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = as many pages tall as it needs


_LEADING_ZERO_RE = re.compile(r"^-?0\d")


def _coerce_csv_value(value: str) -> Any:
    """Convert a raw CSV string to int/float when it unambiguously
    represents a number, so SUM/pivot/chart formulas can use imported
    data as real numbers instead of silently treating it as text.
    Zero-padded strings (zip codes, IDs) are kept as text, matching
    Excel's own CSV-import convention."""
    if value == "":
        return None
    if _LEADING_ZERO_RE.match(value):
        return value
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


# ---------------------------------------------------------------------------
# Public engine functions
# ---------------------------------------------------------------------------


def create_workbook(
    output_path: str,
    sheet_name: str = "Sheet1",
    open_after: bool = True,
    return_content: bool = False,
) -> dict[str, Any]:
    """Create a blank Excel workbook with one sheet."""
    progress: list[dict[str, Any]] = []
    try:
        path = resolve_output_path(output_path, "workbook.xlsx")
        _ensure_parent(path)
        progress.append(info("Creating blank workbook", path.name))

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        wb.save(str(path))
        progress.append(ok(f"Saved {path.name}", f"sheet: {sheet_name}"))

        if open_after:
            open_file(path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_workbook",
            "output": str(path),
            "output_name": path.name,
            "sheet_name": sheet_name,
            "progress": progress,
        }
        embed_content(result, path, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": "Check that output_path is a valid file path and you have write permission.",
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_from_data(
    output_path: str,
    sheet_name: str,
    headers: list[Any],
    rows: list[list[Any]],
    open_after: bool = True,
    return_content: bool = False,
) -> dict[str, Any]:
    """Create an Excel workbook from headers and data rows."""
    progress: list[dict[str, Any]] = []
    try:
        path = resolve_output_path(output_path, "workbook.xlsx")
        _ensure_parent(path)

        col_count = len(headers)
        row_count = len(rows)
        progress.append(
            info(
                "Creating workbook from data",
                f"{row_count} rows, {col_count} columns",
            )
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        _write_headers(ws, headers)
        _write_rows(ws, rows, start_row=2)
        progress.append(
            ok(
                f"Wrote {row_count} data rows",
                "headers in row 1, bold",
            )
        )

        _fit_columns(ws)
        wb.save(str(path))
        progress.append(ok(f"Saved {path.name}"))

        if open_after:
            open_file(path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_from_data",
            "output": str(path),
            "output_name": path.name,
            "sheet_name": sheet_name,
            "row_count": row_count,
            "column_count": col_count,
            "progress": progress,
        }
        embed_content(result, path, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "Ensure headers is a list of strings and rows is a list of lists. Check that output_path is writable."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_report(
    output_path: str,
    title: str,
    sheets: list[dict[str, Any]],
    open_after: bool = True,
    return_content: bool = False,
) -> dict[str, Any]:
    """Create a multi-sheet Excel workbook with a Cover sheet."""
    progress: list[dict[str, Any]] = []
    try:
        path = resolve_output_path(output_path, "workbook.xlsx")
        _ensure_parent(path)

        sheet_count = len(sheets)
        progress.append(
            info(
                "Creating multi-sheet report",
                f"{sheet_count} data sheet(s) + Cover",
            )
        )

        wb = openpyxl.Workbook()

        # Cover sheet — reuse the default active sheet
        cover = wb.active
        if cover is None:
            cover = wb.create_sheet()
        cover.title = "Cover"
        cover["A1"] = title
        cover["A1"].font = Font(bold=True, size=16)
        progress.append(ok("Created Cover sheet", title))

        # Data sheets
        for sheet_def in sheets:
            name = sheet_def.get("name", "Sheet")
            headers = sheet_def.get("headers", [])
            rows = sheet_def.get("rows", [])
            ws = wb.create_sheet(title=name)
            _write_headers(ws, headers)
            _write_rows(ws, rows, start_row=2)
            progress.append(
                ok(
                    f"Created sheet '{name}'",
                    f"{len(rows)} rows, {len(headers)} columns",
                )
            )

        for sheet in wb.worksheets:
            _fit_columns(sheet)
            _fit_to_one_page_wide(sheet)
        wb.save(str(path))
        progress.append(ok(f"Saved {path.name}", f"{sheet_count + 1} sheets total"))

        if open_after:
            open_file(path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_report",
            "output": str(path),
            "output_name": path.name,
            "title": title,
            "sheets_created": sheet_count + 1,  # includes Cover
            "progress": progress,
        }
        embed_content(result, path, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "sheets must be a list of dicts with 'name', 'headers', and 'rows' keys. "
                "Check that output_path is writable."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_from_template(
    template_path: str,
    output_path: str,
    substitutions: dict[str, Any],
    open_after: bool = True,
    return_content: bool = False,
) -> dict[str, Any]:
    """Copy a .xlsx template, apply {key: value} substitutions, save to output_path."""
    progress: list[dict[str, Any]] = []
    try:
        src = Path(template_path).resolve()
        if not src.exists():
            progress.append(fail("Template not found", str(src)))
            return {
                "success": False,
                "error": f"File not found: {src}",
                "hint": "Check that template_path is an absolute path to an existing .xlsx file.",
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }
        if src.suffix.lower() not in {".xlsx", ".xlsm"}:
            progress.append(fail("Wrong file type", src.suffix))
            return {
                "success": False,
                "error": f"Expected .xlsx file, got {src.suffix}",
                "hint": "template_path must point to a .xlsx or .xlsm file.",
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        dst = resolve_output_path(output_path, "workbook.xlsx")
        _ensure_parent(dst)

        progress.append(info("Loading template", src.name))
        wb = openpyxl.load_workbook(str(src))

        replaced_count = 0
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and cell.value in substitutions:
                        cell.value = substitutions[cell.value]  # type: ignore[reportArgumentType]
                        replaced_count += 1

        progress.append(
            ok(
                f"Replaced {replaced_count} cell value(s)",
                f"{len(substitutions)} substitution key(s) searched",
            )
        )

        wb.save(str(dst))
        progress.append(ok(f"Saved {dst.name}"))

        if open_after:
            open_file(dst)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_from_template",
            "template": str(src),
            "output": str(dst),
            "output_name": dst.name,
            # Named to match docx and pptx: reading one tier's response must not
            # KeyError on another's. The cell-level detail is in progress.
            "substitutions_applied": replaced_count,
            "progress": progress,
        }
        embed_content(result, dst, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": ("Ensure template_path points to a valid .xlsx file and output_path is a writable destination."),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_from_csv(
    csv_path: str,
    output_path: str,
    sheet_name: str = "Data",
    delimiter: str = ",",
    has_header: bool = True,
    open_after: bool = True,
    return_content: bool = False,
) -> dict[str, Any]:
    """Import a CSV file into a new Excel workbook."""
    progress: list[dict[str, Any]] = []
    try:
        import csv

        src = Path(csv_path).resolve()
        if not src.exists():
            progress.append(fail("CSV file not found", str(src)))
            return {
                "success": False,
                "error": f"File not found: {csv_path}",
                "hint": "Check that csv_path is an absolute path to an existing CSV file.",
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        out_path = resolve_output_path(output_path, "workbook.xlsx")
        _ensure_parent(out_path)
        progress.append(info("Reading CSV", src.name))

        all_rows: list[list[Any]] = []
        with open(str(src), newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            for row in reader:
                all_rows.append(list(row))

        if len(all_rows) == 0:
            progress.append(warn("CSV file is empty"))

        row_count = len(all_rows) - (1 if has_header else 0)
        col_count = max((len(r) for r in all_rows), default=0)
        progress.append(ok(f"Read {len(all_rows)} row(s)", f"{col_count} column(s)"))

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        for r_idx, row_data in enumerate(all_rows, start=1):
            is_header_row = has_header and r_idx == 1
            for c_idx, value in enumerate(row_data, start=1):
                cell_value = value if is_header_row else _coerce_csv_value(value)
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                if is_header_row:
                    cell.font = Font(bold=True)

        _fit_columns(ws)
        wb.save(str(out_path))
        progress.append(ok(f"Saved {out_path.name}", f"{row_count} data row(s)"))

        if open_after:
            open_file(out_path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_from_csv",
            "output": str(out_path),
            "output_name": out_path.name,
            "row_count": row_count,
            "column_count": col_count,
            "progress": progress,
        }
        embed_content(result, out_path, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "Check that csv_path points to a readable CSV file and output_path is a writable .xlsx destination."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


_ITEM_KEYS: dict[str, tuple[str, ...]] = {
    "description": ("description", "desc", "item", "name", "product", "details", "service"),
    "quantity": ("quantity", "qty", "count", "units", "hours", "amount"),
    "unit_price": ("unit_price", "unitprice", "price", "rate", "unit_cost", "cost"),
}


def _pick(item: dict, field: str):
    """First value in the item under any name this field is known by."""
    lowered = {str(k).strip().lower().replace(" ", "_"): v for k, v in item.items()}
    for alias in _ITEM_KEYS[field]:
        if alias in lowered and lowered[alias] not in (None, ""):
            return lowered[alias]
    return None


def _read_invoice_items(items: list) -> tuple[list[tuple[str, float, float]], list[tuple[int, str]]]:
    """Split items into ones that carry a price and ones that do not.

    `items` is a bare `list` in the schema, so the key names exist nowhere a
    caller can read -- the only mention was inside the error hint for an *empty*
    list. A sweep sent {"description":..., "qty":..., "price":...} and got a
    complete, correctly formatted invoice in which every quantity and price was
    0, subtotal 0.0, under success: true. Zeroing money silently is worse than
    refusing, so the aliases are honoured and anything still unpriced is named.
    """
    priced: list[tuple[str, float, float]] = []
    unpriced: list[tuple[int, str]] = []
    for index, item in enumerate(items):
        if not isinstance(item, dict):
            unpriced.append((index, f"not an object ({type(item).__name__})"))
            continue
        quantity = _pick(item, "quantity")
        unit_price = _pick(item, "unit_price")
        if quantity is None and unit_price is None:
            unpriced.append((index, ", ".join(str(k) for k in item)))
            continue
        try:
            qty_value = float(quantity if quantity is not None else 0)
            price_value = float(unit_price if unit_price is not None else 0)
        except (TypeError, ValueError):
            unpriced.append((index, ", ".join(str(k) for k in item)))
            continue
        priced.append((str(_pick(item, "description") or ""), qty_value, price_value))
    return priced, unpriced


def create_invoice(
    output_path: str,
    company_name: str,
    client_name: str,
    invoice_number: str,
    items: list,
    tax_rate: float = 0.0,
    currency: str = "USD",
    open_after: bool = True,
    return_content: bool = False,
) -> dict[str, Any]:
    """Create a formatted invoice .xlsx with items, totals, and tax formula."""
    progress: list[dict[str, Any]] = []
    try:
        from openpyxl.styles import PatternFill  # type: ignore[import-untyped]

        if not isinstance(items, list) or len(items) == 0:
            progress.append(fail("items must be a non-empty list"))
            return {
                "success": False,
                "error": "items must be a non-empty list",
                "hint": ('Pass a list like [{"description":"Widget","quantity":2,"unit_price":50.0}].'),
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        # tax_rate is a fraction: the label multiplies it by 100 and the formula
        # multiplies the subtotal by it directly. Nothing said so -- the schema
        # has no descriptions and the docstring has no room -- so a sweep passed
        # 5.0 meaning 5%, and got a finished invoice reading "Tax (500.0%)" with
        # a total of 15,060,713 against a subtotal of 2,510,119. success: true.
        # No tax anywhere is above 100%, so a rate above 1 is a unit mistake and
        # is worth refusing rather than quietly billing five times the goods.
        if tax_rate < 0 or tax_rate > 1:
            progress.append(fail(f"tax_rate {tax_rate} is not a fraction"))
            return {
                "success": False,
                "error": f"tax_rate must be a fraction between 0 and 1, got {tax_rate}",
                "hint": (
                    f"Pass 0.05 for 5%, not 5. For {tax_rate}% use tax_rate={tax_rate / 100:g}."
                    if tax_rate > 1
                    else "Pass 0.05 for 5%. Use 0 for no tax."
                ),
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        priced, unpriced = _read_invoice_items(items)
        if unpriced:
            first_index, first_keys = unpriced[0]
            progress.append(fail(f"item {first_index} carries no quantity or price"))
            return {
                "success": False,
                "error": (
                    f"{len(unpriced)} of {len(items)} invoice item(s) have no quantity or unit "
                    f"price. Item {first_index} has keys: {first_keys or 'none'}"
                ),
                "hint": (
                    'Each item needs a number under "quantity" and one under "unit_price", as in '
                    '[{"description":"Widget","quantity":2,"unit_price":50.0}]. '
                    '"qty", "price", "rate" and "cost" are accepted as well.'
                ),
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        out_path = resolve_output_path(output_path, "workbook.xlsx")
        _ensure_parent(out_path)
        progress.append(info("Creating invoice", out_path.name))

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Invoice"

        # Header block
        ws["A1"] = company_name
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = "INVOICE"
        ws["A3"] = f"Invoice #: {invoice_number}"
        ws["A4"] = f"Client: {client_name}"

        # Column headers in row 6
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        col_headers = ["Description", "Quantity", "Unit Price", "Total"]
        for c_idx, label in enumerate(col_headers, start=1):
            cell = ws.cell(row=6, column=c_idx, value=label)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        progress.append(ok("Written invoice header"))

        # Item rows starting at row 7
        item_start_row = 7
        subtotal = 0.0
        for i, (description, quantity, unit_price) in enumerate(priced):
            row_num = item_start_row + i
            ws.cell(row=row_num, column=1, value=description)
            ws.cell(row=row_num, column=2, value=quantity)
            ws.cell(row=row_num, column=3, value=unit_price)
            # Total formula: =B{row}*C{row}
            ws.cell(row=row_num, column=4, value=f"=B{row_num}*C{row_num}")
            subtotal += float(quantity) * float(unit_price)

        last_item_row = item_start_row + len(items) - 1
        progress.append(ok(f"Written {len(items)} item row(s)"))

        # Subtotal row
        subtotal_row = last_item_row + 2
        ws.cell(row=subtotal_row, column=3, value="Subtotal")
        ws.cell(row=subtotal_row, column=3).font = Font(bold=True)
        ws.cell(
            row=subtotal_row,
            column=4,
            value=f"=SUM(D{item_start_row}:D{last_item_row})",
        )

        current_row = subtotal_row

        # Tax row (only if tax_rate > 0)
        if tax_rate > 0:
            tax_row = subtotal_row + 1
            ws.cell(row=tax_row, column=3, value=f"Tax ({tax_rate * 100:.1f}%)")
            ws.cell(row=tax_row, column=4, value=f"=D{subtotal_row}*{tax_rate}")
            current_row = tax_row

        # Total row
        total_row = current_row + 1
        ws.cell(row=total_row, column=3, value=f"Total ({currency})")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        if tax_rate > 0:
            ws.cell(
                row=total_row,
                column=4,
                value=f"=D{subtotal_row}+D{current_row}",
            )
        else:
            ws.cell(row=total_row, column=4, value=f"=D{subtotal_row}")
        ws.cell(row=total_row, column=4).font = Font(bold=True)

        progress.append(ok("Written subtotal, tax, and total rows"))

        _fit_columns(ws)
        _fit_to_one_page_wide(ws)
        wb.save(str(out_path))
        progress.append(ok(f"Saved {out_path.name}"))

        if open_after:
            open_file(out_path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_invoice",
            "output": str(out_path),
            "output_name": out_path.name,
            "item_count": len(items),
            "subtotal": round(subtotal, 2),
            "currency": currency,
            "progress": progress,
        }
        embed_content(result, out_path, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "Ensure items is a list of dicts with 'description', 'quantity', "
                "and 'unit_price' keys. Check that output_path is writable."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }
