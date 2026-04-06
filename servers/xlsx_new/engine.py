"""XLSX New engine — create Excel workbooks from scratch. Zero MCP imports."""
from __future__ import annotations

import logging
import sys
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# sys.path bootstrap so 'shared' is importable when run directly
# ---------------------------------------------------------------------------
_ROOT = Path(__file__).resolve().parents[2]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from shared.platform_utils import open_file
from shared.progress import fail, info, ok, warn

logging.basicConfig(stream=sys.stderr, level=logging.WARNING)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _ensure_parent(path: Path) -> None:
    """Create parent directories if they do not exist."""
    path.parent.mkdir(parents=True, exist_ok=True)


def _token_estimate(obj: Any) -> int:
    return len(str(obj)) // 4


def _write_headers(ws: Any, headers: list[Any]) -> None:
    """Write header row in row 1, bold."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)


def _write_rows(ws: Any, rows: list[list[Any]], start_row: int = 2) -> None:
    """Write data rows starting at start_row."""
    for row_idx, row_data in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


# ---------------------------------------------------------------------------
# Public engine functions
# ---------------------------------------------------------------------------


def create_workbook(
    output_path: str,
    sheet_name: str = "Sheet1",
    open_after: bool = True,
) -> dict[str, Any]:
    """Create a blank Excel workbook with one sheet."""
    progress: list[dict[str, Any]] = []
    try:
        path = Path(output_path).resolve()
        _ensure_parent(path)
        progress.append(info("Creating blank workbook", path.name))

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        wb.save(str(path))
        progress.append(ok(f"Saved {path.name}", f"sheet: {sheet_name}"))

        if open_after:
            open_file(path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_workbook",
            "output": str(path),
            "output_name": path.name,
            "sheet_name": sheet_name,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": "Check that output_path is a valid file path and you have write permission.",
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_from_data(
    output_path: str,
    sheet_name: str,
    headers: list[Any],
    rows: list[list[Any]],
    open_after: bool = True,
) -> dict[str, Any]:
    """Create an Excel workbook from headers and data rows."""
    progress: list[dict[str, Any]] = []
    try:
        path = Path(output_path).resolve()
        _ensure_parent(path)

        col_count = len(headers)
        row_count = len(rows)
        progress.append(
            info(
                "Creating workbook from data",
                f"{row_count} rows, {col_count} columns",
            )
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        _write_headers(ws, headers)
        _write_rows(ws, rows, start_row=2)
        progress.append(
            ok(
                f"Wrote {row_count} data rows",
                f"headers in row 1, bold",
            )
        )

        wb.save(str(path))
        progress.append(ok(f"Saved {path.name}"))

        if open_after:
            open_file(path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_from_data",
            "output": str(path),
            "output_name": path.name,
            "sheet_name": sheet_name,
            "row_count": row_count,
            "column_count": col_count,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "Ensure headers is a list of strings and rows is a list of lists. "
                "Check that output_path is writable."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_report(
    output_path: str,
    title: str,
    sheets: list[dict[str, Any]],
    open_after: bool = True,
) -> dict[str, Any]:
    """Create a multi-sheet Excel workbook with a Cover sheet."""
    progress: list[dict[str, Any]] = []
    try:
        path = Path(output_path).resolve()
        _ensure_parent(path)

        sheet_count = len(sheets)
        progress.append(
            info(
                "Creating multi-sheet report",
                f"{sheet_count} data sheet(s) + Cover",
            )
        )

        wb = openpyxl.Workbook()

        # Cover sheet — reuse the default active sheet
        cover = wb.active
        if cover is None:
            cover = wb.create_sheet()
        cover.title = "Cover"
        cover["A1"] = title
        cover["A1"].font = Font(bold=True, size=16)
        progress.append(ok("Created Cover sheet", title))

        # Data sheets
        for sheet_def in sheets:
            name = sheet_def.get("name", "Sheet")
            headers = sheet_def.get("headers", [])
            rows = sheet_def.get("rows", [])
            ws = wb.create_sheet(title=name)
            _write_headers(ws, headers)
            _write_rows(ws, rows, start_row=2)
            progress.append(
                ok(
                    f"Created sheet '{name}'",
                    f"{len(rows)} rows, {len(headers)} columns",
                )
            )

        wb.save(str(path))
        progress.append(ok(f"Saved {path.name}", f"{sheet_count + 1} sheets total"))

        if open_after:
            open_file(path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_report",
            "output": str(path),
            "output_name": path.name,
            "title": title,
            "sheets_created": sheet_count + 1,  # includes Cover
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "sheets must be a list of dicts with 'name', 'headers', and 'rows' keys. "
                "Check that output_path is writable."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_from_template(
    template_path: str,
    output_path: str,
    replacements: dict[str, Any],
    open_after: bool = True,
) -> dict[str, Any]:
    """Copy a .xlsx template, replace matching cell values, save to output_path."""
    progress: list[dict[str, Any]] = []
    try:
        src = Path(template_path).resolve()
        if not src.exists():
            progress.append(fail(f"Template not found", str(src)))
            return {
                "success": False,
                "error": f"File not found: {src}",
                "hint": "Check that template_path is an absolute path to an existing .xlsx file.",
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }
        if src.suffix.lower() not in {".xlsx", ".xlsm"}:
            progress.append(fail("Wrong file type", src.suffix))
            return {
                "success": False,
                "error": f"Expected .xlsx file, got {src.suffix}",
                "hint": "template_path must point to a .xlsx or .xlsm file.",
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        dst = Path(output_path).resolve()
        _ensure_parent(dst)

        progress.append(info(f"Loading template", src.name))
        wb = openpyxl.load_workbook(str(src))

        replaced_count = 0
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and cell.value in replacements:
                        cell.value = replacements[cell.value]
                        replaced_count += 1

        progress.append(
            ok(
                f"Replaced {replaced_count} cell value(s)",
                f"{len(replacements)} replacement key(s) searched",
            )
        )

        wb.save(str(dst))
        progress.append(ok(f"Saved {dst.name}"))

        if open_after:
            open_file(dst)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_from_template",
            "template": str(src),
            "output": str(dst),
            "output_name": dst.name,
            "cells_replaced": replaced_count,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "Ensure template_path points to a valid .xlsx file and "
                "output_path is a writable destination."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_from_csv(
    csv_path: str,
    output_path: str,
    sheet_name: str = "Data",
    delimiter: str = ",",
    has_header: bool = True,
    open_after: bool = True,
) -> dict[str, Any]:
    """Import a CSV file into a new Excel workbook."""
    progress: list[dict[str, Any]] = []
    try:
        import csv

        src = Path(csv_path).resolve()
        if not src.exists():
            progress.append(fail("CSV file not found", str(src)))
            return {
                "success": False,
                "error": f"File not found: {csv_path}",
                "hint": "Check that csv_path is an absolute path to an existing CSV file.",
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        out_path = Path(output_path).resolve()
        _ensure_parent(out_path)
        progress.append(info(f"Reading CSV", src.name))

        all_rows: list[list[Any]] = []
        with open(str(src), newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            for row in reader:
                all_rows.append(list(row))

        if len(all_rows) == 0:
            progress.append(warn("CSV file is empty"))

        row_count = len(all_rows) - (1 if has_header else 0)
        col_count = max((len(r) for r in all_rows), default=0)
        progress.append(ok(f"Read {len(all_rows)} row(s)", f"{col_count} column(s)"))

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        for r_idx, row_data in enumerate(all_rows, start=1):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if has_header and r_idx == 1:
                    cell.font = Font(bold=True)

        wb.save(str(out_path))
        progress.append(ok(f"Saved {out_path.name}", f"{row_count} data row(s)"))

        if open_after:
            open_file(out_path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_from_csv",
            "output": str(out_path),
            "output_name": out_path.name,
            "row_count": row_count,
            "column_count": col_count,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "Check that csv_path points to a readable CSV file and "
                "output_path is a writable .xlsx destination."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }


def create_invoice(
    output_path: str,
    company_name: str,
    client_name: str,
    invoice_number: str,
    items: list,
    tax_rate: float = 0.0,
    currency: str = "USD",
    open_after: bool = True,
) -> dict[str, Any]:
    """Create a formatted invoice .xlsx with items, totals, and tax formula."""
    progress: list[dict[str, Any]] = []
    try:
        from openpyxl.styles import PatternFill  # type: ignore[import-untyped]

        if not isinstance(items, list) or len(items) == 0:
            progress.append(fail("items must be a non-empty list"))
            return {
                "success": False,
                "error": "items must be a non-empty list",
                "hint": (
                    'Pass a list like [{"description":"Widget","quantity":2,"unit_price":50.0}].'
                ),
                "progress": progress,
                "token_estimate": _token_estimate(progress),
            }

        out_path = Path(output_path).resolve()
        _ensure_parent(out_path)
        progress.append(info("Creating invoice", out_path.name))

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Invoice"

        # Header block
        ws["A1"] = company_name
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = "INVOICE"
        ws["A3"] = f"Invoice #: {invoice_number}"
        ws["A4"] = f"Client: {client_name}"

        # Column headers in row 6
        header_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        col_headers = ["Description", "Quantity", "Unit Price", "Total"]
        for c_idx, label in enumerate(col_headers, start=1):
            cell = ws.cell(row=6, column=c_idx, value=label)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        progress.append(ok("Written invoice header"))

        # Item rows starting at row 7
        item_start_row = 7
        subtotal = 0.0
        for i, item in enumerate(items):
            row_num = item_start_row + i
            description = item.get("description", "") if isinstance(item, dict) else str(item)
            quantity = item.get("quantity", 0) if isinstance(item, dict) else 0
            unit_price = item.get("unit_price", 0.0) if isinstance(item, dict) else 0.0
            ws.cell(row=row_num, column=1, value=description)
            ws.cell(row=row_num, column=2, value=quantity)
            ws.cell(row=row_num, column=3, value=unit_price)
            # Total formula: =B{row}*C{row}
            ws.cell(row=row_num, column=4, value=f"=B{row_num}*C{row_num}")
            subtotal += float(quantity) * float(unit_price)

        last_item_row = item_start_row + len(items) - 1
        progress.append(ok(f"Written {len(items)} item row(s)"))

        # Subtotal row
        subtotal_row = last_item_row + 2
        ws.cell(row=subtotal_row, column=3, value="Subtotal")
        ws.cell(row=subtotal_row, column=3).font = Font(bold=True)
        ws.cell(
            row=subtotal_row,
            column=4,
            value=f"=SUM(D{item_start_row}:D{last_item_row})",
        )

        current_row = subtotal_row

        # Tax row (only if tax_rate > 0)
        if tax_rate > 0:
            tax_row = subtotal_row + 1
            ws.cell(row=tax_row, column=3, value=f"Tax ({tax_rate * 100:.1f}%)")
            ws.cell(row=tax_row, column=4, value=f"=D{subtotal_row}*{tax_rate}")
            current_row = tax_row

        # Total row
        total_row = current_row + 1
        ws.cell(row=total_row, column=3, value=f"Total ({currency})")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        if tax_rate > 0:
            ws.cell(
                row=total_row,
                column=4,
                value=f"=D{subtotal_row}+D{current_row}",
            )
        else:
            ws.cell(row=total_row, column=4, value=f"=D{subtotal_row}")
        ws.cell(row=total_row, column=4).font = Font(bold=True)

        progress.append(ok("Written subtotal, tax, and total rows"))

        wb.save(str(out_path))
        progress.append(ok(f"Saved {out_path.name}"))

        if open_after:
            open_file(out_path)
            progress.append(ok("Opened in default application"))

        result: dict[str, Any] = {
            "success": True,
            "op": "create_invoice",
            "output": str(out_path),
            "output_name": out_path.name,
            "item_count": len(items),
            "subtotal": round(subtotal, 2),
            "currency": currency,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result

    except Exception as exc:
        progress.append(fail(str(exc)))
        return {
            "success": False,
            "error": str(exc),
            "hint": (
                "Ensure items is a list of dicts with 'description', 'quantity', "
                "and 'unit_price' keys. Check that output_path is writable."
            ),
            "progress": progress,
            "token_estimate": _token_estimate(progress),
        }
