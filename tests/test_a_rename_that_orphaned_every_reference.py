"""rename_sheet renamed the tab and left everything pointing at the old name.

    rename_sheet(book.xlsx, "Data", "Renamed")  -> {"success": true}

    chart refs:   ["'Data'!B1", "'Data'!$A$2:$A$4", "'Data'!$B$2:$B$4"]
    defined name: "'Data'!$B$2:$B$3"
    Summary!A1:   =Data!B2
    sheets:       ['Renamed']

Three kinds of reference, every one of them naming a sheet that no longer
exists, and the tool reported a bare success. Excel draws that chart empty and
shows #REF! in the formula. The implementation was `wb[old].title = new` and
save -- openpyxl renames the tab and nothing else.

Found by reading xlsx_charts/engine.py, where chart data ranges are built as
`Reference(ws, range_string=f"'{ws.title}'!{cell_range}")`, and asking what
happens when a *different server* renames that sheet. Reproduced before it was
believed.

The trap in fixing it is over-matching: renaming "Data" must not touch
"Data2!A1", and a name needing quotes ('Q3 Revenue') appears in two forms
depending on where it is written. Both are covered below, because a fix that
corrupts an unrelated sheet's formulas would be worse than the defect.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from xlsx_basic.helpers import rename_sheet  # type: ignore[reportMissingImports]
from xlsx_charts.engine import add_chart  # type: ignore[reportMissingImports]


def _chart_refs(path: Path) -> list[str]:
    """Reference strings straight out of the chart part, not via openpyxl."""
    refs: list[str] = []
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if "charts/chart" in name:
                refs += re.findall(r"<f>([^<]*)</f>", z.read(name).decode("utf-8", "replace"))
    return refs


@pytest.fixture()
def book(tmp_path: Path) -> Path:
    """A sheet referenced by a chart, a defined name, and another sheet."""
    p = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["month", "value"])
    for row in (("Jan", 1), ("Feb", 2), ("Mar", 3)):
        ws.append(list(row))
    decoy = wb.create_sheet("Data2")
    decoy["A1"] = "=Data2!B9"
    summary = wb.create_sheet("Summary")
    summary["A1"] = "=Data!B2"
    wb.defined_names.add(DefinedName("MyRange", attr_text="'Data'!$B$2:$B$3"))
    wb.save(str(p))
    wb.close()
    assert add_chart(str(p), "Data", "bar", "A1:B4", title="T", anchor_cell="D2", open_after=False)["success"]
    return p


class TestEveryReferenceFollowsTheRename:
    def test_the_chart_points_at_the_new_name(self, book: Path) -> None:
        assert rename_sheet(str(book), "Data", "Renamed", open_after=False)["success"] is True
        refs = _chart_refs(book)
        assert refs, "the chart lost its references entirely"
        assert not [r for r in refs if "Data" in r and "Data2" not in r], refs
        assert all("Renamed" in r for r in refs), refs

    def test_the_defined_name_follows(self, book: Path) -> None:
        rename_sheet(str(book), "Data", "Renamed", open_after=False)
        wb = openpyxl.load_workbook(str(book))
        texts = [d.attr_text for d in wb.defined_names.values()]
        wb.close()
        assert texts == ["Renamed!$B$2:$B$3"], texts

    def test_the_cross_sheet_formula_follows(self, book: Path) -> None:
        rename_sheet(str(book), "Data", "Renamed", open_after=False)
        wb = openpyxl.load_workbook(str(book))
        value = wb["Summary"]["A1"].value
        wb.close()
        assert value == "=Renamed!B2"

    def test_the_reply_says_what_else_moved(self, book: Path) -> None:
        r = rename_sheet(str(book), "Data", "Renamed", open_after=False)
        assert r["references_updated"] == {"formulas": 1, "defined_names": 1, "chart_series": 3}


class TestASheetWhoseNameMerelyStartsTheSame:
    """Renaming "Data" must leave "Data2" alone -- over-matching is worse."""

    def test_the_other_sheet_keeps_its_formula(self, book: Path) -> None:
        rename_sheet(str(book), "Data", "Renamed", open_after=False)
        wb = openpyxl.load_workbook(str(book))
        value = wb["Data2"]["A1"].value
        wb.close()
        assert value == "=Data2!B9", "a sheet sharing the prefix was rewritten"

    def test_that_sheet_still_exists(self, book: Path) -> None:
        rename_sheet(str(book), "Data", "Renamed", open_after=False)
        wb = openpyxl.load_workbook(str(book))
        names = wb.sheetnames
        wb.close()
        assert names == ["Renamed", "Data2", "Summary"]


class TestNamesThatNeedQuoting:
    def test_a_name_with_a_space_is_quoted_in_formulas(self, tmp_path: Path) -> None:
        p = tmp_path / "spaced.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        wb.active["A1"] = 1
        s = wb.create_sheet("Summary")
        s["A1"] = "=Data!A1"
        wb.save(str(p))
        wb.close()
        assert rename_sheet(str(p), "Data", "Q3 Revenue", open_after=False)["success"] is True
        wb = openpyxl.load_workbook(str(p))
        value = wb["Summary"]["A1"].value
        wb.close()
        assert value == "='Q3 Revenue'!A1", value

    def test_a_quoted_source_name_is_matched(self, tmp_path: Path) -> None:
        p = tmp_path / "wasquoted.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Old Name"
        wb.active["A1"] = 1
        s = wb.create_sheet("Summary")
        s["A1"] = "='Old Name'!A1"
        wb.save(str(p))
        wb.close()
        assert rename_sheet(str(p), "Old Name", "New", open_after=False)["success"] is True
        wb = openpyxl.load_workbook(str(p))
        value = wb["Summary"]["A1"].value
        wb.close()
        assert value == "=New!A1", value


class TestARenameWithNothingToFollow:
    def test_it_reports_zero_rather_than_failing(self, tmp_path: Path) -> None:
        p = tmp_path / "plain.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.active["A1"] = "x"
        wb.save(str(p))
        wb.close()
        r = rename_sheet(str(p), "Sheet1", "Renamed", open_after=False)
        assert r["success"] is True
        assert r["references_updated"] == {"formulas": 0, "defined_names": 0, "chart_series": 0}
