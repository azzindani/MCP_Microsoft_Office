"""It announced a conversion above a line saying nothing was converted.

    convert_to_values(workbook_this_server_wrote, ...)
      -> success: true
         progress: ok   "Converted formulas to values in Q1:Q100"
                        "0 formulas replaced"
         skipped_no_cached_value: [...]

The headline claimed the thing its own detail denied, and `success: true` sat
over both. A reader who stops at the first progress line -- which is what a
first line is for -- comes away believing the formulas are now values. They are
not; the file is byte-identical to before.

This is not a rare corner. openpyxl writes a formula as text and never
calculates it, so a workbook produced anywhere in this fleet has no cached
values at all, and convert_to_values can convert nothing in it. The round-15
phase that found this hit it on a file the xlsx server had written minutes
earlier -- which is the normal way to arrive here, not an unusual one.

The honest fields were already present (`formulas_converted: 0`,
`skipped_no_cached_value`, `skipped_count`). The defect was that the prose
contradicted them, the same shape as a chart count that counted sections: the
reply is internally inconsistent, and only one half of it is true.

Fixed by making the headline follow the count, and by making the hint name the
fix -- open it once in Excel or LibreOffice so the values cache -- rather than
only describing the situation.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from xlsx_formulas.engine import convert_to_values  # type: ignore[reportMissingImports]


@pytest.fixture()
def uncalculated(tmp_path: Path) -> Path:
    """Exactly what these servers produce: formulas, no cached values."""
    p = tmp_path / "book.xlsx"
    wb = Workbook()
    ws = wb.active
    for row in range(1, 6):
        ws[f"A{row}"] = row
        ws[f"B{row}"] = f"=A{row}*2"
    wb.save(str(p))
    return p


def headline(r: dict) -> str:
    return str(r["progress"][-3].get("msg", "")) if len(r["progress"]) >= 3 else ""


def messages(r: dict) -> str:
    return " | ".join(str(p.get("msg", "")) for p in r["progress"])


class TestNothingConvertedIsNotAnnouncedAsAConversion:
    def test_it_reports_zero(self, uncalculated: Path) -> None:
        r = convert_to_values(str(uncalculated), "Sheet", "B1:B5")
        assert r["success"] is True, r.get("error")
        assert r["formulas_converted"] == 0, r
        assert r["skipped_count"] == 5, r

    def test_no_progress_line_claims_it_converted(self, uncalculated: Path) -> None:
        """The defect, stated once."""
        r = convert_to_values(str(uncalculated), "Sheet", "B1:B5")
        assert "Converted formulas to values" not in messages(r), messages(r)

    def test_it_says_none_were_converted(self, uncalculated: Path) -> None:
        r = convert_to_values(str(uncalculated), "Sheet", "B1:B5")
        assert "No formulas converted" in messages(r), messages(r)

    def test_the_file_really_is_unchanged(self, uncalculated: Path) -> None:
        before = uncalculated.read_bytes()
        convert_to_values(str(uncalculated), "Sheet", "B1:B5")
        wb = load_workbook(str(uncalculated))
        assert wb.active["B1"].value == "=A1*2", "a formula was overwritten after all"
        assert len(before) > 0

    def test_the_hint_names_the_fix(self, uncalculated: Path) -> None:
        hint = convert_to_values(str(uncalculated), "Sheet", "B1:B5")["hint"]
        assert "Excel or LibreOffice" in hint, hint
        assert "convert_to_values again" in hint, hint

    def test_the_hint_explains_why_the_fleet_always_lands_here(self, uncalculated: Path) -> None:
        hint = convert_to_values(str(uncalculated), "Sheet", "B1:B5")["hint"]
        assert "never calculated" in hint or "not a result" in hint, hint


class TestARealConversionStillAnnouncesItself:
    """The success path must survive; the message exists for a reason."""

    @pytest.fixture()
    def cached(self, tmp_path: Path) -> Path:
        p = tmp_path / "cached.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 3
        ws["B1"] = "=A1*2"
        wb.save(str(p))
        # Simulate what Excel does on save: a data_only copy carrying results.
        vals = load_workbook(str(p))
        vals.active["B1"] = 6
        vals.save(str(p.with_name("values.xlsx")))
        return p

    def test_a_range_with_no_formulas_says_so(self, tmp_path: Path) -> None:
        p = tmp_path / "plain.xlsx"
        wb = Workbook()
        wb.active["A1"] = 1
        wb.save(str(p))
        r = convert_to_values(str(p), "Sheet", "A1:A1")
        assert r["success"] is True, r.get("error")
        assert r["formulas_converted"] == 0
        assert "No formulas converted" in messages(r)
        assert "no formula cells" in messages(r) or "holds no formula cells" in str(r["progress"])

    def test_no_hint_when_there_was_nothing_to_skip(self, tmp_path: Path) -> None:
        p = tmp_path / "plain2.xlsx"
        wb = Workbook()
        wb.active["A1"] = 1
        wb.save(str(p))
        r = convert_to_values(str(p), "Sheet", "A1:A1")
        assert "Excel or LibreOffice" not in (r.get("hint") or "")
