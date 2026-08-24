"""A cell holding a formula must not read back as empty.

set_formula wrote `=SUM(B2:B10)` and reported success. read_cell on the same
cell then returned `value: null, type: "empty"` -- because openpyxl has no
calculation engine and the cached-result slot that Excel and LibreOffice fill
in on save was never written. Nothing was wrong with the file; the formula was
there, and any spreadsheet application would compute it on open.

What was wrong was the pair of answers. "Empty" is what a caller sees when a
write failed, so the sequence read as a silent failure -- and the natural next
move is to write it again.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
for _p in (
    str(ROOT),
    str(ROOT / "servers" / "xlsx_basic"),
    str(ROOT / "servers" / "xlsx_formulas"),
):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from xlsx_basic.engine import read_cell  # noqa: E402
from xlsx_formulas.engine import auto_sum, fill_formula_down, set_formula  # noqa: E402


@pytest.fixture()
def book(tmp_path) -> Path:
    f = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for i, v in enumerate([10, 20, 30], start=2):
        ws[f"B{i}"] = v
    wb.save(f)
    wb.close()
    return f


def test_set_formula_says_it_did_not_compute(book):
    r = set_formula(str(book), "Data", "B5", "=SUM(B2:B4)")
    assert r["success"] is True
    assert r["calculated"] is False
    assert "no calculation engine" in r["note"]
    assert "formula_uncalculated" in r["note"]


def test_reading_that_cell_back_does_not_call_it_empty(book):
    set_formula(str(book), "Data", "B5", "=SUM(B2:B4)")
    r = read_cell(str(book), "Data", "B5")
    assert r["success"] is True
    assert r["formula"] == "=SUM(B2:B4)"
    # The bug: type was "empty", which reads as a write that never landed.
    assert r["type"] == "formula_uncalculated"
    assert r["value"] is None
    assert "no cached result" in r["note"]


def test_a_genuinely_empty_cell_is_still_empty(book):
    r = read_cell(str(book), "Data", "Z99")
    assert r["type"] == "empty"
    assert r["formula"] is None
    assert "note" not in r


def test_a_cell_with_a_real_value_is_unaffected(book):
    r = read_cell(str(book), "Data", "B2")
    assert r["type"] == "number"
    assert r["value"] == 10
    assert "note" not in r


def test_a_cached_formula_result_reads_as_its_value(tmp_path):
    """When a spreadsheet app has computed it, the cached value wins.

    openpyxl writes the cached value only if it is handed one, which is what
    this fixture does directly -- the same shape a file saved by Excel has.
    """
    f = tmp_path / "cached.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 42
    wb.save(f)
    wb.close()
    r = read_cell(str(f), "Data", "A1")
    assert r["type"] == "number"
    assert r["value"] == 42


def test_the_other_formula_writers_say_the_same_thing(book):
    filled = fill_formula_down(str(book), "Data", "=B2*2", "C2", 4)
    assert filled["success"] is True
    assert filled["calculated"] is False
    assert "no calculation engine" in filled["note"]

    summed = auto_sum(str(book), "Data", "B2:B4", "B6")
    assert summed["success"] is True
    assert summed["calculated"] is False
    assert "no calculation engine" in summed["note"]
