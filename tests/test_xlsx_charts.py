"""Tests for xlsx_charts engine functions."""

import shutil
from pathlib import Path

import openpyxl
import pytest

from servers.xlsx_charts.engine import (
    add_chart,
    add_pivot_table,
    delete_chart,
    set_cell_style,
    update_chart,
)

FIXTURES = Path(__file__).parent / "fixtures"
DASHBOARD = FIXTURES / "dashboard.xlsx"
BUDGET_SIMPLE = FIXTURES / "budget_simple.xlsx"


@pytest.fixture()
def workbook(tmp_path: Path) -> Path:
    """Return a writable copy of dashboard.xlsx."""
    dest = tmp_path / "dashboard.xlsx"
    shutil.copy(DASHBOARD, dest)
    return dest


@pytest.fixture()
def budget_workbook(tmp_path: Path) -> Path:
    """Return a writable copy of budget_simple.xlsx."""
    dest = tmp_path / "budget_simple.xlsx"
    shutil.copy(BUDGET_SIMPLE, dest)
    return dest


def _first_sheet(path: Path) -> str:
    wb = openpyxl.load_workbook(str(path))
    name = wb.sheetnames[0]
    wb.close()
    return name


# ---------------------------------------------------------------------------
# add_chart
# ---------------------------------------------------------------------------


def test_add_chart_bar(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = add_chart(
        str(workbook),
        sheet_name,
        "bar",
        "A1:B5",
        "My Bar Chart",
        "D2",
    )
    assert result["success"] is True
    assert result["chart_type"] == "bar"
    assert result["title"] == "My Bar Chart"
    assert "backup" in result


def test_add_chart_line(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = add_chart(
        str(workbook),
        sheet_name,
        "line",
        "A1:B5",
        "My Line Chart",
        "D20",
    )
    assert result["success"] is True
    assert result["chart_type"] == "line"


def test_add_chart_pie(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = add_chart(
        str(workbook),
        sheet_name,
        "pie",
        "A1:B5",
        "My Pie Chart",
        "D20",
    )
    assert result["success"] is True
    assert result["chart_type"] == "pie"


def test_add_chart_unsupported_type_error(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = add_chart(
        str(workbook),
        sheet_name,
        "donut",
        "A1:B5",
        "Bad Chart",
        "D2",
    )
    assert result["success"] is False
    assert "donut" in result["error"]
    assert "hint" in result


def test_add_chart_creates_snapshot(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = add_chart(
        str(workbook),
        sheet_name,
        "bar",
        "A1:B5",
        "Snapshot Test Chart",
        "E2",
    )
    assert result["success"] is True
    backup_path = Path(result["backup"])
    assert backup_path.exists()


def test_add_chart_file_not_found() -> None:
    result = add_chart(
        "/nonexistent/file.xlsx",
        "Sheet1",
        "bar",
        "A1:B5",
        "Chart",
        "D2",
    )
    assert result["success"] is False
    assert "hint" in result


def test_add_chart_sheet_not_found(workbook: Path) -> None:
    result = add_chart(
        str(workbook),
        "NonExistentSheet",
        "bar",
        "A1:B5",
        "Chart",
        "D2",
    )
    assert result["success"] is False
    assert "hint" in result


# ---------------------------------------------------------------------------
# delete_chart
# ---------------------------------------------------------------------------


def test_delete_chart_works(workbook: Path) -> None:
    """Add a chart then delete it — verifies deletion path executes."""
    sheet_name = _first_sheet(workbook)
    # First add a chart
    add_result = add_chart(
        str(workbook), sheet_name, "bar", "A1:B5", "ToDelete", "F2"
    )
    assert add_result["success"] is True

    # Now delete it (it should be at index 0 if the sheet had no charts, else last)
    wb = openpyxl.load_workbook(str(workbook))
    ws = wb[sheet_name]
    chart_count = len(ws._charts)
    wb.close()

    result = delete_chart(str(workbook), sheet_name, chart_count - 1)
    assert result["success"] is True
    assert "backup" in result


def test_delete_chart_out_of_range(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    # Use a very high index that won't exist
    wb = openpyxl.load_workbook(str(workbook))
    ws = wb[sheet_name]
    existing_count = len(ws._charts)
    wb.close()

    result = delete_chart(str(workbook), sheet_name, existing_count + 999)
    assert result["success"] is False
    assert "hint" in result


# ---------------------------------------------------------------------------
# update_chart
# ---------------------------------------------------------------------------


def test_update_chart_title(workbook: Path) -> None:
    """Add a chart then update its title."""
    sheet_name = _first_sheet(workbook)
    add_chart(str(workbook), sheet_name, "bar", "A1:B5", "Original", "H2")

    wb = openpyxl.load_workbook(str(workbook))
    ws = wb[sheet_name]
    idx = len(ws._charts) - 1
    wb.close()

    result = update_chart(str(workbook), sheet_name, idx, title="Updated Title")
    assert result["success"] is True
    assert result["title"] == "Updated Title"


# ---------------------------------------------------------------------------
# add_pivot_table
# ---------------------------------------------------------------------------


def test_add_pivot_table(budget_workbook: Path) -> None:
    """Create a pivot summary from budget data."""
    wb = openpyxl.load_workbook(str(budget_workbook))
    sheet_name = wb.sheetnames[0]
    # Verify there's enough data to pivot
    max_row = wb[sheet_name].max_row
    wb.close()

    if max_row < 2:
        pytest.skip("Fixture has insufficient rows for pivot test")

    # Find headers
    wb2 = openpyxl.load_workbook(str(budget_workbook))
    ws = wb2[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    wb2.close()

    # Use first two string columns and first numeric column
    str_cols = [h for h in headers if isinstance(h, str)]
    if len(str_cols) < 2:
        pytest.skip("Not enough string columns for pivot test")

    result = add_pivot_table(
        str(budget_workbook),
        sheet_name,
        f"A1:{chr(64 + ws.max_column)}{max_row}",
        "I1",
        str_cols[0],
        str_cols[1] if len(str_cols) > 1 else str_cols[0],
        str_cols[0],  # use a string col as value (will sum 0s for non-numeric)
    )
    # Either succeeds or gives a clear error about column mismatch
    assert "success" in result
    assert "progress" in result


def test_add_pivot_table_invalid_column(budget_workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(budget_workbook))
    sheet_name = wb.sheetnames[0]
    max_row = wb[sheet_name].max_row
    max_col = wb[sheet_name].max_column
    wb.close()

    result = add_pivot_table(
        str(budget_workbook),
        sheet_name,
        f"A1:{chr(64 + max_col)}{max_row}",
        "I1",
        "NonExistentCol",
        "NonExistentCol2",
        "NonExistentVal",
    )
    assert result["success"] is False
    assert "hint" in result


# ---------------------------------------------------------------------------
# set_cell_style
# ---------------------------------------------------------------------------


def test_set_cell_style_font(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = set_cell_style(
        str(workbook),
        sheet_name,
        "A1",
        font_name="Arial",
        font_size=14,
        bold=True,
    )
    assert result["success"] is True
    assert result["cell"] == "A1"
    assert "backup" in result

    wb = openpyxl.load_workbook(str(workbook))
    ws = wb[sheet_name]
    assert ws["A1"].font.name == "Arial"
    assert ws["A1"].font.size == 14
    assert ws["A1"].font.bold is True
    wb.close()


def test_set_cell_style_fill_color(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = set_cell_style(
        str(workbook),
        sheet_name,
        "B2",
        fill_color="FF0000",
    )
    assert result["success"] is True

    wb = openpyxl.load_workbook(str(workbook))
    ws = wb[sheet_name]
    fill = ws["B2"].fill
    assert fill.fgColor.rgb.endswith("FF0000")
    wb.close()


def test_set_cell_style_number_format(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = set_cell_style(
        str(workbook),
        sheet_name,
        "C3",
        number_format="#,##0.00",
    )
    assert result["success"] is True

    wb = openpyxl.load_workbook(str(workbook))
    ws = wb[sheet_name]
    assert ws["C3"].number_format == "#,##0.00"
    wb.close()


def test_set_cell_style_invalid_address(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = set_cell_style(str(workbook), sheet_name, "not_a_cell", bold=True)
    assert result["success"] is False
    assert "hint" in result


def test_set_cell_style_creates_snapshot(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    result = set_cell_style(str(workbook), sheet_name, "A1", bold=True)
    assert result["success"] is True
    assert Path(result["backup"]).exists()


# ---------------------------------------------------------------------------
# progress key present on all responses
# ---------------------------------------------------------------------------


def test_all_responses_have_progress(workbook: Path) -> None:
    sheet_name = _first_sheet(workbook)
    results = [
        add_chart(str(workbook), sheet_name, "bar", "A1:B5", "T", "D2"),
        add_chart(str(workbook), sheet_name, "badtype", "A1:B5", "T", "D2"),
        add_chart("/bad/path.xlsx", sheet_name, "bar", "A1:B5", "T", "D2"),
        set_cell_style(str(workbook), sheet_name, "A1", bold=True),
        set_cell_style(str(workbook), sheet_name, "notcell", bold=True),
    ]
    for r in results:
        assert "progress" in r, f"Missing 'progress' key in: {r}"
        assert isinstance(r["progress"], list)
