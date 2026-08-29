"""It said to restore a snapshot. Nothing had been written.

    set_cell(cell_address="A0")
      -> error: Row numbers must be between 1 and 1048576. Row number supplied was 0
         hint : Use restore_version to undo if a snapshot was taken.

The call never reached the save. openpyxl rejects the coordinate while the
arguments are still being validated, so the workbook on disk is untouched --
and the hint answers a typo with a destructive rollback that would discard
whatever legitimate edits the caller had made since their last snapshot.

Found by round 18, whose axis was: make each tool fail the way a careful caller
plausibly would, then do EXACTLY what the hint says. That is what a model with
nothing else to go on does, and it is what the sweep did -- three times in a
row, on set_cell, set_range and insert_row. Every retry failed, because
restoring a file cannot fix a bad argument. One phase went as far as calling a
DIFFERENT server's restore_version (ml-basic's) on an Office workbook to obey
it.

The cause is the last line of `hint_for_error`, its fallback for anything that
is not a PermissionError, a FileNotFoundError or a `.mcp_versions` path. 70 call
sites across 8 servers route through that function, and the sentence turned up
in three separate reports on two different servers.

Fixing it by exception type rather than by message keeps it a choke point: a
ValueError or TypeError out of these engines is an argument being validated,
which always happens before the save. Everything else keeps the undo advice,
because a partial write really can leave a file needing one.

The ordering matters and is asserted below: the `.mcp_versions` guard raises a
plain ValueError, and a path that never resolved arrives as one too. Both are
checked ahead of the new branch and must keep their own, more specific answers.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
for p in (str(ROOT), str(ROOT / "shared"), str(ROOT / "servers" / "xlsx_basic")):
    if p not in sys.path:
        sys.path.insert(0, p)

from shared.file_utils import hint_for_error  # noqa: E402
from xlsx_basic import engine  # noqa: E402


@pytest.fixture
def book(tmp_path) -> str:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "keep me"
    p = tmp_path / "book.xlsx"
    wb.save(str(p))
    return str(p)


class TestTheHintNoLongerTellsYouToUndoNothing:
    def test_a_bad_coordinate_is_not_answered_with_a_restore(self, book):
        r = engine.set_cell(book, "Sheet", "A0", "x")
        assert r["success"] is False
        assert "restore_version" not in r["hint"], r["hint"]

    def test_it_says_nothing_was_written(self, book):
        r = engine.set_cell(book, "Sheet", "A0", "x")
        assert "nothing was written" in r["hint"].lower(), r["hint"]

    def test_it_points_at_the_argument(self, book):
        r = engine.set_cell(book, "Sheet", "A0", "x")
        assert "argument" in r["hint"].lower(), r["hint"]

    def test_the_error_still_names_the_offending_value(self, book):
        # The hint says what to do; the error must still say what was wrong,
        # or "fix the value named in the error" names nothing.
        r = engine.set_cell(book, "Sheet", "A0", "x")
        assert "0" in r["error"]

    def test_the_file_really_was_untouched(self, book):
        import openpyxl

        engine.set_cell(book, "Sheet", "A0", "x")
        assert openpyxl.load_workbook(book)["Sheet"]["A1"].value == "keep me"

    def test_set_range_gets_the_same_answer(self, book):
        r = engine.set_range(book, "Sheet", "A0", [["x"]])
        assert r["success"] is False
        assert "restore_version" not in r["hint"], r["hint"]

    def test_insert_row_gets_the_same_answer(self, book):
        r = engine.insert_row(book, "Sheet", -1)
        assert r["success"] is False
        assert "restore_version" not in r["hint"], r["hint"]

    def test_a_valid_call_still_writes(self, book):
        import openpyxl

        r = engine.set_cell(book, "Sheet", "B2", "hello")
        assert r["success"] is True
        assert openpyxl.load_workbook(book)["Sheet"]["B2"].value == "hello"


class TestTheOtherBranchesAreUnchanged:
    """The fix must not swallow the cases the fallback exists for."""

    def test_an_os_error_still_offers_the_undo(self, tmp_path):
        # A failure that is NOT argument validation can genuinely leave a
        # half-written file, and there the undo advice is the right answer.
        assert "restore_version" in hint_for_error(OSError("disk went away"), tmp_path / "f.xlsx")

    def test_a_snapshot_path_keeps_the_timestamp_route(self, tmp_path):
        # This one is a ValueError too, and it must NOT fall into the new
        # branch -- it has a better answer of its own.
        e = ValueError("Path '/x/.mcp_versions/y.bak' is inside .mcp_versions/.")
        assert "get_history" in hint_for_error(e, tmp_path / "f.xlsx")

    def test_an_unresolvable_path_still_says_so(self):
        # Also a ValueError, also checked earlier, also keeps its own answer.
        e = ValueError("Cannot resolve workspace alias 'workspace:nope'")
        assert "could not be resolved" in hint_for_error(e, None)

    def test_permission_error_is_untouched(self, tmp_path):
        p = tmp_path / "f.xlsx"
        p.write_text("x")
        assert "restore_version" not in hint_for_error(PermissionError(13, "denied"), p)

    def test_file_not_found_is_untouched(self, tmp_path):
        assert "does not exist" in hint_for_error(FileNotFoundError(2, "nope"), tmp_path / "f.xlsx")

    def test_a_type_error_is_treated_as_an_argument_error(self, tmp_path):
        h = hint_for_error(TypeError("'<' not supported between instances of 'str' and 'int'"), tmp_path / "f.xlsx")
        assert "restore_version" not in h
        assert "nothing was written" in h.lower()
