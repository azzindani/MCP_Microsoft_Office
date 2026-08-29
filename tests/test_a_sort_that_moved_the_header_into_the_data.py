"""Round 19b: seven ways an Office tool answered wrongly about what it did.

**sort_sheet sorted the header into the body.** `has_header` skipped
`all_rows[0]` -- the *physical* first row, not the header -- and the write-back
started at a fixed row 2. On a sheet whose row 1 is blank, which is what
`insert_row(1)` leaves behind and what any workbook with a title or spacer row
already looks like, the real header was treated as data:

    1  ⌀      ⌀            1  ⌀      ⌀
    2  name   qty   sort   2  alpha  3
    3  beta   2     ---->  3  beta   2
    4  alpha  3            4  gamma  1
    5  gamma  1            5  name   qty     <- the header, sorted as a value

`success: true`, `rows_sorted: 4`, no warning anywhere. Two ordinary calls in
sequence -- insert a row, sort -- corrupted the file silently. The same function
already carried a long comment about a *different* silent-corruption bug fixed
in an earlier round; its sibling was left in place.

**sort_sheet leaked raw exceptions.** `column="qty"` -- the header name, which
the docstring's `column='A'` invites -- is a *valid* column string resolving to
index 12347, so the guard passed and indexing every row raised
`list index out of range`. A column holding a header string plus numbers raised
`'<' not supported between instances of 'int' and 'str'`. Neither named an
argument.

**The round-18 hint promised what the error could not deliver.** Its sentence
ends "Fix the value named in the error and call again." Three errors this round
named nothing: the two above plus
`invalid literal for int() with base 16: 're'` from `color_hex="red"`.

**hint_for_error was too narrow.** Gated on `ValueError, TypeError`, it let
`IndexError` and PIL's `UnidentifiedImageError` fall through to
"Use restore_version to undo if a snapshot was taken." -- the exact advice
round 18 existed to remove, reached again on two different servers.

**Every failed argument call still wrote a snapshot** and then contradicted
itself: `hint` said there was nothing to restore while `backup` named a file and
`progress` said "✔ Snapshot saved". `discard_unused_snapshot` had answered this
since round 15 and no production caller ever used it.

**add_sheet hinted at a tool that does not exist** -- "delete the existing sheet
first"; there is no `delete_sheet` anywhere in the Office fleet.

**A heap address reached a client.**
`cannot identify image file <_io.BytesIO object at 0x7170edaa6a70>`.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
for p in (
    str(ROOT),
    str(ROOT / "shared"),
    str(ROOT / "servers" / "xlsx_basic"),
    str(ROOT / "servers" / "pptx_design"),
):
    if p not in sys.path:
        sys.path.insert(0, p)

from shared.file_utils import (  # noqa: E402
    drop_snapshot_if_unwritten,
    hint_for_error,
    image_problem,
    scrub_repr,
)
from shared.version_control import snapshot  # noqa: E402
from xlsx_basic import engine as basic_engine  # noqa: E402
from xlsx_basic import helpers as basic_helpers  # noqa: E402

RESTORE_ADVICE = "restore_version"
NOTHING_WRITTEN = "Nothing was written"


def _book(path: Path, rows: list[list[object]], *, leading_blank: bool = False) -> Path:
    wb = Workbook()
    ws = wb.active
    if leading_blank:
        ws.append([None, None])
    for row in rows:
        ws.append(row)
    wb.save(str(path))
    return path


def _values(path: Path) -> list[list[object]]:
    ws = load_workbook(str(path)).active
    return [[c.value for c in row] for row in ws.iter_rows()]


class TestTheHeaderStaysWhereItIs:
    def test_a_leading_blank_row_does_not_promote_the_header_to_data(self, tmp_path):
        book = _book(
            tmp_path / "b.xlsx",
            [["name", "qty"], ["beta", 2], ["alpha", 3], ["gamma", 1]],
            leading_blank=True,
        )
        r = basic_helpers.sort_sheet(str(book), "Sheet", "A")
        assert r["success"] is True
        rows = _values(book)
        assert rows[0] == [None, None], "the blank row moved"
        assert rows[1] == ["name", "qty"], f"header was sorted into the data: {rows}"
        assert [r[0] for r in rows[2:]] == ["alpha", "beta", "gamma"]

    def test_the_ordinary_two_call_sequence_that_corrupted_the_file(self, tmp_path):
        # insert_row(1) then sort_sheet -- both succeed, both are reasonable.
        book = _book(tmp_path / "b.xlsx", [["name", "qty"], ["beta", 2], ["alpha", 3]])
        assert basic_engine.insert_row(str(book), "Sheet", 1)["success"] is True
        r = basic_helpers.sort_sheet(str(book), "Sheet", "A")
        assert r["success"] is True
        assert ["name", "qty"] in _values(book), "header lost by insert-then-sort"

    def test_a_sheet_with_no_leading_blank_still_sorts(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["name", "qty"], ["beta", 2], ["alpha", 3]])
        basic_helpers.sort_sheet(str(book), "Sheet", "A")
        rows = _values(book)
        assert rows[0] == ["name", "qty"]
        assert [r[0] for r in rows[1:]] == ["alpha", "beta"]

    def test_has_header_false_still_sorts_every_row(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["b", 2], ["a", 1]])
        basic_helpers.sort_sheet(str(book), "Sheet", "A", has_header=False)
        assert [r[0] for r in _values(book)] == ["a", "b"]

    def test_a_sheet_of_only_blanks_is_not_an_error(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [[None, None], [None, None]], leading_blank=True)
        r = basic_helpers.sort_sheet(str(book), "Sheet", "A")
        assert r["success"] is True and r["rows_sorted"] == 0


class TestTheSortNeverRaisesAtTheCaller:
    def test_a_column_of_mixed_types_sorts_instead_of_raising(self, tmp_path):
        # The header string sits in the sorted range whenever row 1 is blank,
        # which is exactly how '<' int vs str was reached live.
        book = _book(
            tmp_path / "b.xlsx",
            [["name", "qty"], ["beta", 2], ["alpha", "n/a"], ["gamma", 1]],
            leading_blank=True,
        )
        r = basic_helpers.sort_sheet(str(book), "Sheet", "B")
        assert r["success"] is True, r.get("error")
        assert "not supported between instances" not in str(r)

    def test_numbers_sort_before_text_and_stay_in_order(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["h"], [10], ["zz"], [2], ["aa"]])
        basic_helpers.sort_sheet(str(book), "Sheet", "A")
        assert [r[0] for r in _values(book)][1:] == [2, 10, "aa", "zz"]

    def test_a_header_name_passed_as_the_column_is_answered_not_raised(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["name", "qty"], ["beta", 2]])
        r = basic_helpers.sort_sheet(str(book), "Sheet", "qty")
        assert r["success"] is False
        assert "list index out of range" not in r["error"]
        assert "qty" in r["error"], r["error"]
        assert "letter" in r["hint"].lower()
        assert NOTHING_WRITTEN in r["hint"]

    def test_blanks_sink_to_the_bottom_in_both_directions(self, tmp_path):
        rows = [["h", "x"], ["b", "x"], [None, "x"], ["a", "x"]]
        for ascending in (True, False):
            book = _book(tmp_path / f"b{ascending}.xlsx", rows)
            basic_helpers.sort_sheet(str(book), "Sheet", "A", ascending=ascending)
            column = [r[0] for r in _values(book)][1:]
            assert column[-1] is None, f"blank floated up, ascending={ascending}: {column}"
            assert None not in column[:-1]


class TestTheHintCoversEveryArgumentError:
    @pytest.mark.parametrize(
        "exc",
        [
            ValueError("Row numbers must be between 1 and 1048576"),
            TypeError("'<' not supported between instances of 'int' and 'str'"),
            IndexError("list index out of range"),
            KeyError("no such thing"),
        ],
    )
    def test_it_never_recommends_a_restore_for_an_argument_error(self, exc):
        hint = hint_for_error(exc, Path("/tmp/x.xlsx"))
        assert RESTORE_ADVICE not in hint, hint
        assert NOTHING_WRITTEN in hint

    def test_pillow_s_unidentified_image_is_an_argument_error_too(self):
        pil = pytest.importorskip("PIL")
        hint = hint_for_error(pil.UnidentifiedImageError("cannot identify"), Path("/tmp/x.pptx"))
        assert RESTORE_ADVICE not in hint, hint

    def test_a_real_io_failure_keeps_the_restore_route(self):
        # OSError as a whole must NOT be swept in: a disk filling up mid-save
        # leaves a partial write, and there the snapshot is the only good copy.
        assert RESTORE_ADVICE in hint_for_error(OSError("No space left on device"), Path("/tmp/x.xlsx"))

    def test_it_names_the_argument_when_the_call_site_knows_it(self):
        hint = hint_for_error(TypeError("'<' not supported"), Path("/tmp/x.xlsx"), argument="column")
        assert "column" in hint

    def test_it_does_not_claim_the_error_names_a_value_when_asked_to_name_one(self):
        hint = hint_for_error(TypeError("'<' not supported"), Path("/tmp/x.xlsx"), argument="column")
        assert "named in the error" not in hint

    def test_a_permission_error_is_untouched(self, tmp_path):
        f = tmp_path / "x.xlsx"
        f.write_text("x")
        assert NOTHING_WRITTEN not in hint_for_error(PermissionError("denied"), f)


class TestNothingAdvertisesASnapshotThatIsNotThere:
    def test_a_failed_argument_call_leaves_no_backup_behind(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["name"], ["a"]])
        before = book.read_bytes()
        r = basic_helpers.sort_sheet(str(book), "Sheet", "qty")
        assert r["success"] is False
        assert not r.get("backup"), r.get("backup")
        assert book.read_bytes() == before
        versions = tmp_path / ".mcp_versions"
        assert not versions.exists() or not list(versions.glob("*.bak"))

    def test_the_helper_removes_an_unused_snapshot_and_says_so_in_progress(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["a"]])
        bak = snapshot(str(book))
        progress = [{"icon": "✔", "status": "ok", "msg": "Snapshot saved", "message": "Snapshot saved", "detail": "x"}]
        assert drop_snapshot_if_unwritten(bak, book, progress) is None
        assert not Path(bak).exists()
        assert progress[0]["msg"] != "Snapshot saved"
        assert "discarded" in progress[0]["msg"].lower()

    def test_a_snapshot_of_a_file_that_did_change_is_kept(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["a"]])
        bak = snapshot(str(book))
        book.write_bytes(book.read_bytes() + b"changed")
        assert drop_snapshot_if_unwritten(bak, book, []) == bak
        assert Path(bak).exists()

    def test_it_is_a_no_op_without_a_backup(self, tmp_path):
        assert drop_snapshot_if_unwritten(None, tmp_path / "b.xlsx", []) is None


class TestNoResponseCarriesAHeapAddress:
    def test_an_object_repr_is_reduced_to_its_type(self):
        out = scrub_repr(Exception("cannot identify image file <_io.BytesIO object at 0x7170edaa6a70>"))
        assert "0x" not in out
        assert "<BytesIO>" in out

    def test_an_ordinary_message_is_untouched(self):
        assert scrub_repr(ValueError("Sheet 'Data' not found")) == "Sheet 'Data' not found"

    def test_a_png_that_is_not_a_png_is_named_as_such(self, tmp_path):
        fake = tmp_path / "logo.png"
        fake.write_bytes(b"this is not an image")
        problem = image_problem(fake)
        assert problem and "not a readable image" in problem
        assert "0x" not in problem

    def test_a_real_image_passes(self, tmp_path):
        Image = pytest.importorskip("PIL.Image")
        good = tmp_path / "logo.png"
        Image.new("RGB", (4, 4)).save(str(good))
        assert image_problem(good) is None

    def test_a_wrong_extension_is_still_refused(self, tmp_path):
        f = tmp_path / "logo.svg"
        f.write_text("<svg/>")
        assert "Unsupported image format" in (image_problem(f) or "")


class TestEveryHintNamesSomethingThatExists:
    def test_add_sheet_does_not_send_the_caller_to_a_delete_sheet_tool(self, tmp_path):
        book = _book(tmp_path / "b.xlsx", [["a"]])
        r = basic_engine.add_sheet(str(book), "Sheet")
        assert r["success"] is False
        assert "delete_sheet" not in r["hint"]
        assert "delete the existing sheet" not in r["hint"]
        # It must still say what to do instead, with tools that are really here.
        assert "rename_sheet" in r["hint"] or "list_sheets" in r["hint"]


class TestAColourThatIsNotAColour:
    """`color_hex="red"` answered `invalid literal for int() with base 16: 're'`.

    RGBColor leaves the parse to int(), so the message named 're' -- half the
    value the caller passed -- and no argument at all, under a hint telling
    them to fix the value the error named. The check now runs before the
    snapshot, so a colour typo costs the deck nothing.
    """

    @pytest.fixture
    def deck(self, tmp_path):
        pptx = pytest.importorskip("pptx")
        path = tmp_path / "d.pptx"
        prs = pptx.Presentation()
        prs.slides.add_slide(prs.slide_layouts[0])
        prs.save(str(path))
        return path

    def _call(self, deck, **kw):
        from pptx_design import engine as design_engine

        return design_engine.set_font_all_slides(str(deck), **kw)

    def test_a_colour_name_is_answered_not_leaked(self, deck):
        r = self._call(deck, color_hex="red")
        assert r["success"] is False
        assert "base 16" not in r["error"], r["error"]
        assert "color_hex" in r["error"]
        assert NOTHING_WRITTEN in r["hint"]

    def test_the_deck_keeps_no_snapshot_for_it(self, deck):
        self._call(deck, color_hex="red")
        versions = deck.parent / ".mcp_versions"
        assert not versions.exists() or not list(versions.glob("*.bak"))

    def test_a_hash_prefixed_colour_still_works(self, deck):
        r = self._call(deck, color_hex="#FF0000", font_name="Arial")
        assert r["success"] is True, r.get("error")

    def test_a_bare_six_digit_colour_still_works(self, deck):
        assert self._call(deck, color_hex="00FF00")["success"] is True
