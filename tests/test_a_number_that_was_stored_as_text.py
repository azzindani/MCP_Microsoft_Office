"""set_cell could not write a number.

    set_cell(book.xlsx, "Summary", "B1", "16833")
      -> {"success": true, "value": "16833"}

    openpyxl: B1.value == '16833', B1.data_type == 's'

The MCP signature is `value: str` (xlsx_basic/server.py), the engine does a bare
`ws[addr] = value`, and openpyxl does not coerce. So every number written
through this tool lands as text: =SUM() over the column returns 0, a chart
built on it plots nothing, and Excel shows the green "number stored as text"
flag on each cell.

A round-16 phase wrote 16833 into a summary sheet, read it back, and recorded
"B1 stored as string not numeric (set_cell limitation)" -- filed as a footnote,
not a defect, because the round's axis was whether the file holds what the reply
claimed and it does. It holds the digits. It just does not hold a number.

The fix cannot be float(value) on anything that parses. The values people most
need kept as text are exactly the ones that parse: 07030 is a New Jersey ZIP
code, not 7030, and float("1E5") is 100000.0 rather than a part number. So the
rule here is deliberately narrow -- a plain canonical decimal and nothing else:

    accepted   16833   -42   3.14   0   0.5   -0.25
    left text  07030   +42   " 42 "   1E5   2019-10-16   TRUE   ""

Exponent form is refused on purpose even though it is unambiguously numeric to
Python: a spreadsheet full of silently-converted part codes is a worse outcome
than a caller having to send a float.

The response now names what was actually stored, so a caller who sent "07030"
and got text can see that rather than discover it in a broken SUM three steps
later.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from xlsx_basic.engine import set_cell, set_range  # type: ignore[reportMissingImports]


@pytest.fixture()
def book(tmp_path: Path) -> str:
    p = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Summary"
    wb.save(str(p))
    wb.close()
    return str(p)


def cell(path: str, addr: str = "B1"):
    wb = openpyxl.load_workbook(path)
    c = wb["Summary"][addr]
    got = (c.value, c.data_type)
    wb.close()
    return got


class TestADigitStringBecomesANumber:
    """The defect itself: the thing a spreadsheet exists to do."""

    @pytest.mark.parametrize(
        ("sent", "expected"),
        [("16833", 16833), ("-42", -42), ("0", 0), ("3.14", 3.14), ("-0.25", -0.25), ("0.5", 0.5)],
    )
    def test_it_is_stored_as_a_number(self, book: str, sent: str, expected: float) -> None:
        r = set_cell(book, "Summary", "B1", sent, open_after=False)
        assert r["success"] is True, r.get("error")
        value, dtype = cell(book)
        assert dtype == "n", f"{sent!r} was stored as text -- SUM over this column returns 0"
        assert value == expected

    def test_a_column_of_them_is_summable(self, book: str) -> None:
        """The consequence, not the mechanism."""
        for i, v in enumerate(["10", "20", "30"], start=1):
            set_cell(book, "Summary", f"A{i}", v, open_after=False)
        wb = openpyxl.load_workbook(book)
        ws = wb["Summary"]
        values = [ws.cell(row=i, column=1).value for i in range(1, 4)]
        wb.close()
        assert all(isinstance(v, (int, float)) for v in values), values
        assert sum(values) == 60


class TestTextThatMerelyLooksNumericSurvives:
    """Coercing these is a worse bug than the one being fixed."""

    @pytest.mark.parametrize(
        "sent",
        ["07030", "+42", " 42 ", "1E5", "2019-10-16", "TRUE", "false", "", "12.3.4", "1,234", "--5", "3."],
    )
    def test_it_is_left_as_text(self, book: str, sent: str) -> None:
        r = set_cell(book, "Summary", "B1", sent, open_after=False)
        assert r["success"] is True, r.get("error")
        value, dtype = cell(book)
        if sent == "":
            assert value in ("", None)
            return
        assert dtype == "s", f"{sent!r} was converted to a number -- that destroys the value"
        assert value == sent


class TestTheReplySaysWhatItStored:
    """A caller who sent '07030' should not learn it stayed text from a broken SUM."""

    def test_a_number_reports_number(self, book: str) -> None:
        r = set_cell(book, "Summary", "B1", "16833", open_after=False)
        assert r["stored_type"] == "number"

    def test_text_reports_text(self, book: str) -> None:
        r = set_cell(book, "Summary", "B1", "07030", open_after=False)
        assert r["stored_type"] == "text"

    def test_the_reported_value_is_the_stored_value(self, book: str) -> None:
        """Echoing the input string while storing a number would be its own lie."""
        r = set_cell(book, "Summary", "B1", "16833", open_after=False)
        assert r["value"] == 16833


class TestNonStringsAreUntouched:
    """The engine is also called directly, where a real int already arrives."""

    def test_an_int_stays_an_int(self, book: str) -> None:
        set_cell(book, "Summary", "B1", 16833, open_after=False)
        assert cell(book) == (16833, "n")

    def test_a_float_stays_a_float(self, book: str) -> None:
        set_cell(book, "Summary", "B1", 3.14, open_after=False)
        assert cell(book) == (3.14, "n")

    def test_a_bool_is_not_turned_into_a_number(self, book: str) -> None:
        set_cell(book, "Summary", "B1", True, open_after=False)
        value, _ = cell(book)
        assert value is True


class TestSetRangeAgreesWithSetCell:
    """Two tools writing cells must not disagree about what a digit string is."""

    def test_digit_strings_become_numbers(self, book: str) -> None:
        r = set_range(book, "Summary", "A1", [["10", "20"], ["30", "40"]], open_after=False)
        assert r["success"] is True, r.get("error")
        wb = openpyxl.load_workbook(book)
        ws = wb["Summary"]
        types = [ws.cell(row=rr, column=cc).data_type for rr in (1, 2) for cc in (1, 2)]
        values = [ws.cell(row=rr, column=cc).value for rr in (1, 2) for cc in (1, 2)]
        wb.close()
        # Assert the types before summing: while these are text, sum() raises a
        # TypeError that buries the actual finding under an arithmetic error.
        assert types == ["n"] * 4, types
        assert sum(values) == 100

    def test_a_zip_code_survives_here_too(self, book: str) -> None:
        set_range(book, "Summary", "A1", [["07030"]], open_after=False)
        assert cell(book, "A1") == ("07030", "s")
