"""Office had three template tools and three different ideas of what a key is.

Round 7 fixed the docx and pptx template fills, where one key ate another:
`platform` matched inside `{platform_spend}` and destroyed it. The xlsx fill was
left alone because it is written differently -- it compared each cell's *whole*
value against the substitutions dict, so no key could consume another.

That immunity came at a price nothing reported. Whole-value equality means:

    template cell B1 = "{platform}"
    create_from_template(..., {"platform": "Google Ads"})

    {"success": true, "substitutions_applied": 0}
    B1 still reads "{platform}"

Zero replacements, no warning, no unmatched key, success: true. The caller is
handed a workbook with every placeholder still in it and told it worked. The
docx path resolves that same call and fills the document. A key matching
nothing at all -- a typo, a stale template -- was equally silent, where docx
warns per key.

And a placeholder inside prose was unreachable by construction:

    B3 = "Report for {platform} campaigns"   ->  never substituted, any key

So the shared planner now backs all three tools. What xlsx keeps that the
others cannot is type: a cell that is *nothing but* one placeholder takes the
substitution value as given, so a number stays a number Excel can sum rather
than text that looks like one.

The collision safety is not lost in the trade -- ordered_pairs sorts targets
longest-first and substitute_once replaces in a single pass, which is what made
the docx fix safe. The prefix trap is tested here directly.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "servers" / "xlsx_new"))

from xlsx_new.engine import create_from_template  # noqa: E402


@pytest.fixture()
def template(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Platform"
    ws["B1"] = "{platform}"
    ws["A2"] = "Spend"
    ws["B2"] = "{platform_spend}"
    ws["A3"] = "Summary"
    ws["B3"] = "Report for {platform} campaigns"
    ws["A4"] = "Rows"
    ws["B4"] = 16834
    dst = tmp_path / "template.xlsx"
    wb.save(str(dst))
    wb.close()
    return dst


def fill(template: Path, subs: dict, name: str = "out") -> tuple[dict, dict[str, object]]:
    out = template.parent / f"{name}.xlsx"
    result = create_from_template(str(template), str(out), subs, open_after=False)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws is not None
    cells = {c: ws[c].value for c in ("B1", "B2", "B3", "B4")}
    wb.close()
    return result, cells


BARE = {"platform": "Google Ads", "platform_spend": 1939003.26}
BRACED = {"{platform}": "Google Ads", "{platform_spend}": 1939003.26}


class TestABareKeyReachesADelimitedTemplate:
    """The defect: this exact call replaced nothing and said success."""

    def test_it_replaces_the_placeholder(self, template: Path):
        _, cells = fill(template, BARE)
        assert cells["B1"] == "Google Ads"

    def test_it_counts_what_it_did(self, template: Path):
        result, _ = fill(template, BARE)
        assert result["substitutions_applied"] == 3, result

    def test_no_key_is_reported_unmatched(self, template: Path):
        result, _ = fill(template, BARE)
        assert result["unmatched_keys"] == []

    def test_it_says_which_form_it_matched(self, template: Path):
        result, _ = fill(template, BARE)
        msgs = " ".join(str(p.get("msg", "")) for p in result["progress"])
        assert "{platform}" in msgs


class TestNoKeyEatsAnother:
    """`platform` is a prefix of `platform_spend`; round 7's docx defect."""

    def test_the_longer_placeholder_survives(self, template: Path):
        _, cells = fill(template, BARE)
        assert cells["B2"] == 1939003.26

    def test_it_is_not_corrupted_into_the_shorter_value(self, template: Path):
        _, cells = fill(template, BARE)
        assert cells["B2"] != "{Google Ads_spend}"
        assert "Google Ads_spend" not in str(cells["B2"])

    def test_the_order_the_caller_wrote_them_in_does_not_matter(self, template: Path):
        reversed_order = {"platform_spend": 1939003.26, "platform": "Google Ads"}
        _, cells = fill(template, reversed_order, name="rev")
        assert cells["B1"] == "Google Ads"
        assert cells["B2"] == 1939003.26


class TestAPlaceholderInProseIsReachable:
    def test_it_substitutes_inside_a_sentence(self, template: Path):
        _, cells = fill(template, BARE)
        assert cells["B3"] == "Report for Google Ads campaigns"

    def test_the_surrounding_words_are_kept(self, template: Path):
        _, cells = fill(template, BARE)
        assert str(cells["B3"]).startswith("Report for") and str(cells["B3"]).endswith("campaigns")


class TestTypeSurvivesAWholeCellPlaceholder:
    """What xlsx has that docx and pptx cannot: a cell Excel can still sum."""

    def test_a_number_stays_a_number(self, template: Path):
        _, cells = fill(template, BARE)
        assert isinstance(cells["B2"], float), type(cells["B2"])

    def test_and_is_not_stringified(self, template: Path):
        _, cells = fill(template, BARE)
        assert cells["B2"] == pytest.approx(1939003.26)

    def test_a_number_inside_prose_is_rendered_as_text(self, template: Path):
        """No choice there -- but the sentence must survive it."""
        _, cells = fill(template, {"platform": 42}, name="num")
        assert cells["B3"] == "Report for 42 campaigns"

    def test_a_non_placeholder_cell_is_untouched(self, template: Path):
        _, cells = fill(template, BARE)
        assert cells["B4"] == 16834


class TestBracedKeysStillWork:
    """The form the old whole-value match required must not regress."""

    @pytest.mark.parametrize("cell,expected", [("B1", "Google Ads"), ("B2", 1939003.26)])
    def test_it_fills_them(self, template: Path, cell: str, expected: object):
        _, cells = fill(template, BRACED, name="braced")
        assert cells[cell] == expected

    def test_it_reaches_prose_too(self, template: Path):
        _, cells = fill(template, BRACED, name="braced2")
        assert cells["B3"] == "Report for Google Ads campaigns"

    def test_nothing_is_left_unmatched(self, template: Path):
        result, _ = fill(template, BRACED, name="braced3")
        assert result["unmatched_keys"] == []


class TestAKeyThatMatchesNothingIsReported:
    def test_it_is_listed(self, template: Path):
        result, _ = fill(template, {"totally_absent_key": "X"}, name="absent")
        assert result["unmatched_keys"] == ["totally_absent_key"]

    def test_it_is_warned_about_in_progress(self, template: Path):
        result, _ = fill(template, {"totally_absent_key": "X"}, name="absent2")
        msgs = " ".join(str(p.get("msg", "")) for p in result["progress"])
        assert "totally_absent_key" in msgs

    def test_it_does_not_rewrite_the_template(self, template: Path):
        _, cells = fill(template, {"totally_absent_key": "X"}, name="absent3")
        assert cells["B1"] == "{platform}"
        assert cells["B3"] == "Report for {platform} campaigns"

    def test_the_good_keys_in_the_same_call_still_apply(self, template: Path):
        result, cells = fill(template, {"platform": "Meta", "nope": "X"}, name="mixed")
        assert cells["B1"] == "Meta"
        assert result["unmatched_keys"] == ["nope"]

    def test_success_is_still_true(self, template: Path):
        """A missing key is a warning, not a failed run -- same as docx."""
        result, _ = fill(template, {"nope": "X"}, name="absent4")
        assert result["success"] is True


class TestABareWordIsNotAPlaceholder:
    """Round 7's rule, now enforced on workbooks too.

    A sweep passed rows="16,834" at a {placeholder} template whose body read
    "205 duplicate rows were identified" and got "205 duplicate 16,834 were
    identified" under success: true. Once any key resolves to a delimited
    placeholder the template plainly uses a convention, so a key without one is
    a caller mistake rather than licence to substitute into a sentence.

    A template with no delimiters *anywhere* is the other case -- it has no
    other reading, so bare keys still apply. Both are checked here because the
    xlsx tool now shares the planner that decides between them.
    """

    PROSE = "205 duplicate rows were identified"

    def sheet_value(self, path: Path, cell: str) -> object:
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        assert ws is not None
        value = ws[cell].value
        wb.close()
        return value

    @pytest.fixture()
    def prose_template(self, tmp_path: Path) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Platform: {platform}"
        ws["A2"] = self.PROSE
        dst = tmp_path / "prose.xlsx"
        wb.save(str(dst))
        wb.close()
        return dst

    def test_the_sentence_survives(self, prose_template: Path):
        fill(prose_template, {"platform": "Google Ads", "rows": "16,834"}, name="prose_out")
        assert self.sheet_value(prose_template.parent / "prose_out.xlsx", "A2") == self.PROSE

    def test_the_real_placeholder_is_still_filled(self, prose_template: Path):
        fill(prose_template, {"platform": "Google Ads", "rows": "16,834"}, name="prose_out2")
        assert self.sheet_value(prose_template.parent / "prose_out2.xlsx", "A1") == "Platform: Google Ads"

    def test_the_bare_key_is_reported_unmatched(self, prose_template: Path):
        result, _ = fill(prose_template, {"platform": "X", "rows": "16,834"}, name="prose_out3")
        assert result["unmatched_keys"] == ["rows"]

    def test_a_template_with_no_placeholders_still_takes_bare_keys(self, tmp_path: Path):
        """Matches the docx tool exactly -- a workbook with no delimiters
        anywhere has no other reading."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Platform: PLATFORM"
        tpl = tmp_path / "plain.xlsx"
        wb.save(str(tpl))
        wb.close()
        fill(tpl, {"PLATFORM": "Google Ads"}, name="plain_out")
        assert self.sheet_value(tmp_path / "plain_out.xlsx", "A1") == "Platform: Google Ads"


class TestAllThreeTemplateToolsShareThePlanner:
    def test_no_office_template_engine_hand_rolls_its_own(self):
        root = Path(__file__).parent.parent / "servers"
        offenders = []
        for engine in root.glob("*_new/*/engine.py"):
            text = engine.read_text(encoding="utf-8")
            if "def create_from_template" in text and "resolve_targets" not in text:
                offenders.append(str(engine.relative_to(root)))
        assert not offenders, offenders
