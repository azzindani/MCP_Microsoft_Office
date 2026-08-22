"""One response carried 33,000 cell references.

convert_to_values skips formula cells that have no cached result -- openpyxl
writes formulas without one, so a workbook these tools built and never opened in
Excel has a cached value for none of them. It then listed every skipped address
in full, twice: once joined into a progress warning and once as the
`skipped_no_cached_value` field.

A sweep converting a filled-down CTR column over the 16,834-row ad dataset
skipped all of them:

    formulas_converted: 0
    skipped_no_cached_value: [ ...16,834 addresses... ]

At 3,000 cells that is already a 46,801-character response with a token_estimate
of 11,689; the sweep's 16,834 would be roughly 260 KB and 65,000 tokens, against
the ~10,000-12,000 token context this server is built for. The tool reported
success and the file was correct -- only the size was wrong, which no structural
check looks at.

The count is the information; the addresses are a sample. Capped with
get_max_cells() like every other range read in the repo, with the true total
kept in skipped_count and a truncated flag beside it.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from shared.platform_utils import get_max_cells
from xlsx_formulas.engine import (  # type: ignore[reportMissingImports]
    convert_to_values,
    fill_formula_down,
)
from xlsx_new.engine import create_from_data  # type: ignore[reportMissingImports]

ROWS = 1200


@pytest.fixture()
def filled_workbook(tmp_path: Path) -> str:
    """A column of formulas with no cached values -- the shape openpyxl writes."""
    out = tmp_path / "big.xlsx"
    create_from_data(str(out), "Data", ["a", "b"], [[i, i * 2] for i in range(ROWS)], open_after=False)
    fill_formula_down(str(out), "Data", "=A2*B2", "C2", ROWS + 1)
    return str(out)


@pytest.fixture()
def response(filled_workbook: str) -> dict:
    return convert_to_values(filled_workbook, "Data", f"C2:C{ROWS + 1}")


class TestTheResponseStaysSmall:
    def test_it_succeeds(self, response: dict):
        assert response["success"] is True, response.get("error")

    def test_the_listed_addresses_are_capped(self, response: dict):
        assert len(response["skipped_no_cached_value"]) <= get_max_cells()

    def test_the_true_total_is_still_reported(self, response: dict):
        assert response["skipped_count"] == ROWS

    def test_it_says_the_list_was_cut(self, response: dict):
        assert response["truncated"] is True

    def test_the_whole_response_is_a_readable_size(self, response: dict):
        """65,000 tokens is six times this server's target context."""
        assert response["token_estimate"] < 2000, response["token_estimate"]

    def test_the_response_is_not_hundreds_of_kilobytes(self, response: dict):
        assert len(json.dumps(response, default=str)) < 20000

    def test_the_warning_does_not_repeat_the_whole_list(self, response: dict):
        warns = [p for p in response["progress"] if p.get("status") == "warn"]
        assert warns, response["progress"]
        assert len(str(warns[0].get("detail", ""))) < 4000

    def test_the_warning_still_names_the_count(self, response: dict):
        warns = [p for p in response["progress"] if p.get("status") == "warn"]
        assert str(ROWS) in str(warns[0].get("message", "")), warns


class TestItStillExplainsItself:
    def test_the_hint_says_why_nothing_converted(self, response: dict):
        assert "cached value" in response["hint"], response["hint"]

    def test_the_addresses_it_does_show_are_real(self, response: dict):
        shown = response["skipped_no_cached_value"]
        assert shown[0] == "C2", shown[:3]
        assert all(a.startswith("C") for a in shown), shown[:5]

    def test_the_formulas_are_left_alone(self, filled_workbook: str, response: dict):
        import openpyxl

        ws = openpyxl.load_workbook(filled_workbook).active
        assert ws is not None
        assert str(ws["C2"].value).startswith("="), ws["C2"].value


class TestASmallRangeIsNotTruncated:
    def test_a_short_list_is_returned_whole(self, tmp_path: Path):
        out = tmp_path / "small.xlsx"
        create_from_data(str(out), "Data", ["a", "b"], [[1, 2], [3, 4]], open_after=False)
        fill_formula_down(str(out), "Data", "=A2*B2", "C2", 3)
        r = convert_to_values(str(out), "Data", "C2:C3")
        assert r["success"] is True, r.get("error")
        assert r["truncated"] is False
        assert r["skipped_count"] == len(r["skipped_no_cached_value"])
