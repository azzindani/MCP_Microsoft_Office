"""Tests for xlsx_new engine functions not already covered by test_scenarios.py
(create_invoice and create_from_csv are exercised there)."""

from pathlib import Path

import openpyxl

from xlsx_new.engine import (
    create_from_data,
    create_from_template,
    create_report,
    create_workbook,
)

FIXTURES = Path(__file__).parent / "fixtures"
BUDGET_SIMPLE = FIXTURES / "budget_simple.xlsx"


# ---------------------------------------------------------------------------
# create_workbook
# ---------------------------------------------------------------------------


def test_create_workbook_with_custom_sheet_name(tmp_path: Path) -> None:
    out = tmp_path / "wb.xlsx"
    result = create_workbook(str(out), sheet_name="Data", open_after=False)
    assert result["success"] is True
    assert result["sheet_name"] == "Data"

    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["Data"]


def test_create_workbook_default_sheet_name(tmp_path: Path) -> None:
    out = tmp_path / "wb.xlsx"
    result = create_workbook(str(out), open_after=False)
    assert result["success"] is True
    assert result["sheet_name"] == "Sheet1"


# ---------------------------------------------------------------------------
# create_from_data
# ---------------------------------------------------------------------------


def test_create_from_data_writes_headers_and_rows(tmp_path: Path) -> None:
    out = tmp_path / "wb.xlsx"
    headers = ["Region", "Revenue"]
    rows = [["North", 1000], ["South", 2000]]
    result = create_from_data(str(out), "Sales", headers, rows, open_after=False)
    assert result["success"] is True
    assert result["row_count"] == 2
    assert result["column_count"] == 2

    wb = openpyxl.load_workbook(str(out))
    ws = wb["Sales"]
    assert [c.value for c in ws[1]] == headers
    assert ws["A2"].value == "North"
    assert ws["B3"].value == 2000


def test_create_from_data_headers_are_bold(tmp_path: Path) -> None:
    out = tmp_path / "wb.xlsx"
    create_from_data(str(out), "Sales", ["A", "B"], [[1, 2]], open_after=False)
    wb = openpyxl.load_workbook(str(out))
    assert wb["Sales"]["A1"].font.bold is True


# ---------------------------------------------------------------------------
# create_report
# ---------------------------------------------------------------------------


def test_create_report_builds_cover_plus_data_sheets(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    sheets = [
        {"name": "Q1", "headers": ["Region", "Revenue"], "rows": [["North", 100]]},
        {"name": "Q2", "headers": ["Region", "Revenue"], "rows": [["South", 200]]},
    ]
    result = create_report(str(out), "Quarterly Report", sheets, open_after=False)
    assert result["success"] is True
    assert result["sheets_created"] == 3  # Cover + Q1 + Q2

    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["Cover", "Q1", "Q2"]
    assert wb["Cover"]["A1"].value == "Quarterly Report"
    assert wb["Cover"]["A1"].font.bold is True
    assert wb["Q1"]["A2"].value == "North"


def test_create_report_empty_sheets_list(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    result = create_report(str(out), "Empty Report", [], open_after=False)
    assert result["success"] is True
    assert result["sheets_created"] == 1  # Cover only

    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["Cover"]


# ---------------------------------------------------------------------------
# create_from_template
# ---------------------------------------------------------------------------


def test_create_from_template_replaces_matching_cell_values(tmp_path: Path) -> None:
    out = tmp_path / "filled.xlsx"
    result = create_from_template(str(BUDGET_SIMPLE), str(out), {"North": "N. Region"}, open_after=False)
    assert result["success"] is True
    assert result["cells_replaced"] == 1

    wb = openpyxl.load_workbook(str(out))
    ws = wb["Q3 Revenue"]
    values = [c.value for row in ws.iter_rows() for c in row]
    assert "N. Region" in values
    assert "North" not in values


def test_create_from_template_missing_template(tmp_path: Path) -> None:
    out = tmp_path / "filled.xlsx"
    result = create_from_template(str(tmp_path / "ghost.xlsx"), str(out), {"A": "B"}, open_after=False)
    assert result["success"] is False
    assert "not found" in result["error"].lower()


def test_create_from_template_wrong_file_type(tmp_path: Path) -> None:
    bad = tmp_path / "notes.txt"
    bad.write_text("hello")
    out = tmp_path / "filled.xlsx"
    result = create_from_template(str(bad), str(out), {"A": "B"}, open_after=False)
    assert result["success"] is False


# ---------------------------------------------------------------------------
# Global contract: every tool response has a progress field
# ---------------------------------------------------------------------------


def test_all_responses_have_progress_field(tmp_path: Path) -> None:
    results = [
        create_workbook(str(tmp_path / "a.xlsx"), open_after=False),
        create_from_data(str(tmp_path / "b.xlsx"), "S", ["H"], [[1]], open_after=False),
        create_report(str(tmp_path / "c.xlsx"), "T", [], open_after=False),
        create_from_template(str(BUDGET_SIMPLE), str(tmp_path / "d.xlsx"), {"North": "X"}, open_after=False),
    ]
    for r in results:
        assert "progress" in r, f"Missing 'progress' in: {r}"
        assert isinstance(r["progress"], list)
