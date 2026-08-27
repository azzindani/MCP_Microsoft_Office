"""copy_sheet copied the cells and left the chart and the picture behind.

    copy_sheet(book.xlsx, "S", "S_copy")  -> {"success": true}

    S      _charts: 1  _images: 1
    S_copy _charts: 0  _images: 0

openpyxl's copy_worksheet documents that it copies cells, styles and dimensions
and *not* charts, images or other drawing objects. copy_sheet passed that
omission straight through as a plain success. The cell data was perfect, so
nothing looked wrong until the file was opened -- which is the whole point of
round 16's axis: does the artifact hold what the reply claimed?

The copy's chart is retargeted at the copy, not left plotting the original's
data. That is what Excel does when you duplicate a sheet by hand, and it reuses
the same _retarget_sheet_name() written for the rename_sheet defect
(test_a_rename_that_orphaned_every_reference.py) rather than growing a second
way to rewrite a sheet reference.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage

from xlsx_basic.helpers import copy_sheet  # type: ignore[reportMissingImports]

PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
    "890000000d4944415478da63f8cfc0f00f0004010100b4f4d9cf0000000049454e44ae426082"
)


@pytest.fixture()
def book(tmp_path: Path) -> Path:
    p = tmp_path / "book.xlsx"
    img = tmp_path / "dot.png"
    img.write_bytes(PNG)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["month", "value"])
    for row in (("Jan", 1), ("Feb", 2), ("Mar", 3)):
        ws.append(list(row))
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    ws.add_chart(chart, "E2")
    ws.add_image(XLImage(str(img)), "H2")
    wb.save(str(p))
    wb.close()
    return p


def _refs(path: Path) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if "charts/chart" in name:
                out[name] = re.findall(r"<f>([^<]*)</f>", z.read(name).decode("utf-8", "replace"))
    return out


class TestTheCopyKeepsItsDrawings:
    def test_the_chart_comes_with_it(self, book: Path) -> None:
        assert copy_sheet(str(book), "S", "S_copy", open_after=False)["success"] is True
        wb = openpyxl.load_workbook(str(book))
        charts = len(wb["S_copy"]._charts)
        wb.close()
        assert charts == 1, "the copied sheet has no chart"

    def test_the_image_comes_with_it(self, book: Path) -> None:
        copy_sheet(str(book), "S", "S_copy", open_after=False)
        wb = openpyxl.load_workbook(str(book))
        images = len(wb["S_copy"]._images)
        wb.close()
        assert images == 1, "the copied sheet has no image"

    def test_the_original_still_has_its_own(self, book: Path) -> None:
        copy_sheet(str(book), "S", "S_copy", open_after=False)
        wb = openpyxl.load_workbook(str(book))
        counts = (len(wb["S"]._charts), len(wb["S"]._images))
        wb.close()
        assert counts == (1, 1), "copying moved the drawings instead of duplicating them"

    def test_the_reply_says_what_it_carried_over(self, book: Path) -> None:
        r = copy_sheet(str(book), "S", "S_copy", open_after=False)
        assert r["drawings_copied"] == {"charts": 1, "images": 1}


class TestTheCopiedChartPlotsTheCopy:
    """Excel's behaviour when you duplicate a sheet by hand."""

    def test_the_new_chart_reads_from_the_new_sheet(self, book: Path) -> None:
        copy_sheet(str(book), "S", "S_copy", open_after=False)
        refs = _refs(book)
        assert len(refs) == 2, refs
        copied = [v for k, v in refs.items() if any("S_copy" in f for f in v)]
        assert copied, f"no chart points at the copy: {refs}"
        assert all("S_copy" in f for f in copied[0]), copied

    def test_the_original_chart_is_untouched(self, book: Path) -> None:
        copy_sheet(str(book), "S", "S_copy", open_after=False)
        refs = _refs(book)
        original = [v for v in refs.values() if not any("S_copy" in f for f in v)]
        assert original, f"the original chart was retargeted: {refs}"
        assert all(re.search(r"'?S'?!", f) for f in original[0]), original


class TestASheetWithNoDrawings:
    def test_it_reports_zero_and_still_copies_cells(self, tmp_path: Path) -> None:
        p = tmp_path / "plain.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "S"
        wb.active["A1"] = "hello"
        wb.save(str(p))
        wb.close()
        r = copy_sheet(str(p), "S", "S_copy", open_after=False)
        assert r["success"] is True
        assert r["drawings_copied"] == {"charts": 0, "images": 0}
        wb = openpyxl.load_workbook(str(p))
        value = wb["S_copy"]["A1"].value
        wb.close()
        assert value == "hello"
