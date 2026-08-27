"""find_duplicates counted its results after throwing most of them away.

    find_duplicates(book.xlsx, "Sheet1", "A")
      -> {"duplicate_count": 100, "truncated": true, ...}

on a column holding 5,000 distinct repeated values. The cap itself is fine and
it is disclosed -- `truncated: true` is right there. What was wrong is the
number beside it:

    if len(duplicates) > 100:
        duplicates = duplicates[:100]
        truncated = True
    ...
    "duplicate_count": len(duplicates),      # counted AFTER the slice

so the one field a caller reads to answer "how many duplicates are in this
column?" said 100, and said it precisely when the column was worst. Same shape
as test_a_headline_that_contradicted_its_own_detail.py.

A round-16 phase called find_duplicates, got exactly 100 results back, and
recorded that as correct -- which for its own axis it was, since the file was
never in question. Reading the source is what found it.

Second half of the same defect: the 100 was hardcoded while the per-value row
cap two lines above already used get_max_search_results(). One limit from the
platform helper and one typed in will drift, and MCP_CONSTRAINED_MODE never
reached the typed one.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from shared.platform_utils import get_max_search_results  # type: ignore[reportMissingImports]
from xlsx_basic.helpers import find_duplicates  # type: ignore[reportMissingImports]

CAP = get_max_search_results()


def _book(tmp_path: Path, distinct_dupes: int) -> str:
    """A column with `distinct_dupes` values that each appear twice."""
    p = tmp_path / "dupes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "id"
    row = 2
    for i in range(distinct_dupes):
        for _ in range(2):
            ws.cell(row=row, column=1).value = f"v{i}"
            row += 1
    wb.save(str(p))
    wb.close()
    return str(p)


class TestTheCountSurvivesTheCap:
    def test_it_reports_every_duplicate_it_found(self, tmp_path: Path) -> None:
        total = CAP + 37
        r = find_duplicates(_book(tmp_path, total), "Sheet1", "A")
        assert r["success"] is True, r.get("error")
        assert r["duplicate_count"] == total, "the count was taken after the list was truncated"

    def test_the_list_is_still_capped(self, tmp_path: Path) -> None:
        r = find_duplicates(_book(tmp_path, CAP + 37), "Sheet1", "A")
        assert len(r["duplicates"]) == CAP

    def test_it_says_how_many_it_actually_returned(self, tmp_path: Path) -> None:
        r = find_duplicates(_book(tmp_path, CAP + 37), "Sheet1", "A")
        assert r["duplicates_returned"] == len(r["duplicates"])
        assert r["duplicates_returned"] < r["duplicate_count"]

    def test_truncation_is_still_flagged(self, tmp_path: Path) -> None:
        assert find_duplicates(_book(tmp_path, CAP + 37), "Sheet1", "A")["truncated"] is True


class TestNothingChangesBelowTheCap:
    def test_count_and_list_agree(self, tmp_path: Path) -> None:
        r = find_duplicates(_book(tmp_path, 3), "Sheet1", "A")
        assert r["duplicate_count"] == 3
        assert len(r["duplicates"]) == 3
        assert r["duplicates_returned"] == 3

    def test_not_flagged_as_truncated(self, tmp_path: Path) -> None:
        assert find_duplicates(_book(tmp_path, 3), "Sheet1", "A")["truncated"] is False

    def test_a_column_with_no_repeats_is_empty_not_truncated(self, tmp_path: Path) -> None:
        p = tmp_path / "uniq.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1).value = "id"
        for i in range(20):
            ws.cell(row=i + 2, column=1).value = f"u{i}"
        wb.save(str(p))
        wb.close()
        r = find_duplicates(str(p), "Sheet1", "A")
        assert r["duplicate_count"] == 0
        assert r["duplicates"] == []
        assert r["truncated"] is False


class TestTheCapComesFromThePlatformHelper:
    """A typed-in limit beside a derived one is the second table that drifts."""

    def test_the_returned_cap_matches_the_helper(self, tmp_path: Path) -> None:
        r = find_duplicates(_book(tmp_path, CAP + 5), "Sheet1", "A")
        assert len(r["duplicates"]) == get_max_search_results()

    def test_the_response_names_the_cap_it_used(self, tmp_path: Path) -> None:
        r = find_duplicates(_book(tmp_path, CAP + 5), "Sheet1", "A")
        assert r["max_duplicates_returned"] == get_max_search_results()

    def test_no_hardcoded_hundred_remains(self) -> None:
        """The literal is what constrained mode could not reach."""
        src = Path(__file__).resolve().parents[1] / "servers/xlsx_basic/xlsx_basic/helpers.py"
        body = src.read_text(encoding="utf-8")
        start = body.index("def find_duplicates")
        end = body.index("def copy_sheet", start)
        assert "> 100" not in body[start:end], "find_duplicates still caps on a typed-in 100"


@pytest.mark.parametrize("column", ["A"])
class TestRowListsAreStillCappedPerValue:
    """The pre-existing cap this defect sat next to -- keep it working."""

    def test_rows_truncated_flag(self, tmp_path: Path, column: str) -> None:
        p = tmp_path / "wide.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1).value = "id"
        for i in range(CAP + 20):
            ws.cell(row=i + 2, column=1).value = "same"
        wb.save(str(p))
        wb.close()
        r = find_duplicates(str(p), "Sheet1", column)
        assert r["rows_truncated"] is True
        assert len(r["duplicates"][0]["rows"]) == CAP
        assert r["duplicates"][0]["count"] == CAP + 20
