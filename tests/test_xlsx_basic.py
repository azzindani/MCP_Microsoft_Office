"""Tests for xlsx_basic engine functions."""

import shutil
from pathlib import Path

import openpyxl
import pytest
from xlsx_basic.engine import (
    add_sheet,
    delete_row,
    get_sheet_summary,
    insert_row,
    list_sheets,
    read_cell,
    read_cell_range,
    search_cells,
    set_cell,
    set_range,
)

FIXTURES = Path(__file__).parent / "fixtures"
BUDGET_SIMPLE = FIXTURES / "budget_simple.xlsx"


@pytest.fixture()
def workbook(tmp_path: Path) -> Path:
    """Return a writable copy of budget_simple.xlsx."""
    dest = tmp_path / "budget_simple.xlsx"
    shutil.copy(BUDGET_SIMPLE, dest)
    return dest


# ---------------------------------------------------------------------------
# list_sheets
# ---------------------------------------------------------------------------


def test_list_sheets_returns_names_and_dimensions():
    result = list_sheets(str(BUDGET_SIMPLE))
    assert result["success"] is True
    names = [s["name"] for s in result["sheets"]]
    assert "Q3 Revenue" in names
    assert "Dashboard" in names
    assert result["sheet_count"] == 2


def test_list_sheets_returns_dimensions():
    result = list_sheets(str(BUDGET_SIMPLE))
    q3 = next(s for s in result["sheets"] if s["name"] == "Q3 Revenue")
    assert q3["max_row"] == 5
    assert q3["max_col"] == 6
    assert q3["last_cell"] == "F5"


def test_list_sheets_file_not_found():
    result = list_sheets("/nonexistent/path/file.xlsx")
    assert result["success"] is False
    assert "not found" in result["error"].lower()
    assert "hint" in result


def test_list_sheets_has_progress():
    result = list_sheets(str(BUDGET_SIMPLE))
    assert "progress" in result
    assert isinstance(result["progress"], list)


# ---------------------------------------------------------------------------
# get_sheet_summary
# ---------------------------------------------------------------------------


def test_get_sheet_summary_returns_headers():
    result = get_sheet_summary(str(BUDGET_SIMPLE), "Q3 Revenue")
    assert result["success"] is True
    headers = [h["value"] for h in result["header_row"]]
    assert "Region" in headers
    assert "Jan" in headers


def test_get_sheet_summary_returns_dimensions():
    result = get_sheet_summary(str(BUDGET_SIMPLE), "Q3 Revenue")
    assert result["dimensions"]["rows"] == 5
    assert result["dimensions"]["cols"] == 6
    assert result["dimensions"]["last_cell"] == "F5"


def test_get_sheet_summary_zero_cell_values():
    """Summary must not dump all cell values — only headers and sample."""
    result = get_sheet_summary(str(BUDGET_SIMPLE), "Q3 Revenue")
    assert result["success"] is True
    # Should not have a full 'data' 2D array
    assert "data" not in result
    # first_col_sample should be ≤ 5 entries (plus possible trailing-message string)
    sample = [x for x in result["first_col_sample"] if isinstance(x, dict)]
    assert len(sample) <= 5


def test_get_sheet_summary_first_col_sample():
    result = get_sheet_summary(str(BUDGET_SIMPLE), "Q3 Revenue")
    sample_values = [x["value"] for x in result["first_col_sample"] if isinstance(x, dict)]
    assert "North" in sample_values


def test_get_sheet_summary_sheet_not_found():
    result = get_sheet_summary(str(BUDGET_SIMPLE), "NonExistent")
    assert result["success"] is False
    assert "not found" in result["error"]
    assert "hint" in result


def test_get_sheet_summary_has_progress():
    result = get_sheet_summary(str(BUDGET_SIMPLE), "Q3 Revenue")
    assert "progress" in result


# ---------------------------------------------------------------------------
# read_cell
# ---------------------------------------------------------------------------


def test_read_cell_value_and_type():
    result = read_cell(str(BUDGET_SIMPLE), "Q3 Revenue", "A1")
    assert result["success"] is True
    assert result["value"] == "Region"
    assert result["type"] == "string"
    assert result["cell"] == "A1"


def test_read_cell_number_type():
    result = read_cell(str(BUDGET_SIMPLE), "Q3 Revenue", "B2")
    assert result["success"] is True
    assert result["value"] == 120000
    assert result["type"] == "number"


def test_read_cell_formula_is_none_for_plain_value():
    result = read_cell(str(BUDGET_SIMPLE), "Q3 Revenue", "A1")
    assert result["success"] is True
    assert result["formula"] is None


def test_read_cell_invalid_address():
    result = read_cell(str(BUDGET_SIMPLE), "Q3 Revenue", "not_a_cell")
    assert result["success"] is False
    assert "Invalid cell address" in result["error"]
    assert "hint" in result


def test_read_cell_sheet_not_found():
    result = read_cell(str(BUDGET_SIMPLE), "Phantom", "A1")
    assert result["success"] is False
    assert "not found" in result["error"]


def test_read_cell_has_progress():
    result = read_cell(str(BUDGET_SIMPLE), "Q3 Revenue", "A1")
    assert "progress" in result


# ---------------------------------------------------------------------------
# read_cell_range
# ---------------------------------------------------------------------------


def test_read_cell_range_returns_2d_array():
    result = read_cell_range(str(BUDGET_SIMPLE), "Q3 Revenue", "A1:B2")
    assert result["success"] is True
    assert result["cell_count"] == 4
    assert len(result["data"]) == 2
    assert len(result["data"][0]) == 2


def test_read_cell_range_values_correct():
    result = read_cell_range(str(BUDGET_SIMPLE), "Q3 Revenue", "A1:A1")
    assert result["success"] is True
    assert result["data"][0][0]["value"] == "Region"
    assert result["data"][0][0]["cell"] == "A1"


def test_read_cell_range_too_large_returns_error():
    # 21 cols × 11 rows = 231 cells — exceeds 200
    result = read_cell_range(str(BUDGET_SIMPLE), "Q3 Revenue", "A1:U11")
    assert result["success"] is False
    assert "max" in result["error"].lower() or "200" in result["error"]
    assert "hint" in result


def test_read_cell_range_invalid_notation():
    result = read_cell_range(str(BUDGET_SIMPLE), "Q3 Revenue", "A1-B2")
    assert result["success"] is False
    assert "Invalid range" in result["error"]


def test_read_cell_range_sheet_not_found():
    result = read_cell_range(str(BUDGET_SIMPLE), "Ghost", "A1:B2")
    assert result["success"] is False
    assert "not found" in result["error"]


def test_read_cell_range_has_progress():
    result = read_cell_range(str(BUDGET_SIMPLE), "Q3 Revenue", "A1:C3")
    assert "progress" in result


# ---------------------------------------------------------------------------
# search_cells
# ---------------------------------------------------------------------------


def test_search_cells_finds_match():
    result = search_cells(str(BUDGET_SIMPLE), "Q3 Revenue", "North")
    assert result["success"] is True
    assert len(result["matches"]) >= 1
    assert any(m["value"] == "North" for m in result["matches"])


def test_search_cells_case_insensitive():
    result = search_cells(str(BUDGET_SIMPLE), "Q3 Revenue", "north")
    assert result["success"] is True
    assert len(result["matches"]) >= 1


def test_search_cells_no_match():
    result = search_cells(str(BUDGET_SIMPLE), "Q3 Revenue", "ZZZ_NOPE_ZZZ")
    assert result["success"] is True
    assert result["matches"] == []
    assert "hint" in result


def test_search_cells_empty_query_error():
    result = search_cells(str(BUDGET_SIMPLE), "Q3 Revenue", "")
    assert result["success"] is False
    assert "empty" in result["error"].lower()


def test_search_cells_sheet_not_found():
    result = search_cells(str(BUDGET_SIMPLE), "BadSheet", "North")
    assert result["success"] is False
    assert "not found" in result["error"]


def test_search_cells_has_progress():
    result = search_cells(str(BUDGET_SIMPLE), "Q3 Revenue", "North")
    assert "progress" in result


# ---------------------------------------------------------------------------
# set_cell
# ---------------------------------------------------------------------------


def test_set_cell_writes_value(workbook: Path):
    result = set_cell(str(workbook), "Q3 Revenue", "B2", 999999)
    assert result["success"] is True
    assert result["cell"] == "B2"
    assert result["value"] == 999999

    # Verify on disk
    wb = openpyxl.load_workbook(str(workbook), data_only=True)
    assert wb["Q3 Revenue"]["B2"].value == 999999
    wb.close()


def test_set_cell_creates_snapshot(workbook: Path):
    result = set_cell(str(workbook), "Q3 Revenue", "A1", "Updated")
    assert result["success"] is True
    assert "backup" in result
    assert result["backup"] is not None
    assert Path(result["backup"]).exists()


def test_set_cell_invalid_address_error(workbook: Path):
    result = set_cell(str(workbook), "Q3 Revenue", "not_a_cell", "value")
    assert result["success"] is False
    assert "Invalid cell address" in result["error"]


def test_set_cell_sheet_not_found(workbook: Path):
    result = set_cell(str(workbook), "Nonexistent", "A1", "value")
    assert result["success"] is False
    assert "not found" in result["error"]
    assert "hint" in result


def test_set_cell_file_not_found():
    result = set_cell("/no/such/file.xlsx", "Sheet1", "A1", "value")
    assert result["success"] is False
    assert "not found" in result["error"].lower()


def test_set_cell_has_progress(workbook: Path):
    result = set_cell(str(workbook), "Q3 Revenue", "A1", "Test")
    assert "progress" in result
    assert isinstance(result["progress"], list)
    assert len(result["progress"]) > 0


# ---------------------------------------------------------------------------
# set_range
# ---------------------------------------------------------------------------


def test_set_range_writes_2d_array(workbook: Path):
    data = [["X", "Y"], ["1", "2"]]
    result = set_range(str(workbook), "Q3 Revenue", "A1", data)
    assert result["success"] is True
    assert result["rows_written"] == 2
    assert result["cells_written"] == 4

    wb = openpyxl.load_workbook(str(workbook), data_only=True)
    ws = wb["Q3 Revenue"]
    assert ws["A1"].value == "X"
    assert ws["B1"].value == "Y"
    assert ws["A2"].value == "1"
    assert ws["B2"].value == "2"
    wb.close()


def test_set_range_creates_snapshot(workbook: Path):
    result = set_range(str(workbook), "Q3 Revenue", "A1", [["test"]])
    assert result["success"] is True
    assert "backup" in result
    assert Path(result["backup"]).exists()


def test_set_range_invalid_start_cell(workbook: Path):
    result = set_range(str(workbook), "Q3 Revenue", "bad_cell", [["v"]])
    assert result["success"] is False
    assert "Invalid cell address" in result["error"]


def test_set_range_has_progress(workbook: Path):
    result = set_range(str(workbook), "Q3 Revenue", "A1", [["v"]])
    assert "progress" in result


# ---------------------------------------------------------------------------
# insert_row
# ---------------------------------------------------------------------------


def test_insert_row_shifts_down(workbook: Path):
    # Row 2 currently has "North", 120000, ...
    result = insert_row(str(workbook), "Q3 Revenue", 2)
    assert result["success"] is True

    wb = openpyxl.load_workbook(str(workbook), data_only=True)
    ws = wb["Q3 Revenue"]
    # Row 2 should now be empty; old row 2 is now row 3
    assert ws["A2"].value is None
    assert ws["A3"].value == "North"
    wb.close()


def test_insert_row_creates_snapshot(workbook: Path):
    result = insert_row(str(workbook), "Q3 Revenue", 1)
    assert result["success"] is True
    assert "backup" in result
    assert Path(result["backup"]).exists()


def test_insert_row_sheet_not_found(workbook: Path):
    result = insert_row(str(workbook), "Ghost", 1)
    assert result["success"] is False
    assert "not found" in result["error"]


def test_insert_row_has_progress(workbook: Path):
    result = insert_row(str(workbook), "Q3 Revenue", 1)
    assert "progress" in result


# ---------------------------------------------------------------------------
# delete_row
# ---------------------------------------------------------------------------


def test_delete_row_shifts_up(workbook: Path):
    # Row 1 = headers; row 2 = North
    result = delete_row(str(workbook), "Q3 Revenue", 1)
    assert result["success"] is True

    wb = openpyxl.load_workbook(str(workbook), data_only=True)
    ws = wb["Q3 Revenue"]
    # After deleting row 1 (headers), row 1 should now be "North"
    assert ws["A1"].value == "North"
    wb.close()


def test_delete_row_creates_snapshot(workbook: Path):
    result = delete_row(str(workbook), "Q3 Revenue", 1)
    assert result["success"] is True
    assert "backup" in result
    assert Path(result["backup"]).exists()


def test_delete_row_out_of_range(workbook: Path):
    result = delete_row(str(workbook), "Q3 Revenue", 999)
    assert result["success"] is False
    assert "out of range" in result["error"].lower()


def test_delete_row_sheet_not_found(workbook: Path):
    result = delete_row(str(workbook), "Ghost", 1)
    assert result["success"] is False
    assert "not found" in result["error"]


def test_delete_row_has_progress(workbook: Path):
    result = delete_row(str(workbook), "Q3 Revenue", 1)
    assert "progress" in result


# ---------------------------------------------------------------------------
# add_sheet
# ---------------------------------------------------------------------------


def test_add_sheet_creates_new(workbook: Path):
    result = add_sheet(str(workbook), "NewSheet")
    assert result["success"] is True
    assert result["sheet"] == "NewSheet"

    wb = openpyxl.load_workbook(str(workbook))
    assert "NewSheet" in wb.sheetnames
    wb.close()


def test_add_sheet_auto_name(workbook: Path):
    result = add_sheet(str(workbook))
    assert result["success"] is True
    assert result["sheet"]  # non-empty name assigned

    wb = openpyxl.load_workbook(str(workbook))
    assert result["sheet"] in wb.sheetnames
    wb.close()


def test_add_sheet_creates_snapshot(workbook: Path):
    result = add_sheet(str(workbook), "AuditSheet")
    assert result["success"] is True
    assert "backup" in result
    assert Path(result["backup"]).exists()


def test_add_sheet_duplicate_name_error(workbook: Path):
    result = add_sheet(str(workbook), "Q3 Revenue")
    assert result["success"] is False
    assert "already exists" in result["error"]


def test_add_sheet_has_progress(workbook: Path):
    result = add_sheet(str(workbook), "ProgressSheet")
    assert "progress" in result


# ---------------------------------------------------------------------------
# Global contract: every tool response has a progress field
# ---------------------------------------------------------------------------


def test_all_responses_have_progress_field(workbook: Path):
    """Every engine function must return a dict with a 'progress' key."""
    results = [
        list_sheets(str(BUDGET_SIMPLE)),
        get_sheet_summary(str(BUDGET_SIMPLE), "Q3 Revenue"),
        read_cell(str(BUDGET_SIMPLE), "Q3 Revenue", "A1"),
        read_cell_range(str(BUDGET_SIMPLE), "Q3 Revenue", "A1:B2"),
        search_cells(str(BUDGET_SIMPLE), "Q3 Revenue", "North"),
        set_cell(str(workbook), "Q3 Revenue", "A1", "test"),
        set_range(str(workbook), "Q3 Revenue", "A1", [["v"]]),
        insert_row(str(workbook), "Q3 Revenue", 1),
        delete_row(str(workbook), "Q3 Revenue", 1),
        add_sheet(str(workbook), "TempSheet"),
    ]
    for r in results:
        assert "progress" in r, f"Missing 'progress' in: {r}"
        assert isinstance(r["progress"], list), f"'progress' is not a list in: {r}"
