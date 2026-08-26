"""It said to close Excel. Excel was not open, and never had been.

    set_cell(root_owned_workbook, ...)
      -> error: [Errno 13] Permission denied: '.../ad_data_work.xlsx'
         hint : 'ad_data_work.xlsx' is open in Word, Excel, or PowerPoint.
                Close it and try again.
         backup: .mcp_versions/ad_data_work_....xlsx.bak

Three things wrong in one response, found by a round-15 phase working on a
container mount where the MCP server runs as a different user than the shell.

ONE, the hint named a specific wrong cause. `hint_for_error` answered every
PermissionError with the Windows file-lock message, which is the right answer
when Excel holds a lock and useless when the file is simply owned by someone
else. The hint rules in CLAUDE.md say name a specific fix; naming a specific
WRONG fix is worse than being vague, because the caller acts on it -- closes an
application that is not running, retries, and gets the same error forever.
os.access answers the real question: if this process cannot write the file, no
amount of closing Excel will help.

TWO, a snapshot was written for an edit that did not happen. Every write tool
snapshots before editing, which is the correct order -- the previous state
cannot be captured afterwards. But when the edit raises, the snapshot stays and
the file's history gains an entry for nothing. Verified here: A1 still holds its
original value and .mcp_versions has a new .bak beside it.

`discard_unused_snapshot` is the fix, and it is deliberately conservative: it
removes the snapshot only when it is still byte-for-byte identical to the file
it came from, which proves nothing was written. A partial write leaves them
different, and then that snapshot is the only copy of the original and must
survive.

THREE -- not fixed here, and worth saying plainly -- the 35 except-blocks that
return `"backup": backup` do not call the new helper yet. They are spread
across five per-server `_error()` builders with different signatures plus
inline dicts in the xlsx servers, so wiring them is a consolidation, not an
edit. The helper and its guard are tested here so that work is mechanical when
it happens.
"""

from __future__ import annotations

import os
import subprocess
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from shared.file_utils import hint_for_error  # type: ignore[reportMissingImports]
from shared.version_control import discard_unused_snapshot, snapshot  # type: ignore[reportMissingImports]


@pytest.fixture()
def book(tmp_path: Path) -> Path:
    p = tmp_path / "book.xlsx"
    wb = Workbook()
    wb.active["A1"] = "before"
    wb.save(str(p))
    return p


class TestTheHintMatchesTheActualCause:
    def test_an_unwritable_file_is_not_blamed_on_excel(self, book: Path, monkeypatch) -> None:
        monkeypatch.setattr(os, "access", lambda *a, **k: False)
        hint = hint_for_error(PermissionError(13, "Permission denied"), book)
        assert "not writable by this process" in hint, hint
        assert "Excel" not in hint, hint

    def test_it_points_at_owner_and_mode(self, book: Path, monkeypatch) -> None:
        monkeypatch.setattr(os, "access", lambda *a, **k: False)
        hint = hint_for_error(PermissionError(13, "Permission denied"), book)
        assert "owner" in hint and "mode" in hint, hint

    def test_it_explains_the_container_case(self, book: Path, monkeypatch) -> None:
        """The situation that actually produced this, so the fix is guessable."""
        monkeypatch.setattr(os, "access", lambda *a, **k: False)
        assert "different user than your shell" in hint_for_error(PermissionError(), book)

    def test_a_writable_file_still_gets_the_lock_message(self, book: Path) -> None:
        """A real Excel lock is writable-by-us; that message must survive."""
        hint = hint_for_error(PermissionError(13, "Permission denied"), book)
        assert "open in Word, Excel, or PowerPoint" in hint, hint

    def test_a_missing_path_still_gets_the_lock_message(self) -> None:
        assert "open in Word" in hint_for_error(PermissionError(), None)

    def test_other_errors_are_untouched(self, book: Path) -> None:
        assert "does not exist" in hint_for_error(FileNotFoundError(), book)


class TestASnapshotForAnEditThatDidNotHappen:
    def test_an_untouched_file_drops_its_snapshot(self, book: Path) -> None:
        bak = snapshot(str(book))
        assert Path(bak).exists()
        assert discard_unused_snapshot(bak, str(book)) is True
        assert not Path(bak).exists()

    def test_a_changed_file_keeps_it(self, book: Path) -> None:
        """A partial write makes that snapshot the only copy of the original."""
        bak = snapshot(str(book))
        wb = load_workbook(str(book))
        wb.active["A1"] = "after"
        wb.save(str(book))
        assert discard_unused_snapshot(bak, str(book)) is False
        assert Path(bak).exists()

    def test_a_same_size_but_different_file_keeps_it(self, tmp_path: Path) -> None:
        """Size alone is not evidence; the content hash is what decides."""
        p = tmp_path / "f.bin"
        p.write_bytes(b"AAAA")
        bak = snapshot(str(p))
        p.write_bytes(b"BBBB")
        assert discard_unused_snapshot(bak, str(p)) is False
        assert Path(bak).exists()

    def test_a_vanished_backup_is_not_an_error(self, book: Path) -> None:
        bak = snapshot(str(book))
        Path(bak).unlink()
        assert discard_unused_snapshot(bak, str(book)) is False

    def test_a_vanished_source_keeps_the_backup(self, book: Path) -> None:
        """If the file is gone the snapshot is precious, not redundant."""
        bak = snapshot(str(book))
        book.unlink()
        assert discard_unused_snapshot(bak, str(book)) is False
        assert Path(bak).exists()


def immutable_supported(p: Path) -> bool:
    return subprocess.run(["chattr", "+i", str(p)], capture_output=True).returncode == 0


class TestTheWholeSequenceAgainstARealFailedWrite:
    """Reproduces what the phase actually hit, rather than a mocked stand-in."""

    def test_a_failed_write_leaves_the_file_alone_and_can_shed_its_snapshot(self, book: Path) -> None:
        if not immutable_supported(book):
            pytest.skip("chattr +i unavailable (needs CAP_LINUX_IMMUTABLE on this filesystem)")
        try:
            from xlsx_basic.engine import set_cell  # type: ignore[reportMissingImports]

            r = set_cell(str(book), "Sheet", "A1", "after")
            assert r["success"] is False, r
            assert load_workbook(str(book)).active["A1"].value == "before"
            bak = r.get("backup")
            assert bak and Path(bak).exists(), "the snapshot this test is about was not taken"
            # The file is untouched, so the snapshot is provably redundant.
            assert discard_unused_snapshot(bak, str(book)) is True
        finally:
            subprocess.run(["chattr", "-i", str(book)], capture_output=True)
