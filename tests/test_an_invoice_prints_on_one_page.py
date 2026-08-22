"""The money printed on page 2, away from the labels naming it.

Rendering a generated invoice showed page 1 carrying the header, the two line
descriptions, their quantities and unit prices, and the words "Subtotal" and
"Total (USD)" with nothing beside them. Page 2 carried a lone column: 1939003.26,
564115.51, 2503118.77, 2503118.77.

This is a regression from the fix directly before it. _fit_columns was added so
generated sheets stopped rendering as "impressioclicks" at openpyxl's default
8.43 characters; it sizes each column to its contents. On the invoice the
description column auto-fits to 41 characters, taking the table to 77 (~7.7in)
against roughly 6.9in of printable A4 width -- so the last column wrapped to its
own page. Readable on screen, severed in print.

Nothing structural could see it: the workbook is valid, every cell holds the
right value, and the totals are correct. It shows up only in a render.

Scaling to one page wide is what a person does in Excel for this, and it is
inert when the content already fits. Raw data exports are deliberately left
alone -- scaling 16 columns of a 16,834-row dump to one page wide would print
nothing legible.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from xlsx_new.engine import (  # type: ignore[reportMissingImports]
    create_from_csv,
    create_from_data,
    create_invoice,
    create_report,
)

LONG_ITEMS = [
    {"description": "Google Ads campaign spends - Aug 2026", "quantity": 1, "unit_price": 1939003.26},
    {"description": "Facebook Ads campaign spends - Aug 2026", "quantity": 1, "unit_price": 564115.51},
]


@pytest.fixture()
def invoice(tmp_path: Path):
    out = tmp_path / "inv.xlsx"
    r = create_invoice(
        str(out),
        client_name="Acme Media Group",
        company_name="Coverage Sweep Analytics",
        invoice_number="INV-2026-0014",
        currency="USD",
        tax_rate=0,
        items=LONG_ITEMS,
        open_after=False,
    )
    assert r["success"] is True, r.get("error")
    return openpyxl.load_workbook(str(out)).active


class TestTheInvoiceIsSetToPrintOnePageWide:
    def test_fit_to_page_is_on(self, invoice):
        assert invoice.sheet_properties.pageSetUpPr.fitToPage is True

    def test_it_is_one_page_wide(self, invoice):
        assert invoice.page_setup.fitToWidth == 1

    def test_it_may_run_as_many_pages_tall_as_it_needs(self, invoice):
        """0 means unbounded height; a long item list must not be squashed."""
        assert invoice.page_setup.fitToHeight == 0

    def test_the_content_really_is_wider_than_a_page(self, invoice):
        """If this stops being true the fixture no longer reproduces the bug."""
        total = sum(d.width for d in invoice.column_dimensions.values() if d.width)
        assert total > 69, total  # ~6.9in of printable A4 at ~10 chars/inch


class TestTheInvoiceStillSaysWhatItSaid:
    def test_every_amount_column_is_present(self, invoice):
        headers = [invoice.cell(row=r, column=c).value for r in range(1, 12) for c in range(1, 5)]
        assert "Description" in headers
        assert "Quantity" in headers
        assert "Total" in headers

    def test_the_unit_prices_are_intact(self, invoice):
        prices = [invoice.cell(row=r, column=3).value for r in range(1, invoice.max_row + 1)]
        assert 1939003.26 in prices and 564115.51 in prices

    def test_the_totals_column_still_sums_the_lines(self, invoice):
        """Column D holds formulas, which is why the rendered PDF showed the
        computed 2503118.77 while openpyxl reads the expression."""
        formulas = [str(invoice.cell(row=r, column=4).value or "") for r in range(1, invoice.max_row + 1)]
        assert any(f.startswith("=SUM(D") for f in formulas), formulas
        assert any(f == "=B7*C7" for f in formulas), formulas

    def test_the_descriptions_are_not_truncated(self, invoice):
        text = " ".join(str(invoice.cell(row=r, column=1).value or "") for r in range(1, invoice.max_row + 1))
        assert "Google Ads campaign spends - Aug 2026" in text


class TestAReportGetsTheSameTreatment:
    def test_every_sheet_is_set_to_fit(self, tmp_path: Path):
        out = tmp_path / "rep.xlsx"
        r = create_report(
            str(out),
            title="Coverage",
            sheets=[
                {"name": "Platform", "headers": ["platform", "spends"], "rows": [["Google Ads", 1]]},
                {"name": "Device", "headers": ["device", "clicks"], "rows": [["Mobile", 2]]},
            ],
            open_after=False,
        )
        assert r["success"] is True, r.get("error")
        wb = openpyxl.load_workbook(str(out))
        for ws in wb.worksheets:
            pp = ws.sheet_properties.pageSetUpPr
            assert pp is not None and pp.fitToPage is True, ws.title
            assert ws.page_setup.fitToWidth == 1, ws.title


class TestARawExportIsLeftAlone:
    """Sixteen columns of a 16,834-row dump scaled to one page wide prints
    nothing anyone can read. These are for screen and filtering."""

    def test_create_from_data_does_not_set_it(self, tmp_path: Path):
        out = tmp_path / "d.xlsx"
        create_from_data(str(out), "Data", ["a", "b"], [[1, 2]], open_after=False)
        ws = openpyxl.load_workbook(str(out)).active
        assert ws is not None
        pp = ws.sheet_properties.pageSetUpPr
        assert pp is None or pp.fitToPage is not True

    def test_create_from_csv_does_not_set_it(self, tmp_path: Path):
        csv_path = tmp_path / "in.csv"
        csv_path.write_text("a,b\n1,2\n", encoding="utf-8")
        out = tmp_path / "c.xlsx"
        r = create_from_csv(str(csv_path), str(out), open_after=False)
        assert r["success"] is True, r.get("error")
        ws = openpyxl.load_workbook(str(out)).active
        assert ws is not None
        pp = ws.sheet_properties.pageSetUpPr
        assert pp is None or pp.fitToPage is not True

    def test_the_columns_are_still_widened_though(self, tmp_path: Path):
        """The readability fix this regressed from must stay in place."""
        out = tmp_path / "d.xlsx"
        create_from_data(str(out), "Data", ["impressions", "clicks"], [[1, 2]], open_after=False)
        ws = openpyxl.load_workbook(str(out)).active
        assert ws is not None
        assert ws.column_dimensions["A"].width > 8.43
