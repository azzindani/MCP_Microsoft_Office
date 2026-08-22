"""A gap in a sheet made two read tools fail with an openpyxl internal.

A coverage sweep built a workbook from the ad dataset (columns A-P), then wrote
a marker into X1. That leaves Q1:W1 blank. get_sheet_summary on that sheet came
back:

    success: false
    error:   'EmptyCell' object has no attribute 'coordinate'

Every read in this module streams the sheet (read_only=True) because a full
load materialises the whole workbook for a 200-cell request. A streaming
worksheet yields EmptyCell for a blank cell: it has .value (None) but no
.coordinate. get_sheet_summary read the address off every cell in row 1,
including the blanks.

read_cell_range had the same line and only escaped the sweep by luck -- the
range it was asked for happened to be full. Any range containing one blank cell
failed, which is most ranges in a real spreadsheet.

xlsx_formulas' convert_to_values already derives coordinates from the range
bounds, with a comment naming this exact exception. The fix here is the same
one, applied through a helper so no caller in this module can reintroduce it:
search_cells was safe only because an unrelated `value is not None` check
happened to filter the blanks out first.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from xlsx_basic.engine import (  # type: ignore[reportMissingImports]
    get_sheet_summary,
    read_cell_range,
    search_cells,
)

LEAK = "EmptyCell"


@pytest.fixture()
def gapped(tmp_path: Path) -> str:
    """Headers in A-C and a marker in F, so D1:E1 are blank -- the sweep's shape."""
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"], ws["B1"], ws["C1"] = "Date", "campaign_platform", "spends"
    ws["F1"] = "sweep_marker"
    ws["A2"], ws["B2"], ws["C2"] = "2019-10-16", "Google Ads", 1939000
    ws["A3"], ws["B3"], ws["C3"] = "2019-10-17", "Facebook Ads", 564100
    wb.save(str(path))
    return str(path)


class TestGetSheetSummary:
    def test_it_succeeds(self, gapped: str):
        r = get_sheet_summary(gapped, "Data")
        assert r["success"] is True, r.get("error")

    def test_no_openpyxl_internal_reaches_the_caller(self, gapped: str):
        assert LEAK not in str(get_sheet_summary(gapped, "Data"))

    def test_the_header_row_still_carries_every_column(self, gapped: str):
        header = get_sheet_summary(gapped, "Data")["header_row"]
        assert [c["cell"] for c in header] == ["A1", "B1", "C1", "D1", "E1", "F1"], header

    def test_the_blank_headers_are_reported_as_blank(self, gapped: str):
        header = {c["cell"]: c["value"] for c in get_sheet_summary(gapped, "Data")["header_row"]}
        assert header["A1"] == "Date"
        assert header["D1"] is None and header["E1"] is None
        assert header["F1"] == "sweep_marker"

    def test_the_first_column_sample_is_addressed_correctly(self, gapped: str):
        sample = get_sheet_summary(gapped, "Data")["first_col_sample"]
        cells = [s["cell"] for s in sample if isinstance(s, dict)]
        assert cells == ["A2", "A3"], sample


class TestReadCellRange:
    def test_a_range_containing_a_blank_cell_is_read(self, gapped: str):
        r = read_cell_range(gapped, "Data", "A1:F1")
        assert r["success"] is True, r.get("error")

    def test_no_openpyxl_internal_reaches_the_caller(self, gapped: str):
        assert LEAK not in str(read_cell_range(gapped, "Data", "A1:F1"))

    def test_every_cell_keeps_its_own_address(self, gapped: str):
        row = read_cell_range(gapped, "Data", "A1:F1")["data"][0]
        assert [c["cell"] for c in row] == ["A1", "B1", "C1", "D1", "E1", "F1"], row

    def test_addresses_are_right_for_a_range_that_does_not_start_at_a1(self, gapped: str):
        """Deriving coordinates from the bounds has to respect the offset."""
        r = read_cell_range(gapped, "Data", "B2:C3")
        assert [[c["cell"] for c in row] for row in r["data"]] == [["B2", "C2"], ["B3", "C3"]]

    def test_the_values_still_line_up_with_the_addresses(self, gapped: str):
        r = read_cell_range(gapped, "Data", "A2:C2")
        got = {c["cell"]: c["value"] for c in r["data"][0]}
        assert got == {"A2": "2019-10-16", "B2": "Google Ads", "C2": 1939000}


class TestSearchCells:
    def test_it_still_finds_the_value(self, gapped: str):
        r = search_cells(gapped, "Data", "Google Ads")
        assert r["success"] is True, r.get("error")
        assert [m["cell"] for m in r["matches"]] == ["B2"], r["matches"]

    def test_it_survives_the_gap(self, gapped: str):
        r = search_cells(gapped, "Data", "sweep_marker")
        assert [m["cell"] for m in r["matches"]] == ["F1"], r["matches"]


class TestAFullSheetIsUnaffected:
    @pytest.fixture()
    def full(self, tmp_path: Path) -> str:
        path = tmp_path / "full.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"
        ws.append(["Platform", "Spend"])
        ws.append(["Google Ads", 1939000])
        wb.save(str(path))
        return str(path)

    def test_summary_addresses_are_unchanged(self, full: str):
        header = get_sheet_summary(full, "Data")["header_row"]
        assert [c["cell"] for c in header] == ["A1", "B1"]

    def test_range_addresses_are_unchanged(self, full: str):
        r = read_cell_range(full, "Data", "A1:B2")
        assert [[c["cell"] for c in row] for row in r["data"]] == [["A1", "B1"], ["A2", "B2"]]
