"""Tests for xlsx_formulas engine functions."""

import shutil
from pathlib import Path

import openpyxl
import pytest
from xlsx_formulas.engine import (
    freeze_panes,
    set_autofilter,
    set_conditional_format,
    set_data_validation,
    set_formula,
    set_named_range,
)

FIXTURES = Path(__file__).parent / "fixtures"
BUDGET_FORMULAS = FIXTURES / "budget_formulas.xlsx"


@pytest.fixture()
def workbook(tmp_path: Path) -> Path:
    """Return a writable copy of budget_formulas.xlsx."""
    dest = tmp_path / "budget_formulas.xlsx"
    shutil.copy(BUDGET_FORMULAS, dest)
    return dest


# ---------------------------------------------------------------------------
# set_formula
# ---------------------------------------------------------------------------


def test_set_formula_writes_formula_string(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_formula(str(workbook), sheet_name, "Z1", "=SUM(A1:A5)")
    assert result["success"] is True
    assert result["formula"] == "=SUM(A1:A5)"
    assert result["cell"] == "Z1"

    # Verify it was saved
    wb2 = openpyxl.load_workbook(str(workbook))
    ws = wb2[sheet_name]
    assert ws["Z1"].value == "=SUM(A1:A5)"
    wb2.close()


def test_set_formula_rejects_missing_equals(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_formula(str(workbook), sheet_name, "Z1", "SUM(A1:A5)")
    assert result["success"] is False
    assert "=" in result["error"]
    assert "hint" in result


def test_set_formula_creates_snapshot(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_formula(str(workbook), sheet_name, "Z1", "=1+1")
    assert result["success"] is True
    assert "backup" in result
    backup_path = Path(result["backup"])
    assert backup_path.exists()


def test_set_formula_invalid_cell_address(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_formula(str(workbook), sheet_name, "not_a_cell", "=1+1")
    assert result["success"] is False
    assert "hint" in result


def test_set_formula_file_not_found() -> None:
    result = set_formula("/nonexistent/file.xlsx", "Sheet1", "A1", "=1")
    assert result["success"] is False
    assert "not found" in result["error"].lower()
    assert "hint" in result


# ---------------------------------------------------------------------------
# set_named_range
# ---------------------------------------------------------------------------


def test_set_named_range_creates_range(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_named_range(str(workbook), sheet_name, "MyTotal", "A1:A10")
    assert result["success"] is True
    assert result["range_name"] == "MyTotal"

    wb2 = openpyxl.load_workbook(str(workbook))
    assert "MyTotal" in wb2.defined_names
    wb2.close()


# ---------------------------------------------------------------------------
# set_conditional_format
# ---------------------------------------------------------------------------


def test_set_conditional_format_greater_than(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_conditional_format(
        str(workbook), sheet_name, "A1:A10", "greater_than", 100.0, "green"
    )
    assert result["success"] is True
    assert result["rule"] == "greater_than"
    assert result["color"] == "green"
    assert "backup" in result


def test_set_conditional_format_less_than(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_conditional_format(str(workbook), sheet_name, "A1:A10", "less_than", 50.0, "red")
    assert result["success"] is True


def test_set_conditional_format_between(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_conditional_format(
        str(workbook), sheet_name, "A1:A10", "between", 10.0, "yellow", value2=50.0
    )
    assert result["success"] is True


def test_set_conditional_format_invalid_rule(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_conditional_format(
        str(workbook), sheet_name, "A1:A10", "not_a_rule", 10.0, "green"
    )
    assert result["success"] is False
    assert "hint" in result


def test_set_conditional_format_invalid_color(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_conditional_format(
        str(workbook), sheet_name, "A1:A10", "greater_than", 10.0, "purple"
    )
    assert result["success"] is False
    assert "hint" in result


# ---------------------------------------------------------------------------
# set_data_validation
# ---------------------------------------------------------------------------


def test_set_data_validation_list(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_data_validation(
        str(workbook), sheet_name, "A1:A10", "list", '"Option1,Option2,Option3"'
    )
    assert result["success"] is True
    assert result["validation_type"] == "list"
    assert "backup" in result


def test_set_data_validation_decimal(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_data_validation(str(workbook), sheet_name, "B1:B10", "decimal", "0", "100")
    assert result["success"] is True


def test_set_data_validation_invalid_type(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_data_validation(str(workbook), sheet_name, "A1:A5", "not_a_type")
    assert result["success"] is False
    assert "hint" in result


# ---------------------------------------------------------------------------
# freeze_panes
# ---------------------------------------------------------------------------


def test_freeze_panes_sets_freeze(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = freeze_panes(str(workbook), sheet_name, "B2")
    assert result["success"] is True
    assert result["frozen"] is True
    assert result["cell_address"] == "B2"
    assert "backup" in result

    wb2 = openpyxl.load_workbook(str(workbook))
    ws = wb2[sheet_name]
    assert ws.freeze_panes == "B2"
    wb2.close()


def test_freeze_panes_unfreeze(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    # First freeze
    freeze_panes(str(workbook), sheet_name, "B2")
    # Then unfreeze
    result = freeze_panes(str(workbook), sheet_name, "")
    assert result["success"] is True
    assert result["frozen"] is False


def test_freeze_panes_invalid_address(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = freeze_panes(str(workbook), sheet_name, "notcell")
    assert result["success"] is False
    assert "hint" in result


# ---------------------------------------------------------------------------
# set_autofilter
# ---------------------------------------------------------------------------


def test_set_autofilter_adds_filter(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    result = set_autofilter(str(workbook), sheet_name, "A1:F1")
    assert result["success"] is True
    assert result["range"] == "A1:F1"
    assert "backup" in result

    wb2 = openpyxl.load_workbook(str(workbook))
    ws = wb2[sheet_name]
    assert ws.auto_filter.ref == "A1:F1"
    wb2.close()


def test_set_autofilter_file_not_found() -> None:
    result = set_autofilter("/nonexistent/file.xlsx", "Sheet1", "A1:D1")
    assert result["success"] is False


# ---------------------------------------------------------------------------
# progress key present on all responses
# ---------------------------------------------------------------------------


def test_all_responses_have_progress(workbook: Path) -> None:
    wb = openpyxl.load_workbook(str(workbook))
    sheet_name = wb.sheetnames[0]
    wb.close()

    results = [
        set_formula(str(workbook), sheet_name, "Z99", "=1"),
        set_formula(str(workbook), sheet_name, "Z99", "no_equals"),
        set_formula("/bad/path.xlsx", "Sheet1", "A1", "=1"),
        set_conditional_format(str(workbook), sheet_name, "A1:A5", "greater_than", 1.0, "green"),
        set_data_validation(str(workbook), sheet_name, "A1:A5", "list", '"A,B"'),
        freeze_panes(str(workbook), sheet_name, "B2"),
        set_autofilter(str(workbook), sheet_name, "A1:D1"),
    ]
    for r in results:
        assert "progress" in r, f"Missing 'progress' key in response: {r}"
        assert isinstance(r["progress"], list)
