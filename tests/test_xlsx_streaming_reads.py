"""Reading one cell must not load the whole sheet into memory.

`read_cell` and `read_cell_range` each opened the workbook twice without
`read_only=True`. openpyxl then builds a Python object for every cell in the
file, so asking for a single cell out of a 16,834 x 16 sheet allocated ~510 MB
against the container's 512 MB limit and killed the server process. All twelve
Office sub-servers share that container, so a single `read_cell` took Word and
PowerPoint down with it. Streaming the identical read peaks at 37 MB and
returns immediately.

Every other read tool in this server already passed `read_only=True`; these two
were the outliers. The write tools legitimately do not -- openpyxl cannot write
a read-only workbook -- so this asserts on the read paths specifically.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest

from xlsx_basic.engine import read_cell, read_cell_range

FIXTURES = Path(__file__).parent / "fixtures"
BUDGET_SIMPLE = FIXTURES / "budget_simple.xlsx"


@pytest.fixture()
def workbook(tmp_path: Path) -> Path:
    dest = tmp_path / "budget_simple.xlsx"
    shutil.copy(BUDGET_SIMPLE, dest)
    return dest


@pytest.fixture()
def load_calls(monkeypatch) -> list[dict]:
    """Record the kwargs of every load_workbook call the tool makes."""
    calls: list[dict] = []
    real = openpyxl.load_workbook

    def spy(*args, **kwargs):
        calls.append(dict(kwargs))
        return real(*args, **kwargs)

    monkeypatch.setattr("xlsx_basic.engine.openpyxl.load_workbook", spy)
    return calls


class TestReadsAreStreamed:
    def test_read_cell_streams_every_load(self, workbook, load_calls):
        sheet = openpyxl.load_workbook(workbook).sheetnames[0]
        load_calls.clear()  # drop this test's own setup read
        result = read_cell(str(workbook), sheet, "A1")

        assert result["success"] is True
        assert load_calls, "read_cell made no load_workbook call"
        for call in load_calls:
            assert call.get("read_only") is True, f"non-streaming load: {call}"

    def test_read_cell_range_streams_every_load(self, workbook, load_calls):
        sheet = openpyxl.load_workbook(workbook).sheetnames[0]
        load_calls.clear()  # drop this test's own setup read
        result = read_cell_range(str(workbook), sheet, "A1:B3")

        assert result["success"] is True
        assert load_calls, "read_cell_range made no load_workbook call"
        for call in load_calls:
            assert call.get("read_only") is True, f"non-streaming load: {call}"

    def test_both_value_and_formula_passes_are_still_made(self, workbook, load_calls):
        """The cached value and the formula never come from the same load, so
        streaming must not be achieved by dropping one of them."""
        sheet = openpyxl.load_workbook(workbook).sheetnames[0]
        load_calls.clear()  # drop this test's own setup read
        read_cell(str(workbook), sheet, "A1")

        data_only_flags = sorted(call.get("data_only") for call in load_calls)
        assert data_only_flags == [False, True], f"expected one of each pass, got {data_only_flags}"


class TestReadsStillReturnTheRightThing:
    """Streaming changes how the file is walked; it must not change answers."""

    def test_values_match_a_direct_read(self, workbook):
        wb = openpyxl.load_workbook(workbook, data_only=True)
        sheet = wb.sheetnames[0]
        expected = wb[sheet]["A1"].value
        wb.close()

        assert read_cell(str(workbook), sheet, "A1")["value"] == expected

    def test_range_shape_is_preserved(self, workbook):
        sheet = openpyxl.load_workbook(workbook).sheetnames[0]
        result = read_cell_range(str(workbook), sheet, "A1:B3")

        assert result["success"] is True
        assert len(result["data"]) == 3
        assert all(len(row) == 2 for row in result["data"])

    def test_a_missing_sheet_still_reports_cleanly(self, workbook):
        result = read_cell(str(workbook), "NoSuchSheet", "A1")
        assert result["success"] is False
        assert "list_sheets" in result["hint"]


class TestConvertToValuesDoesNotDoubleTheMemory:
    """convert_to_values needs a writable workbook to save its edits, which
    cannot stream. But it also opened a *second* full copy with data_only=True
    just to read cached values, doubling the footprint. On a 16,834 x 16 sheet
    the pair ran long enough for the transport to drop mid-call -- the same
    failure read_cell had, and the one remaining tool holding two full loads."""

    @pytest.fixture()
    def formula_book(self, tmp_path: Path) -> Path:
        path = tmp_path / "f.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"], ws["A2"] = 2, 3
        ws["B1"] = "=A1+A2"
        wb.save(path)
        return path

    @pytest.fixture()
    def formula_calls(self, monkeypatch) -> list[dict]:
        calls: list[dict] = []
        real = openpyxl.load_workbook

        def spy(*args, **kwargs):
            calls.append(dict(kwargs))
            return real(*args, **kwargs)

        monkeypatch.setattr("xlsx_formulas.engine.openpyxl.load_workbook", spy)
        return calls

    def test_the_values_copy_streams(self, formula_book, formula_calls):
        from xlsx_formulas.engine import convert_to_values

        sheet = openpyxl.load_workbook(formula_book).sheetnames[0]
        formula_calls.clear()
        result = convert_to_values(str(formula_book), sheet, "A1:B2")

        assert result["success"] is True
        values_loads = [c for c in formula_calls if c.get("data_only") is True]
        assert values_loads, "no data_only load was made"
        for call in values_loads:
            assert call.get("read_only") is True, f"values copy still loads in full: {call}"

    def test_only_one_writable_load_is_held(self, formula_book, formula_calls):
        from xlsx_formulas.engine import convert_to_values

        sheet = openpyxl.load_workbook(formula_book).sheetnames[0]
        formula_calls.clear()
        convert_to_values(str(formula_book), sheet, "A1:B2")

        full_loads = [c for c in formula_calls if not c.get("read_only")]
        assert len(full_loads) == 1, f"expected one writable load, got {len(full_loads)}"

    def test_conversion_still_works(self, formula_book):
        """Streaming changes how the file is walked; the result must not change.
        A formula with no cached value must still be skipped rather than wiped."""
        from xlsx_formulas.engine import convert_to_values

        sheet = openpyxl.load_workbook(formula_book).sheetnames[0]
        result = convert_to_values(str(formula_book), sheet, "A1:B2")

        assert result["success"] is True
        check = openpyxl.load_workbook(formula_book)[sheet]
        # openpyxl never evaluates formulas, so B1 has no cached value and must
        # be left intact rather than blanked.
        assert check["B1"].value == "=A1+A2"
