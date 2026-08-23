"""add_pivot_table required a `cols` column, so the commonest pivot was unbuildable.

"Total spend per platform" is one row axis and one measure. This tool had no
way to say it -- `cols` carried no default, so pydantic refused the call before
the server could explain:

    add_pivot_table(rows="campaign_platform", values="spends")
      1 validation error for add_pivot_tableArguments
      cols
        Field required [type=missing, ...]

The obvious escape made it worse. Passing cols="" reached the header check,
which tested all three names the same way and quoted the caller's own attempt
to opt out back at them:

    {"success": false,
     "error": "Column '' not found in source range headers",
     "hint": "Available headers: Date, product, phase, campaign_platform, ..."}

That is the third repo to grow the same shape -- an error naming a value the
caller never chose. Here the caller did type it, but only because the schema
left no way to mean "no second axis".

Watched live, a sweep model called add_pivot_table, got the validation error,
and immediately re-sent the identical call with cols=device bolted on -- a
column it had no reason to group by, picked to satisfy the signature. The
phase then burned two more attempts.

Also fixed here, both found while confirming the above against the live
endpoint:

* The snapshot was taken before any validation, so a typo'd column name copied
  the whole workbook to .mcp_versions and then refused. Two probe calls left
  two backups of a 16,834-row file having changed nothing. The snapshot now
  happens once the arguments are known to be good, and `backup` is honestly
  None when nothing was written.
* Header matching was exact, against a sheet whose headers are "Date" beside
  "spends". Which convention any one column follows is unknowable to a caller,
  so a case/space-insensitive match resolves it and the progress log says which
  header it landed on.
"""

from __future__ import annotations

import inspect
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string

from xlsx_charts import engine
from xlsx_charts.engine import add_pivot_table

ROWS = [
    ("platform", "device", "spends", "clicks"),
    ("Google", "Mobile", 100.0, 5),
    ("Google", "Desktop", 50.0, 3),
    ("Meta", "Mobile", 25.0, 1),
    ("Meta", "Mobile", 25.0, 1),
]


@pytest.fixture()
def book(tmp_path: Path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Ad Data"
    for row in ROWS:
        ws.append(row)
    dst = tmp_path / "pivot.xlsx"
    wb.save(str(dst))
    wb.close()
    return str(dst)


def cells(path: str, top_left: str, width: int, height: int) -> list[list[object]]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Ad Data"]
    col = column_index_from_string("".join(c for c in top_left if c.isalpha()))
    row = int("".join(c for c in top_left if c.isdigit()))
    out: list[list[object]] = [
        [ws.cell(row=row + r, column=col + c).value for c in range(width)] for r in range(height)
    ]
    wb.close()
    return out


class TestOneAxisIsEnough:
    def test_cols_is_optional_in_the_tool_schema(self):
        """The whole defect: pydantic rejected the call, not the server."""
        from xlsx_charts import server

        sig = inspect.signature(server.add_pivot_table)
        assert sig.parameters["cols"].default == "", sig

    def test_rows_and_values_are_still_required(self):
        from xlsx_charts import server

        sig = inspect.signature(server.add_pivot_table)
        for name in ("rows", "values"):
            assert sig.parameters[name].default is inspect.Parameter.empty, name

    def test_a_one_dimensional_pivot_succeeds(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="spends")
        assert r["success"] is True, r.get("error")

    def test_it_sums_the_measure_per_group(self, book: str):
        add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="spends")
        grid = cells(book, "F1", 2, 3)
        assert grid[0] == ["platform", "spends"]
        assert grid[1] == ["Google", 150.0]
        assert grid[2] == ["Meta", 50.0]

    def test_the_single_column_is_named_after_what_it_sums(self, book: str):
        """Not blank, and not the empty group key it is aggregated under."""
        add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="clicks")
        assert cells(book, "F1", 2, 1)[0] == ["platform", "clicks"]

    def test_it_reports_one_column_group(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="spends")
        assert r["col_groups"] == 1, r


class TestTwoAxesStillWork:
    def test_the_two_dimensional_pivot_is_unchanged(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", "device", "spends")
        assert r["success"] is True, r.get("error")
        assert r["row_groups"] == 2 and r["col_groups"] == 2

    def test_its_numbers_are_right(self, book: str):
        add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", "device", "spends")
        grid = cells(book, "F1", 3, 3)
        header = grid[0]
        google = dict(zip(header[1:], grid[1][1:]))
        assert grid[1][0] == "Google"
        assert google["Mobile"] == 100.0 and google["Desktop"] == 50.0

    def test_engine_order_is_unchanged_for_positional_callers(self, book: str):
        """server.py passes cols before values; existing tests do too."""
        params = list(inspect.signature(engine.add_pivot_table).parameters)
        assert params[:7] == [
            "file_path",
            "sheet_name",
            "source_range",
            "dest_cell",
            "rows",
            "cols",
            "values",
        ]


class TestAMissingNameIsNamed:
    def test_an_empty_values_says_which_argument_is_missing(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="")
        assert r["success"] is False
        assert "values" in r["error"]

    def test_it_does_not_quote_the_empty_string_back(self, book: str):
        """The old message was: Column '' not found in source range headers."""
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="")
        assert "''" not in r["error"], r["error"]

    def test_an_empty_rows_is_refused_too(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "", values="spends")
        assert r["success"] is False and "rows" in r["error"]

    def test_the_hint_says_cols_is_the_optional_one(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="")
        assert "cols" in r["hint"] and "optional" in r["hint"].lower()

    def test_a_real_typo_still_names_the_argument_and_the_headers(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platfrom", values="spends")
        assert r["success"] is False
        assert "platfrom" in r["error"]
        assert "rows=" in r["hint"] and "platform" in r["hint"]

    def test_a_bad_cols_names_cols(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", "devise", "spends")
        assert r["success"] is False and "cols=" in r["hint"]


class TestHeaderCaseIsNotAGuess:
    @pytest.mark.parametrize("given", ["Platform", "PLATFORM", " platform "])
    def test_it_resolves_to_the_header_as_written(self, book: str, given: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", given, values="spends")
        assert r["success"] is True, r.get("error")
        assert cells(book, "F1", 1, 1)[0][0] == "platform"

    def test_it_says_so_in_the_progress_log(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "PLATFORM", values="spends")
        msgs = " ".join(str(p.get("msg", "")) for p in r["progress"])
        assert "PLATFORM" in msgs and "platform" in msgs

    def test_an_exact_match_is_not_announced(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="spends")
        assert "Matched" not in " ".join(str(p.get("msg", "")) for p in r["progress"])


class TestNothingIsSnapshottedForACallThatCannotRun:
    @pytest.mark.parametrize(
        "kwargs",
        [
            {"rows": "platform", "values": ""},
            {"rows": "", "values": "spends"},
            {"rows": "nope", "values": "spends"},
            {"rows": "platform", "cols": "nope", "values": "spends"},
        ],
    )
    def test_a_rejected_call_leaves_no_backup(self, book: str, kwargs: dict):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", **kwargs)
        assert r["success"] is False
        assert r["backup"] is None, r["backup"]

    def test_and_writes_no_version_file(self, book: str, tmp_path: Path):
        add_pivot_table(book, "Ad Data", "A1:D5", "F1", "nope", values="spends")
        versions = tmp_path / ".mcp_versions"
        assert not versions.exists() or not list(versions.glob("*")), list(versions.glob("*"))

    def test_a_call_that_runs_still_takes_one(self, book: str):
        r = add_pivot_table(book, "Ad Data", "A1:D5", "F1", "platform", values="spends")
        assert r["backup"], r
        assert Path(r["backup"]).exists()
