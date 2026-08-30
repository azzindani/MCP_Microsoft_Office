"""set_data_validation added a second rule instead of replacing the first.

Found by round 22's sweep, which calls one tool twice with identical arguments
and diffs the two responses. The responses were byte-identical apart from the
backup timestamp -- correctly so -- while the FILE gained a second
DataValidation on the same sqref. So the only evidence of the duplication was
in the workbook, and nothing a caller could read said it had happened.

Not a contract violation: `EDITS` declares `idempotentHint=False`, and the note
beside that declaration names the appending ops -- "add_row, append_text,
add_slide and their kin". `set_data_validation` is not their kin. Within its own
module two of the four `set_*` tools replaced and two appended:

    set_named_range        wb.defined_names[name] = defn      replaces
    freeze_panes           ws.freeze_panes = addr             replaces
    set_conditional_format ws.conditional_formatting.add(...) appends
    set_data_validation    ws.add_data_validation(dv)         appends

A caller has no way to tell which is which from the names, and `set` reads as
replace in three of the four. Excel also does not enjoy two rules claiming the
same range.

Only the exact-range case is replaced. A rule on an OVERLAPPING but different
range is left alone: deciding that C2:C100 supersedes C50:C200 is a judgement
this tool has no business making silently.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest

FIXTURES = Path(__file__).parent / "fixtures"


@pytest.fixture()
def workbook(tmp_path: Path) -> Path:
    dest = tmp_path / "budget_formulas.xlsx"
    shutil.copy(FIXTURES / "budget_formulas.xlsx", dest)
    return dest


def _sheet_name(path: Path) -> str:
    return openpyxl.load_workbook(str(path)).sheetnames[0]


def _validations(path: Path, sheet: str) -> list:
    return list(openpyxl.load_workbook(str(path))[sheet].data_validations.dataValidation)


def test_the_same_call_twice_leaves_one_rule(workbook: Path) -> None:
    """The defect, in one assertion."""
    from xlsx_formulas.engine import set_data_validation

    sheet = _sheet_name(workbook)
    first = set_data_validation(str(workbook), sheet, "A1:A10", "list", '"One,Two,Three"')
    assert first["success"], first
    second = set_data_validation(str(workbook), sheet, "A1:A10", "list", '"One,Two,Three"')
    assert second["success"], second

    rules = [dv for dv in _validations(workbook, sheet) if str(dv.sqref) == "A1:A10"]
    assert len(rules) == 1, f"{len(rules)} rules on one range after two identical calls"


def test_a_second_call_replaces_rather_than_stacks(workbook: Path) -> None:
    """The surviving rule must be the NEW one, not the old one kept."""
    from xlsx_formulas.engine import set_data_validation

    sheet = _sheet_name(workbook)
    set_data_validation(str(workbook), sheet, "A1:A10", "list", '"One,Two"')
    set_data_validation(str(workbook), sheet, "A1:A10", "list", '"Three,Four"')

    rules = [dv for dv in _validations(workbook, sheet) if str(dv.sqref) == "A1:A10"]
    assert len(rules) == 1
    assert "Three" in str(rules[0].formula1)


def test_a_replacement_is_reported(workbook: Path) -> None:
    """Replacing is a change to something that existed, so it is said out loud."""
    from xlsx_formulas.engine import set_data_validation

    sheet = _sheet_name(workbook)
    set_data_validation(str(workbook), sheet, "A1:A10", "list", '"One,Two"')
    second = set_data_validation(str(workbook), sheet, "A1:A10", "list", '"Three,Four"')

    messages = " ".join(str(entry) for entry in second["progress"])
    assert "Replaced" in messages


def test_a_different_range_is_left_alone(workbook: Path) -> None:
    """Replacement is by exact range, never by overlap.

    Deciding that one range supersedes an overlapping one is a judgement, and
    making it silently would delete a rule the caller never mentioned.
    """
    from xlsx_formulas.engine import set_data_validation

    sheet = _sheet_name(workbook)
    set_data_validation(str(workbook), sheet, "A1:A10", "list", '"One,Two"')
    set_data_validation(str(workbook), sheet, "B1:B10", "list", '"Three,Four"')

    ranges = {str(dv.sqref) for dv in _validations(workbook, sheet)}
    assert "A1:A10" in ranges and "B1:B10" in ranges


def test_the_first_call_reports_no_replacement(workbook: Path) -> None:
    from xlsx_formulas.engine import set_data_validation

    sheet = _sheet_name(workbook)
    first = set_data_validation(str(workbook), sheet, "A1:A10", "list", '"One,Two"')
    assert "Replaced" not in " ".join(str(entry) for entry in first["progress"])
