"""Sheets these tools created were unreadable, and one invoice billed 500% tax.

Both were found by rendering the sweep's workbooks to PDF and looking at them.

openpyxl leaves every column at the default 8.43 characters. create_report's
header row printed as

    platform  spends    impressioclicks

-- "impressions" running into the column beside it -- and every "Google Ads"
cell as "Google Ad". The values are all in the file, so nothing structural
noticed; it only shows up in a render or a print, which is the one thing nobody
had done to these files.

create_invoice was worse. tax_rate is a fraction -- the label multiplies it by
100 and the formula multiplies the subtotal by it directly -- and nothing said
so: the schema carries no descriptions and the docstring has no room. The sweep
passed 5.0 meaning 5% and got back a finished invoice headed "Tax (500.0%)",
subtotal 2,510,119, total 15,060,713, reported as success.
"""

from __future__ import annotations

import csv as csvmod
from pathlib import Path

import openpyxl
import pytest

from xlsx_new.engine import (  # type: ignore[reportMissingImports]
    create_from_csv,
    create_from_data,
    create_invoice,
    create_report,
)

HEADERS = ["campaign_platform", "spends", "impressions", "clicks"]
ROWS = [["Google Ads", 1939003, 776893, 124065], ["Facebook Ads", 564115, 4070612, 77569]]
ITEMS = [{"description": "Google Ads spend", "quantity": 1, "unit_price": 1939003.0}]


def widths(path: str, sheet: str | None = None) -> dict[str, float | None]:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    return {letter: dim.width for letter, dim in ws.column_dimensions.items()}


def header_fits(path: str, sheet: str | None = None) -> bool:
    """Every header must have at least as much room as it has characters."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    for cell in ws[1]:
        if cell.value is None:
            continue
        width = ws.column_dimensions[cell.column_letter].width
        if width is None or width < len(str(cell.value)):
            return False
    return True


class TestCreateFromData:
    def test_the_columns_are_widened(self, tmp_path: Path):
        out = tmp_path / "d.xlsx"
        r = create_from_data(str(out), "Spend", HEADERS, ROWS, open_after=False)
        assert r["success"] is True, r.get("error")
        assert widths(str(out)), "no column widths were set"

    def test_every_header_fits(self, tmp_path: Path):
        out = tmp_path / "d.xlsx"
        create_from_data(str(out), "Spend", HEADERS, ROWS, open_after=False)
        assert header_fits(str(out))

    def test_the_widest_value_fits_too(self, tmp_path: Path):
        out = tmp_path / "d.xlsx"
        create_from_data(str(out), "Spend", HEADERS, ROWS, open_after=False)
        assert widths(str(out))["A"] >= len("campaign_platform")  # type: ignore[operator]

    def test_one_very_long_cell_does_not_run_off_the_page(self, tmp_path: Path):
        out = tmp_path / "d.xlsx"
        create_from_data(str(out), "Spend", ["note"], [["x" * 400]], open_after=False)
        assert widths(str(out))["A"] <= 60  # type: ignore[operator]


class TestCreateReport:
    def test_every_sheet_is_widened(self, tmp_path: Path):
        out = tmp_path / "r.xlsx"
        r = create_report(
            str(out),
            "Ad campaign spend review",
            [
                {"name": "Platforms", "headers": HEADERS, "rows": ROWS},
                {"name": "Totals", "headers": ["metric", "value"], "rows": [["spends", 2503118]]},
            ],
            open_after=False,
        )
        assert r["success"] is True, r.get("error")
        assert header_fits(str(out), "Platforms")
        assert header_fits(str(out), "Totals")


class TestCreateFromCsv:
    def test_the_import_is_widened_too(self, tmp_path: Path):
        source = tmp_path / "ad.csv"
        with source.open("w", newline="", encoding="utf-8") as fh:
            w = csvmod.writer(fh)
            w.writerow(HEADERS)
            w.writerows(ROWS)
        out = tmp_path / "c.xlsx"
        r = create_from_csv(str(source), str(out), open_after=False)
        assert r["success"] is True, r.get("error")
        assert header_fits(str(out))


class TestTheInvoiceTaxRate:
    @pytest.mark.parametrize("rate", [5.0, 21.0, 100.0])
    def test_a_percentage_is_refused(self, rate: float, tmp_path: Path):
        r = create_invoice(
            str(tmp_path / "i.xlsx"),
            "Casava",
            "Marketing",
            "INV-1",
            ITEMS,
            tax_rate=rate,
            open_after=False,
        )
        assert r["success"] is False, f"tax_rate={rate} was accepted"

    def test_the_error_says_it_wants_a_fraction(self, tmp_path: Path):
        r = create_invoice(
            str(tmp_path / "i.xlsx"),
            "Casava",
            "Marketing",
            "INV-1",
            ITEMS,
            tax_rate=5.0,
            open_after=False,
        )
        assert "fraction" in r["error"], r["error"]

    def test_the_hint_does_the_conversion(self, tmp_path: Path):
        r = create_invoice(
            str(tmp_path / "i.xlsx"),
            "Casava",
            "Marketing",
            "INV-1",
            ITEMS,
            tax_rate=5.0,
            open_after=False,
        )
        assert "0.05" in r["hint"], r["hint"]

    def test_no_file_is_written(self, tmp_path: Path):
        out = tmp_path / "i.xlsx"
        create_invoice(str(out), "Casava", "Marketing", "INV-1", ITEMS, tax_rate=5.0, open_after=False)
        assert not out.exists()

    def test_a_negative_rate_is_refused(self, tmp_path: Path):
        r = create_invoice(
            str(tmp_path / "i.xlsx"),
            "Casava",
            "Marketing",
            "INV-1",
            ITEMS,
            tax_rate=-0.1,
            open_after=False,
        )
        assert r["success"] is False

    @pytest.mark.parametrize("rate", [0.0, 0.05, 0.21, 1.0])
    def test_a_real_fraction_still_works(self, rate: float, tmp_path: Path):
        r = create_invoice(
            str(tmp_path / f"i{rate}.xlsx"),
            "Casava",
            "Marketing",
            "INV-1",
            ITEMS,
            tax_rate=rate,
            open_after=False,
        )
        assert r["success"] is True, f"{rate}: {r.get('error')}"

    def test_the_invoice_columns_are_widened(self, tmp_path: Path):
        out = tmp_path / "i.xlsx"
        create_invoice(str(out), "Casava", "Marketing", "INV-1", ITEMS, tax_rate=0.05, open_after=False)
        assert widths(str(out)), "no column widths were set"
