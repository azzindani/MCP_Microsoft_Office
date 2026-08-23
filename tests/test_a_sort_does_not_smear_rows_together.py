"""Four ways an xlsx tool reported success and wrote something wrong.

**sort_sheet smeared rows into each other.** The write-back used

    ws.cell(row=..., column=..., value=val)

and openpyxl's cell() skips the assignment entirely when the value is None:

    cell = self._get_cell(row, column)
    if value is not None:
        cell.value = value

so every blank cell in the sorted data left the *previous* occupant of that
address in place. Sorting three rows by column A turned

    b, 2, ⌀           a, 1, ⌀
    c, ⌀, keep   into  b, 2, keep     <- c's note, now on b's row
    a, 1, ⌀           c, 1, keep     <- b's n, now on c's row

with success:true and the ordering itself perfectly correct. A sweep measured
541 blanks in one column of a 16,834-row sheet come back holding whatever value
had been at that address before. set_range had the same call shape; "" already
cleared correctly there, a JSON null did not.

**set_named_range built an invalid reference.** It prefixed sheet_name
unconditionally, so range_address="Sheet1!$M$2:$M$10" -- the form every Excel
reference a caller has seen -- was stored as "'Sheet1'!Sheet1!$M$2:$M$10".
Excel rejects that definedName; the response said success.

**set_data_validation validated nothing.** openpyxl defaults showErrorMessage
to False, so Excel drew the dropdown and then accepted anything typed over it.

**find_duplicates answered with every row number it had.** The distinct-value
count was capped at 100 and the per-value row list was not capped at all: two
values carrying 15,101 and 1,733 rows put ~16,800 integers into one response
and the transport truncated it. What a caller needs is how many, not which.
"""

from __future__ import annotations

import re
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
for p in (
    str(ROOT),
    str(ROOT / "shared"),
    str(ROOT / "servers" / "xlsx_basic"),
    str(ROOT / "servers" / "xlsx_formulas"),
):
    if p not in sys.path:
        sys.path.insert(0, p)

from xlsx_basic import engine as basic_engine  # noqa: E402
from xlsx_basic import helpers as basic_helpers  # noqa: E402
from xlsx_formulas import engine as formulas_engine  # noqa: E402


@pytest.fixture
def gappy(tmp_path):
    """A sheet whose blanks are real, plus a cell only one row carries."""
    p = tmp_path / "gappy.xlsx"
    wb = Workbook()
    ws = sheet(wb)
    ws.append(["name", "n", "note"])
    ws.append(["b", 2, None])
    ws.append(["c", None, "keep"])
    ws.append(["a", 1, None])
    ws["E2"] = "marker-on-row-b"
    wb.save(p)
    return p


def sheet(wb: Workbook, title: str = "S"):
    """wb.active is Optional to pyright; every workbook here has one."""
    ws = wb.active
    assert ws is not None
    ws.title = title
    return ws


def grid(path: Path, rows: int = 4, cols: int = 5) -> list[list]:
    ws = load_workbook(path)["S"]
    return [[c.value for c in r] for r in ws.iter_rows(min_row=1, max_row=rows, max_col=cols)]


class TestSortKeepsEachRowWhole:
    def test_blanks_stay_blank_and_values_stay_on_their_own_row(self, gappy):
        r = basic_helpers.sort_sheet(str(gappy), "S", "A", ascending=True, has_header=True)
        assert r["success"] is True, r.get("error")
        assert grid(gappy) == [
            ["name", "n", "note", None, None],
            ["a", 1, None, None, None],
            ["b", 2, None, None, "marker-on-row-b"],
            ["c", None, "keep", None, None],
        ]

    def test_descending_too(self, gappy):
        basic_helpers.sort_sheet(str(gappy), "S", "A", ascending=False, has_header=True)
        assert [row[0] for row in grid(gappy)[1:]] == ["c", "b", "a"]

    def test_no_value_is_duplicated_across_rows(self, gappy):
        basic_helpers.sort_sheet(str(gappy), "S", "A", ascending=True, has_header=True)
        markers = [c for row in grid(gappy) for c in row if c == "marker-on-row-b"]
        assert len(markers) == 1, "a value that belonged to one row appears on two"

    def test_the_multiset_of_values_is_unchanged(self, gappy):
        before = sorted(str(c) for row in grid(gappy) for c in row)
        basic_helpers.sort_sheet(str(gappy), "S", "A", ascending=True, has_header=True)
        assert sorted(str(c) for row in grid(gappy) for c in row) == before


class TestSetRangeClearsWhatItIsToldTo:
    @pytest.mark.parametrize("blank", ["", None])
    def test_a_blank_overwrites_rather_than_leaving_the_old_value(self, tmp_path, blank):
        p = tmp_path / "r.xlsx"
        wb = Workbook()
        ws = sheet(wb)
        ws["A1"], ws["B1"] = "old-a", "old-b"
        wb.save(p)
        r = basic_engine.set_range(str(p), "S", "A1", [["new-a", blank]])
        assert r["success"] is True, r.get("error")
        ws2 = load_workbook(p)["S"]
        assert ws2["A1"].value == "new-a"
        assert ws2["B1"].value is None, "the old value survived a blank write"


class TestNamedRangeReferenceIsValid:
    @pytest.mark.parametrize("address", ["Sheet1!$M$2:$M$10", "'Sheet1'!$M$2:$M$10", "$M$2:$M$10"])
    def test_the_sheet_is_qualified_exactly_once(self, tmp_path, address):
        p = tmp_path / "n.xlsx"
        wb = Workbook()
        sheet(wb, "Sheet1")
        wb.save(p)
        r = formulas_engine.set_named_range(str(p), "Sheet1", "Total", address)
        assert r["success"] is True, r.get("error")
        assert r["reference"] == "'Sheet1'!$M$2:$M$10"
        stored = load_workbook(p).defined_names["Total"].value
        assert stored == "'Sheet1'!$M$2:$M$10", stored
        assert stored.count("!") == 1

    def test_an_address_for_another_sheet_is_refused_not_rewritten(self, tmp_path):
        p = tmp_path / "n.xlsx"
        wb = Workbook()
        sheet(wb, "Sheet1")
        wb.create_sheet("Other")
        wb.save(p)
        r = formulas_engine.set_named_range(str(p), "Sheet1", "Total", "Other!$M$2:$M$10")
        assert r["success"] is False
        assert "Other" in r["error"] and "Sheet1" in r["error"]
        assert "$M$2:$M$10" in r["hint"], "the hint should show the bare range to pass"


class TestDataValidationActuallyRejects:
    def test_excel_is_told_to_show_an_error(self, tmp_path):
        p = tmp_path / "v.xlsx"
        wb = Workbook()
        sheet(wb)
        wb.save(p)
        r = formulas_engine.set_data_validation(str(p), "S", "A2:A100", "list", '"a,b,c"')
        assert r["success"] is True, r.get("error")
        xml = zipfile.ZipFile(p).read("xl/worksheets/sheet1.xml").decode()
        node = re.search(r"<dataValidation [^>]*>", xml)
        assert node, "no dataValidation written"
        assert 'showErrorMessage="1"' in node.group(0), node.group(0)


class TestFindDuplicatesAnswersHowManyNotWhich:
    @pytest.fixture
    def many(self, tmp_path):
        p = tmp_path / "d.xlsx"
        wb = Workbook()
        ws = sheet(wb)
        ws.append(["k"])
        for i in range(500):
            ws.append(["repeated" if i % 2 else "other"])
        wb.save(p)
        return p

    def test_the_row_list_is_capped_and_says_so(self, many):
        r = basic_helpers.find_duplicates(str(many), "S", "A", has_header=True)
        assert r["success"] is True, r.get("error")
        entry = next(e for e in r["duplicates"] if e["value"] == "repeated")
        assert entry["count"] == 250, "the true total must still be reported"
        assert len(entry["rows"]) == r["max_rows_per_value"] < 250
        assert entry["rows_truncated"] is True
        assert r["rows_truncated"] is True

    def test_a_short_list_is_not_marked_truncated(self, tmp_path):
        p = tmp_path / "s.xlsx"
        wb = Workbook()
        ws = sheet(wb)
        ws.append(["k"])
        for v in ["a", "b", "a"]:
            ws.append([v])
        wb.save(p)
        r = basic_helpers.find_duplicates(str(p), "S", "A", has_header=True)
        assert r["rows_truncated"] is False
        assert r["duplicates"] == [{"value": "a", "count": 2, "rows": [2, 4]}]


class TestASheetThatCannotBeSortedSaysSo:
    def test_merged_cells_are_refused_rather_than_scrambled(self, tmp_path):
        # A merged region spans rows, so reordering the rows underneath cannot
        # preserve it, and its non-anchor cells are read-only. The None-skip
        # used to hide that: the sort "succeeded" and left the merged block
        # sitting over whichever rows landed beneath it.
        p = tmp_path / "m.xlsx"
        wb = Workbook()
        ws = sheet(wb)
        ws.append(["name", "n"])
        ws.append(["b", 2])
        ws.append(["a", 1])
        ws.merge_cells("D2:D3")
        wb.save(p)
        r = basic_helpers.sort_sheet(str(p), "S", "A", ascending=True, has_header=True)
        assert r["success"] is False
        assert "merged" in r["error"].lower(), r["error"]
        assert "D2:D3" in r["error"]
        assert "Unmerge" in r["hint"]
